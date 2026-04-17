"""
리포트 생성 모듈
- Excel 리포트 (12개 시트)
- PDF 리포트 (워드클라우드 이미지 포함)
규칙: merge_cells 금지, wrap_text 금지, Arial 폰트, 파란 헤더(4472C4)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ─── 스타일 상수 ───
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

# 긴 텍스트 최대 길이 (wrap_text 금지이므로 말줄임)
MAX_TEXT_LEN = 80


def _truncate(text: str, max_len: int = MAX_TEXT_LEN) -> str:
    """텍스트 말줄임"""
    if not text:
        return ''
    text = str(text).replace('\n', ' ').replace('\r', '')
    if len(text) > max_len:
        return text[:max_len - 3] + '...'
    return text


def _apply_header_style(ws, row: int, col_count: int):
    """헤더 행에 스타일 적용"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def _apply_data_style(ws, row: int, col_count: int, bold: bool = False):
    """데이터 행에 스타일 적용"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BOLD_FONT if bold else DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def _setup_sheet(ws, headers: list, col_widths: list = None, title_row: str = None):
    """시트 기본 설정: 제목행(옵션) + 헤더 + 스타일"""
    start_row = 1
    if title_row:
        ws.cell(row=1, column=1, value=title_row).font = Font(name="Arial", bold=True, size=12)
        start_row = 3

    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    _apply_header_style(ws, start_row, len(headers))

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row}"
    return start_row + 1  # 데이터 시작 행


def create_summary_sheet(wb: Workbook, comprehensive: dict, grade_analysis: dict):
    """종합요약 시트"""
    ws = wb.active
    ws.title = '종합요약'

    headers = ['항목', '내용']
    data_row = _setup_sheet(ws, headers, [25, 80], '종합 분석 요약')

    rows = [
        ('종합등급', f"{comprehensive.get('overall_grade', '')} ({comprehensive.get('overall_score', 0)}점/100점)"),
        ('내신 점수', f"{comprehensive.get('score_breakdown', {}).get('내신', 0)}점"),
        ('세특 점수', f"{comprehensive.get('score_breakdown', {}).get('세특', 0)}점"),
        ('창체 점수', f"{comprehensive.get('score_breakdown', {}).get('창체', 0)}점"),
        ('행특 점수', f"{comprehensive.get('score_breakdown', {}).get('행특', 0)}점"),
        ('출결 점수', f"{comprehensive.get('score_breakdown', {}).get('출결', 0)}점"),
        ('', ''),
        ('추천 전형', comprehensive.get('recommended_jeonhyung', {}).get('추천전형', '')),
        ('전형 판단근거', _truncate(comprehensive.get('recommended_jeonhyung', {}).get('판단근거', ''))),
        ('', ''),
        ('성적 추이', f"{grade_analysis.get('trend_pattern', '')} - {grade_analysis.get('trend_detail', '')}"),
        ('성장 패턴', comprehensive.get('growth_story', {}).get('pattern', '')),
        ('성장 스토리', _truncate(comprehensive.get('growth_story', {}).get('서술', ''), 200)),
        ('', ''),
    ]

    # 핵심 강점
    for i, strength in enumerate(comprehensive.get('core_strengths', []), 1):
        rows.append((f'핵심강점 {i}', _truncate(strength, 200)))

    rows.append(('', ''))
    for i, area in enumerate(comprehensive.get('areas_to_improve', []), 1):
        rows.append((f'보완영역 {i}', _truncate(area, 200)))

    for r, (label, value) in enumerate(rows, data_row):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        _apply_data_style(ws, r, 2, bold=bool(label and '등급' in label))


def create_grade_sheet(wb: Workbook, grade_analysis: dict, extracted: dict):
    """내신분석 시트"""
    ws = wb.create_sheet('내신분석')

    # 학기별 성적 테이블
    headers = ['학기', '전교과 평균', '주요교과 평균']
    data_row = _setup_sheet(ws, headers, [12, 15, 15], '내신 성적 분석')

    for semester in ['1-1', '1-2', '2-1', '2-2', '3-1', '3-2']:
        avg = grade_analysis.get('semester_averages', {}).get(semester, {})
        if avg:
            ws.cell(row=data_row, column=1, value=semester)
            ws.cell(row=data_row, column=2, value=avg.get('전교과', ''))
            ws.cell(row=data_row, column=3, value=avg.get('주요교과', ''))
            _apply_data_style(ws, data_row, 3)
            data_row += 1

    # 전체 평균
    overall = grade_analysis.get('overall_average', {})
    data_row += 1
    ws.cell(row=data_row, column=1, value='전체')
    ws.cell(row=data_row, column=2, value=overall.get('전교과', ''))
    ws.cell(row=data_row, column=3, value=overall.get('주요교과', ''))
    _apply_data_style(ws, data_row, 3, bold=True)

    # 환산 & 추이
    data_row += 2
    ws.cell(row=data_row, column=1, value='9등급 환산').font = BOLD_FONT
    ws.cell(row=data_row, column=2, value=grade_analysis.get('grade_converted_9', ''))
    data_row += 1
    ws.cell(row=data_row, column=1, value='추이 패턴').font = BOLD_FONT
    ws.cell(row=data_row, column=2, value=f"{grade_analysis.get('trend_pattern', '')} - {grade_analysis.get('trend_detail', '')}")

    # 교과이수충실도
    cf = grade_analysis.get('course_fulfillment', {})
    data_row += 2
    ws.cell(row=data_row, column=1, value='교과이수충실도').font = BOLD_FONT
    ws.cell(row=data_row, column=2, value=f"{cf.get('score', 0)}점/5점 - {cf.get('detail', '')}")

    # 소인수 과목
    small = grade_analysis.get('small_class_subjects', [])
    if small:
        data_row += 2
        ws.cell(row=data_row, column=1, value='소인수 과목 (수강자 13명 이하)').font = BOLD_FONT
        data_row += 1
        headers2 = ['학기', '과목명', '수강자수']
        for i, h in enumerate(headers2, 1):
            ws.cell(row=data_row, column=i, value=h)
        _apply_header_style(ws, data_row, 3)
        for s in small:
            data_row += 1
            ws.cell(row=data_row, column=1, value=s['학기'])
            ws.cell(row=data_row, column=2, value=s['과목명'])
            ws.cell(row=data_row, column=3, value=s['수강자수'])
            _apply_data_style(ws, data_row, 3)

    # 과목별 상세 (원점수/평균/맥락)
    context = grade_analysis.get('context_analysis', [])
    if context:
        data_row += 2
        ws.cell(row=data_row, column=1, value='과목별 성적 맥락 분석').font = BOLD_FONT
        data_row += 1
        ctx_headers = ['학기', '과목명', '성취도', '원점수', '과목평균', '차이', '실질위치', '분포분석']
        for i, h in enumerate(ctx_headers, 1):
            ws.cell(row=data_row, column=i, value=h)
        _apply_header_style(ws, data_row, len(ctx_headers))
        for i in range(1, 9):
            ws.column_dimensions[get_column_letter(i)].width = max(
                ws.column_dimensions[get_column_letter(i)].width or 10,
                [12, 15, 8, 8, 8, 8, 15, 25][i-1])

        for c in context:
            data_row += 1
            vals = [c['학기'], c['과목명'], c['성취도'], c['원점수'],
                    c['과목평균'], c['원점수_평균차'], c['실질위치'],
                    _truncate(c['분포분석'], 40)]
            for i, v in enumerate(vals, 1):
                ws.cell(row=data_row, column=i, value=v)
            _apply_data_style(ws, data_row, len(vals))


def create_admission_sheet(wb: Workbook, admission_data: dict):
    """입결비교 시트"""
    ws = wb.create_sheet('입결비교')

    # 안내문
    ws.cell(row=1, column=1, value='입결 비교 분석 (입결 70% 기준)').font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=2, column=1, value=admission_data.get('disclaimer', '')).font = Font(name="Arial", size=9, color="FF0000")
    ws.cell(row=3, column=1, value=f"학생 환산 9등급: {admission_data.get('student_grade_9', '')}").font = BOLD_FONT

    # 비교과 종합등급 표시
    bigyogwa = admission_data.get('bigyogwa_grades', {})
    if bigyogwa:
        ws.cell(row=3, column=4,
                value=f"비교과 종합: {bigyogwa.get('비교과종합', 'N/A')} "
                      f"(세특:{bigyogwa.get('세특', 'N/A')} / 창체:{bigyogwa.get('창체', 'N/A')} / 행특:{bigyogwa.get('행특', 'N/A')})").font = BOLD_FONT

    headers = ['대학', '전형구분', '전형명', '모집단위', '학년도',
               '입결70%', '입결50%', '내환산등급', '차이', '경쟁률']
    col_widths = [15, 10, 20, 20, 8, 10, 10, 10, 8, 8]

    row = 5
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))
    row += 1

    matches = admission_data.get('matches', [])
    for m in matches:
        vals = [m.get('대학', ''), m.get('전형구분', ''), m.get('전형명', ''),
                m.get('모집단위명', ''), m.get('학년도', ''),
                m.get('입결70', ''), m.get('입결50', ''),
                m.get('내환산등급', ''), m.get('차이', ''), m.get('경쟁률', '')]
        for i, v in enumerate(vals, 1):
            ws.cell(row=row, column=i, value=v)
        _apply_data_style(ws, row, len(vals))
        row += 1

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A6'


def create_setuek_sheet(wb: Workbook, setuek_analysis: dict):
    """세특분석 시트"""
    ws = wb.create_sheet('세특분석')

    headers = ['학기', '과목명', '교과연계성', '탐구동기', '탐구과정',
               '결과성찰', '전공적합성', '차별성', '학업태도', '가중평균', '등급',
               '강점', '보완점', '좋은평가문장', '문장평가이유']
    col_widths = [8, 12, 8, 8, 8, 8, 8, 8, 8, 8, 6, 35, 35, 40, 40]
    data_row = _setup_sheet(ws, headers, col_widths, '세특 분석 결과')

    overall = setuek_analysis.get('overall_grade', '')
    ws.cell(row=2, column=1, value=f"종합등급: {overall} (가중평균: {setuek_analysis.get('overall_weighted_score', 0)})").font = BOLD_FONT

    for subj in setuek_analysis.get('subject_scores', []):
        scores = subj.get('scores', {})
        good_sentences = subj.get('좋은평가문장', [])
        sentence_text = ' / '.join([s.get('문장', '') for s in good_sentences]) if good_sentences else ''
        reason_text = ' / '.join([s.get('이유', '') for s in good_sentences]) if good_sentences else ''

        vals = [
            subj.get('학기', ''),
            subj.get('과목명', ''),
            scores.get('교과연계성', {}).get('score', ''),
            scores.get('탐구동기', {}).get('score', ''),
            scores.get('탐구과정', {}).get('score', ''),
            scores.get('결과성찰', {}).get('score', ''),
            scores.get('전공적합성', {}).get('score', ''),
            scores.get('차별성', {}).get('score', ''),
            scores.get('학업태도', {}).get('score', ''),
            subj.get('weighted_score', ''),
            subj.get('grade', ''),
            _truncate(subj.get('강점', ''), 60),
            _truncate(subj.get('보완점', ''), 60),
            _truncate(sentence_text, 80),
            _truncate(reason_text, 80),
        ]
        for i, v in enumerate(vals, 1):
            ws.cell(row=data_row, column=i, value=v)
        _apply_data_style(ws, data_row, len(vals))
        data_row += 1


def create_changche_sheet(wb: Workbook, changche_analysis: dict):
    """창체분석 시트"""
    ws = wb.create_sheet('창체분석')

    ws.cell(row=1, column=1, value='창체 분석 결과').font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=2, column=1, value=f"종합등급: {changche_analysis.get('overall_grade', '')}").font = BOLD_FONT

    row = 4
    for area in ['자율', '동아리', '진로']:
        area_data = changche_analysis.get('영역별', {}).get(area, {})
        ws.cell(row=row, column=1, value=f'{area}활동').font = Font(name="Arial", bold=True, size=11, color="4472C4")
        ws.cell(row=row, column=2, value=f"등급: {area_data.get('overall_grade', '')}").font = BOLD_FONT
        row += 1

        # 영역에 맞는 항목명 가져오기
        from modules.changche_analyzer import CHANGCHE_ITEMS
        item_names = CHANGCHE_ITEMS.get(area, [])
        headers = ['학년'] + item_names + ['종합', '등급', '강점', '보완점',
                   '좋은평가문장', '글자수', '제한', '비율', '분량판정']
        col_widths = [6] + [12]*len(item_names) + [6, 6, 30, 30, 35, 8, 8, 8, 10]

        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        _apply_header_style(ws, row, len(headers))
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(
                ws.column_dimensions[get_column_letter(i)].width or 8, w)
        row += 1

        yearly = area_data.get('학년별', {})
        volume = changche_analysis.get('volume_ratio', {}).get(area, {})
        for year in ['1', '2', '3']:
            yd = yearly.get(year, {})
            if not yd:
                continue
            scores = yd.get('scores', {})
            good_sentences = yd.get('좋은평가문장', [])
            sentence_text = ' / '.join([s.get('문장', '') for s in good_sentences]) if good_sentences else ''
            vol = volume.get(year, {})

            vals = [f'{year}학년']
            for item in item_names:
                vals.append(scores.get(item, {}).get('score', ''))
            vals.extend([
                yd.get('total', ''),
                yd.get('grade', ''),
                _truncate(yd.get('강점', ''), 50),
                _truncate(yd.get('보완점', ''), 50),
                _truncate(sentence_text, 60),
                vol.get('글자수', ''),
                vol.get('제한', ''),
                vol.get('비율%', ''),
                vol.get('판정', ''),
            ])
            for i, v in enumerate(vals, 1):
                ws.cell(row=row, column=i, value=v)
            _apply_data_style(ws, row, len(vals))
            row += 1

        row += 1


def create_haengtuk_sheet(wb: Workbook, haengtuk_analysis: dict):
    """행특분석 시트"""
    ws = wb.create_sheet('행특분석')

    from modules.haengtuk_analyzer import HAENGTUK_ITEMS
    headers = ['학년'] + HAENGTUK_ITEMS + ['종합', '등급', '강점', '보완점', '좋은평가문장', '문장평가이유']
    col_widths = [6] + [12]*5 + [6, 6, 35, 35, 40, 40]
    data_row = _setup_sheet(ws, headers, col_widths, '행특 분석 결과')

    ws.cell(row=2, column=1, value=f"종합등급: {haengtuk_analysis.get('overall_grade', '')} / 성장추이: {haengtuk_analysis.get('growth_trend', '')}").font = BOLD_FONT

    yearly = haengtuk_analysis.get('학년별', {})
    for year in ['1', '2', '3']:
        yd = yearly.get(year, {})
        if not yd:
            continue
        scores = yd.get('scores', {})
        good_sentences = yd.get('좋은평가문장', [])
        sentence_text = ' / '.join([s.get('문장', '') for s in good_sentences]) if good_sentences else ''
        reason_text = ' / '.join([s.get('이유', '') for s in good_sentences]) if good_sentences else ''

        vals = [f'{year}학년']
        for item in HAENGTUK_ITEMS:
            vals.append(scores.get(item, {}).get('score', ''))
        vals.extend([
            yd.get('total', ''),
            yd.get('grade', ''),
            _truncate(yd.get('강점', ''), 60),
            _truncate(yd.get('보완점', ''), 60),
            _truncate(sentence_text, 80),
            _truncate(reason_text, 80),
        ])
        for i, v in enumerate(vals, 1):
            ws.cell(row=data_row, column=i, value=v)
        _apply_data_style(ws, data_row, len(vals))
        data_row += 1


def create_linkage_sheet(wb: Workbook, comprehensive: dict):
    """연계성분석 시트"""
    ws = wb.create_sheet('연계성분석')

    headers = ['분석 항목', '연계도', '상세']
    data_row = _setup_sheet(ws, headers, [25, 8, 80], '연계성 분석')

    # 세특 과목간 연계
    sl = comprehensive.get('setuek_linkage', {})
    ws.cell(row=data_row, column=1, value='세특 과목 간 연계')
    ws.cell(row=data_row, column=2, value=sl.get('score', ''))
    cases = sl.get('사례', [])
    ws.cell(row=data_row, column=3, value=_truncate(' / '.join(cases), 150))
    _apply_data_style(ws, data_row, 3)
    data_row += 1

    # 개별 사례
    for case in cases:
        ws.cell(row=data_row, column=1, value='')
        ws.cell(row=data_row, column=2, value='')
        ws.cell(row=data_row, column=3, value=_truncate(case))
        _apply_data_style(ws, data_row, 3)
        data_row += 1

    data_row += 1

    # 창체-세특 연계
    cl = comprehensive.get('changche_setuek_linkage', {})
    ws.cell(row=data_row, column=1, value='창체-세특 영역 간 연계')
    ws.cell(row=data_row, column=2, value=cl.get('score', ''))
    cases2 = cl.get('사례', [])
    ws.cell(row=data_row, column=3, value=_truncate(' / '.join(cases2), 150))
    _apply_data_style(ws, data_row, 3)
    data_row += 1

    for case in cases2:
        ws.cell(row=data_row, column=1, value='')
        ws.cell(row=data_row, column=2, value='')
        ws.cell(row=data_row, column=3, value=_truncate(case))
        _apply_data_style(ws, data_row, 3)
        data_row += 1

    data_row += 1

    # 성장 스토리
    gs = comprehensive.get('growth_story', {})
    ws.cell(row=data_row, column=1, value='성장 스토리')
    ws.cell(row=data_row, column=2, value=gs.get('pattern', ''))
    ws.cell(row=data_row, column=3, value=_truncate(gs.get('서술', ''), 200))
    _apply_data_style(ws, data_row, 3)


def create_jeonhyung_sheet(wb: Workbook, comprehensive: dict):
    """전형적합도 시트"""
    ws = wb.create_sheet('전형적합도')

    headers = ['항목', '내용']
    data_row = _setup_sheet(ws, headers, [20, 80], '전형 적합도 분석')

    jh = comprehensive.get('recommended_jeonhyung', {})
    rows = [
        ('추천 전형', jh.get('추천전형', '')),
        ('판단 근거', _truncate(jh.get('판단근거', ''), 150)),
    ]

    for label, value in rows:
        ws.cell(row=data_row, column=1, value=label)
        ws.cell(row=data_row, column=2, value=value)
        _apply_data_style(ws, data_row, 2)
        data_row += 1


def create_university_eval_sheet(wb: Workbook, comprehensive: dict):
    """대학평가요소 시트"""
    ws = wb.create_sheet('대학평가요소')

    headers = ['평가 영역', '평가 항목', '분석 결과']
    data_row = _setup_sheet(ws, headers, [18, 18, 60], '대학 평가요소별 매핑')

    mapping = comprehensive.get('university_evaluation_mapping', {})
    for domain, items in mapping.items():
        for item_name, result in items.items():
            ws.cell(row=data_row, column=1, value=domain)
            ws.cell(row=data_row, column=2, value=item_name)
            ws.cell(row=data_row, column=3, value=_truncate(result))
            _apply_data_style(ws, data_row, 3)
            data_row += 1


def create_keyword_sheet(wb: Workbook, comprehensive: dict):
    """키워드분석 시트"""
    ws = wb.create_sheet('키워드분석')

    keyword_data = comprehensive.get('keyword_analysis', {})
    keywords = keyword_data.get('keywords', [])
    by_year = keyword_data.get('by_year', {})

    # 키워드 빈도표
    headers = ['키워드', '빈도', '역량 카테고리', '출현 영역']
    col_widths = [15, 8, 15, 25]
    data_row = _setup_sheet(ws, headers, col_widths, '핵심 키워드 분석')

    for kw in keywords:
        areas_str = ', '.join(kw.get('출현영역', []))
        vals = [kw.get('키워드', ''), kw.get('빈도', 0),
                kw.get('카테고리', ''), areas_str]
        for i, v in enumerate(vals, 1):
            ws.cell(row=data_row, column=i, value=v)
        _apply_data_style(ws, data_row, len(vals))
        data_row += 1

    # 학년별 키워드 변화
    data_row += 2
    ws.cell(row=data_row, column=1, value='학년별 키워드 변화').font = Font(name="Arial", bold=True, size=11, color="4472C4")
    data_row += 1
    year_headers = ['학년', '주요 키워드']
    for i, h in enumerate(year_headers, 1):
        ws.cell(row=data_row, column=i, value=h)
    _apply_header_style(ws, data_row, 2)
    data_row += 1

    for year in ['1', '2', '3']:
        year_kws = by_year.get(year, [])
        ws.cell(row=data_row, column=1, value=f'{year}학년')
        ws.cell(row=data_row, column=2, value=', '.join(year_kws[:15]))
        _apply_data_style(ws, data_row, 2)
        data_row += 1


def create_remedial_sheet(wb: Workbook, comprehensive: dict):
    """역량별보완법 시트"""
    ws = wb.create_sheet('역량별보완법')

    headers = ['역량', '세부 항목', '현재 점수', '보완 대상', '점수 소스', '우선순위']
    col_widths = [12, 15, 10, 10, 35, 10]
    data_row = _setup_sheet(ws, headers, col_widths, '역량별 보완법 상세')

    ws.cell(row=2, column=1, value='기준: 3.0점 이하 항목이 보완 대상').font = Font(name="Arial", size=9, color="FF0000")

    remedial = comprehensive.get('remedial_data', {})
    for domain in ['학업역량', '진로역량', '공동체역량']:
        items = remedial.get(domain, {})
        for item_name, item_data in items.items():
            is_target = item_data.get('보완대상', False)
            priority = '상' if is_target and item_data.get('점수', 0) <= 2.0 else ('중' if is_target else '-')

            vals = [
                domain,
                item_name,
                item_data.get('점수', ''),
                'O' if is_target else '',
                _truncate(item_data.get('소스', ''), 60),
                priority,
            ]
            for i, v in enumerate(vals, 1):
                ws.cell(row=data_row, column=i, value=v)
            _apply_data_style(ws, data_row, len(vals), bold=is_target)
            data_row += 1


def create_attendance_sheet(wb: Workbook, grade_analysis: dict, extracted: dict):
    """출결·봉사 시트"""
    ws = wb.create_sheet('출결봉사')

    # 출결
    ws.cell(row=1, column=1, value='출결 현황').font = Font(name="Arial", bold=True, size=12)
    headers = ['학년', '수업일수', '질병결석', '미인정결석', '기타결석', '지각', '조퇴', '결과']
    col_widths = [6, 10, 10, 10, 10, 8, 8, 8]

    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    row += 1

    attendance = extracted.get('attendance', {})
    for year in ['1', '2', '3']:
        data = attendance.get(year, {})
        if not isinstance(data, dict):
            continue
        absence = data.get('결석', {})
        vals = [
            f'{year}학년',
            data.get('수업일수', ''),
            absence.get('질병', 0),
            absence.get('미인정', 0),
            absence.get('기타', 0),
            _sum_subcategory(data.get('지각', {})),
            _sum_subcategory(data.get('조퇴', {})),
            _sum_subcategory(data.get('결과', {})),
        ]
        for i, v in enumerate(vals, 1):
            ws.cell(row=row, column=i, value=v)
        _apply_data_style(ws, row, len(vals))
        row += 1

    ws.cell(row=row + 1, column=1, value=f"출결 점수: {grade_analysis.get('attendance_score_100', 100)}점/100점").font = BOLD_FONT

    # 봉사활동
    row += 3
    ws.cell(row=row, column=1, value='봉사활동 시수').font = Font(name="Arial", bold=True, size=12)
    row += 1
    vol_headers = ['학년', '시수']
    for i, h in enumerate(vol_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, 2)
    row += 1

    volunteer = extracted.get('volunteer_hours', {})
    for year in ['1', '2', '3']:
        ws.cell(row=row, column=1, value=f'{year}학년')
        ws.cell(row=row, column=2, value=volunteer.get(year, 0))
        _apply_data_style(ws, row, 2)
        row += 1


def _sum_subcategory(data) -> int:
    """하위 카테고리 합산"""
    if isinstance(data, dict):
        return sum(data.values())
    if isinstance(data, (int, float)):
        return int(data)
    return 0


def generate_excel_report(all_results: dict, output_path: str) -> str:
    """Excel 리포트 생성 메인 함수"""
    wb = Workbook()

    extracted = all_results.get('extracted_data', {})
    grade_analysis = all_results.get('grade_analysis', {})
    admission = all_results.get('admission_matching', {})
    setuek = all_results.get('setuek_analysis', {})
    changche = all_results.get('changche_analysis', {})
    haengtuk = all_results.get('haengtuk_analysis', {})
    comprehensive = all_results.get('comprehensive_analysis', {})

    create_summary_sheet(wb, comprehensive, grade_analysis)
    create_grade_sheet(wb, grade_analysis, extracted)
    create_admission_sheet(wb, admission)
    create_setuek_sheet(wb, setuek)
    create_changche_sheet(wb, changche)
    create_haengtuk_sheet(wb, haengtuk)
    create_linkage_sheet(wb, comprehensive)
    create_keyword_sheet(wb, comprehensive)
    create_jeonhyung_sheet(wb, comprehensive)
    create_university_eval_sheet(wb, comprehensive)
    create_remedial_sheet(wb, comprehensive)
    create_attendance_sheet(wb, grade_analysis, extracted)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel 리포트 생성 완료: {output_path}")
    return output_path


def generate_pdf_report(all_results: dict, output_path: str) -> str:
    """PDF 리포트 생성"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # 한글 폰트 등록
    font_registered = False
    font_paths = [
        'C:/Windows/Fonts/malgun.ttf',
        'C:/Windows/Fonts/NanumGothic.ttf',
        '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                font_name = 'KoreanFont'
                pdfmetrics.registerFont(TTFont(font_name, fp))
                font_registered = True
                break
            except Exception:
                continue

    if not font_registered:
        font_name = 'Helvetica'
        print("경고: 한글 폰트를 찾을 수 없습니다. PDF에 한글이 깨질 수 있습니다.")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                           leftMargin=15*mm, rightMargin=15*mm,
                           topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('KTitle', parent=styles['Title'],
                                  fontName=font_name, fontSize=16)
    heading_style = ParagraphStyle('KHeading', parent=styles['Heading2'],
                                    fontName=font_name, fontSize=12)
    body_style = ParagraphStyle('KBody', parent=styles['Normal'],
                                 fontName=font_name, fontSize=9, leading=12)

    blue = HexColor('#4472C4')
    elements = []

    comprehensive = all_results.get('comprehensive_analysis', {})
    grade_analysis = all_results.get('grade_analysis', {})

    # 제목
    student_name = all_results.get('extracted_data', {}).get('student_info', {}).get('name', '학생')
    elements.append(Paragraph(f'{student_name} 학생부 경쟁력 리포트', title_style))
    elements.append(Spacer(1, 10*mm))

    # 종합 요약
    elements.append(Paragraph('종합 요약', heading_style))
    summary_data = [
        ['항목', '결과'],
        ['종합등급', f"{comprehensive.get('overall_grade', '')} ({comprehensive.get('overall_score', 0)}점)"],
        ['추천전형', comprehensive.get('recommended_jeonhyung', {}).get('추천전형', '')],
        ['성적추이', grade_analysis.get('trend_pattern', '')],
        ['성장패턴', comprehensive.get('growth_story', {}).get('pattern', '')],
    ]
    t = Table(summary_data, colWidths=[40*mm, 130*mm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8*mm))

    # 핵심 강점
    elements.append(Paragraph('핵심 강점', heading_style))
    for i, s in enumerate(comprehensive.get('core_strengths', []), 1):
        elements.append(Paragraph(f'{i}. {_truncate(s, 200)}', body_style))
    elements.append(Spacer(1, 5*mm))

    # 보완 영역
    elements.append(Paragraph('보완 필요 영역', heading_style))
    for i, a in enumerate(comprehensive.get('areas_to_improve', []), 1):
        elements.append(Paragraph(f'{i}. {_truncate(a, 200)}', body_style))
    elements.append(Spacer(1, 5*mm))

    # 점수 구성
    elements.append(Paragraph('영역별 점수', heading_style))
    breakdown = comprehensive.get('score_breakdown', {})
    score_data = [
        ['영역', '점수', '가중치', '기여점수'],
        ['내신', f"{breakdown.get('내신', 0)}", '30%', f"{breakdown.get('내신', 0)*0.3:.1f}"],
        ['세특', f"{breakdown.get('세특', 0)}", '30%', f"{breakdown.get('세특', 0)*0.3:.1f}"],
        ['창체', f"{breakdown.get('창체', 0)}", '25%', f"{breakdown.get('창체', 0)*0.25:.1f}"],
        ['행특', f"{breakdown.get('행특', 0)}", '10%', f"{breakdown.get('행특', 0)*0.1:.1f}"],
        ['출결', f"{breakdown.get('출결', 0)}", '5%', f"{breakdown.get('출결', 0)*0.05:.1f}"],
    ]
    t2 = Table(score_data, colWidths=[30*mm, 30*mm, 30*mm, 30*mm])
    t2.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#CCCCCC')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 5*mm))

    # 워드클라우드 이미지 삽입
    wordcloud_path = comprehensive.get('wordcloud_path', '')
    if wordcloud_path and os.path.exists(wordcloud_path):
        from reportlab.platypus import Image
        elements.append(Spacer(1, 8*mm))
        elements.append(Paragraph('핵심 키워드 워드클라우드', heading_style))
        try:
            img = Image(wordcloud_path, width=160*mm, height=80*mm)
            elements.append(img)
        except Exception as e:
            elements.append(Paragraph(f'(워드클라우드 이미지 삽입 실패: {e})', body_style))
        elements.append(Spacer(1, 5*mm))

    # 안내문
    elements.append(Paragraph(
        '* 본 리포트는 학생부 데이터를 기반으로 자동 생성된 참고 자료이며, '
        '실제 입시 결과와 다를 수 있습니다.',
        ParagraphStyle('Disclaimer', parent=body_style, fontSize=8, textColor=HexColor('#999999'))
    ))

    doc.build(elements)
    print(f"PDF 리포트 생성 완료: {output_path}")
    return output_path
