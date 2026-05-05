"""
내신 성적 분석 모듈
- 학기별/학년별 평균 등급 산출
- 5등급→9등급 환산
- 성적 추이 패턴 분석
- 원점수/평균/분포 맥락 분석
- 교과 이수 충실도 평가
- 소인수 과목 표기
- 출결 점수 산출
- 대학별 내신 산출 (average_school_grade_db.xlsx 활용)
"""
import os
from typing import Optional

import openpyxl
import yaml


def load_config(config_path: str = None) -> dict:
    """config.yaml 로드"""
    if config_path is None:
        config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                   'config', 'config.yaml')
    with open(config_path, encoding='utf-8') as f:
        return yaml.safe_load(f)


def load_grade_conversion(xlsx_path: str = None, sheet_name: str = None) -> list:
    """5등급→9등급 환산표 로드. 선형보간에 사용할 테이블 반환."""
    if xlsx_path is None:
        xlsx_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                 'config', 'grade_conversion.xlsx')
    config = load_config()
    if sheet_name is None:
        sheet_name = config.get('grade_conversion_source', '부산광역시교육청')

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    table = []
    if sheet_name == '부산광역시교육청':
        # 부산: A열=5등급 평균(숫자, 일부 None), C열=9등급 평균
        for row in ws.iter_rows(min_row=2, values_only=True):
            five = row[0]
            nine = row[2]
            if five is not None and nine is not None:
                table.append({'five': float(five), 'nine': float(nine)})
    elif sheet_name == '경기진협':
        # 경기진협: A열="1.000 (6)" 형식 문자열, C열=환산 9등급
        for row in ws.iter_rows(min_row=2, values_only=True):
            five_str = row[0]
            nine = row[2]
            if five_str is not None and nine is not None:
                five = float(str(five_str).split('(')[0].strip())
                table.append({'five': five, 'nine': float(nine)})

    wb.close()
    table.sort(key=lambda x: x['five'])
    return table


def convert_5to9(five_grade_avg: float, conversion_table: list = None) -> float:
    """5등급 평균을 9등급으로 환산 (선형보간)"""
    if conversion_table is None:
        conversion_table = load_grade_conversion()

    if not conversion_table:
        return five_grade_avg  # fallback

    # 범위 밖 처리
    if five_grade_avg <= conversion_table[0]['five']:
        return conversion_table[0]['nine']
    if five_grade_avg >= conversion_table[-1]['five']:
        return conversion_table[-1]['nine']

    # 선형보간
    for i in range(len(conversion_table) - 1):
        low = conversion_table[i]
        high = conversion_table[i + 1]
        if low['five'] <= five_grade_avg <= high['five']:
            ratio = (five_grade_avg - low['five']) / (high['five'] - low['five'])
            return round(low['nine'] + ratio * (high['nine'] - low['nine']), 2)

    return conversion_table[-1]['nine']


def _get_subject_grade(subject: dict) -> float:
    """과목의 등급 값 반환. 석차등급 있으면 사용, 없으면 성취도 환산."""
    if subject.get('석차등급') is not None:
        return float(subject['석차등급'])
    # 진로선택과목: 성취도 → 등급 환산
    grade_map = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5}
    return grade_map.get(subject.get('성취도', 'C'), 3)


def _is_major_subject(subject: dict, track: str = '공통') -> bool:
    """주요교과 여부 판별"""
    major_groups = {
        '공통': ['국어', '수학', '영어'],
        '인문': ['국어', '수학', '영어', '사회', '한국사'],
        '자연': ['국어', '수학', '영어', '과학'],
    }
    groups = major_groups.get(track, major_groups['공통'])
    return subject.get('교과군', '') in groups


def estimate_track(grades: dict) -> str:
    """이수 과목으로 인문/자연 계열 추정"""
    science_keywords = ['물리학', '화학', '생명과학', '지구과학', '물리', '화학Ⅱ',
                        '생명과학Ⅱ', '지구과학Ⅱ', '미적분', '기하', '확률과통계']
    humanities_keywords = ['정치와법', '경제', '사회문화', '생활과윤리', '윤리와사상',
                          '세계사', '동아시아사', '한국지리', '세계지리']

    science_count = 0
    humanities_count = 0
    for semester_subjects in grades.values():
        for subj in semester_subjects:
            name = subj.get('과목명', '')
            if any(kw in name for kw in science_keywords):
                science_count += 1
            if any(kw in name for kw in humanities_keywords):
                humanities_count += 1

    if science_count > humanities_count:
        return '자연'
    elif humanities_count > science_count:
        return '인문'
    return '공통'


def calc_semester_average(subjects: list, track: str = None) -> dict:
    """학기별 학점 가중 평균 등급 산출. 전교과 + 주요교과 동시 계산."""
    all_weighted_sum = 0
    all_credit_sum = 0
    major_weighted_sum = 0
    major_credit_sum = 0

    for subj in subjects:
        credit = subj.get('학점', 0)
        if credit <= 0:
            continue
        grade = _get_subject_grade(subj)

        all_weighted_sum += grade * credit
        all_credit_sum += credit

        if track and _is_major_subject(subj, track):
            major_weighted_sum += grade * credit
            major_credit_sum += credit

    result = {
        '전교과': round(all_weighted_sum / all_credit_sum, 2) if all_credit_sum > 0 else 0,
        '주요교과': round(major_weighted_sum / major_credit_sum, 2) if major_credit_sum > 0 else 0,
    }
    return result


def calc_yearly_averages(grades: dict, track: str = '공통') -> dict:
    """학년별 평균 산출 (두 학기 합산)"""
    yearly = {}
    for year in ['1', '2', '3']:
        all_subjects = []
        for semester in [f'{year}-1', f'{year}-2']:
            all_subjects.extend(grades.get(semester, []))
        if all_subjects:
            yearly[year] = calc_semester_average(all_subjects, track)
        else:
            yearly[year] = {'전교과': 0, '주요교과': 0}
    return yearly


def analyze_trend(yearly_averages: dict, threshold: float = 0.3) -> tuple:
    """성적 추이 5유형 판별. 등급 숫자가 작을수록 좋은 성적."""
    avgs = []
    for year in ['1', '2', '3']:
        avg = yearly_averages.get(year, {}).get('전교과', 0)
        if avg > 0:
            avgs.append(avg)

    if len(avgs) < 2:
        return '판별불가', '데이터 부족'

    if len(avgs) == 2:
        diff = avgs[1] - avgs[0]
        if diff < -threshold:
            return '상승형', f'{avgs[0]:.2f} → {avgs[1]:.2f} (등급 향상)'
        elif diff > threshold:
            return '하락형', f'{avgs[0]:.2f} → {avgs[1]:.2f} (등급 하락)'
        else:
            return '안정형', f'{avgs[0]:.2f} → {avgs[1]:.2f} (변동 미미)'

    # 3학년 데이터
    d12 = avgs[1] - avgs[0]  # 1→2 변화 (음수=상승)
    d23 = avgs[2] - avgs[1]  # 2→3 변화

    detail = f'{avgs[0]:.2f} → {avgs[1]:.2f} → {avgs[2]:.2f}'

    if d12 < -threshold and d23 < -threshold:
        return '상승형', f'{detail} (지속적 등급 향상)'
    elif d12 > threshold and d23 < -threshold:
        return 'V자반등형', f'{detail} (중간 하락 후 회복)'
    elif d12 < -threshold and d23 > threshold:
        return '역V자형', f'{detail} (중간 상승 후 하락)'
    elif d12 > threshold and d23 > threshold:
        return '하락형', f'{detail} (지속적 등급 하락)'
    else:
        return '안정형', f'{detail} (등급 변동 미미)'


def find_small_class_subjects(grades: dict, threshold: int = 13) -> list:
    """소인수 과목 (수강자수 <= threshold) 목록"""
    result = []
    for semester, subjects in grades.items():
        for subj in subjects:
            if subj.get('수강자수', 999) <= threshold:
                result.append({
                    '학기': semester,
                    '과목명': subj['과목명'],
                    '수강자수': subj['수강자수']
                })
    return result


def analyze_jinro_subjects(grades: dict) -> list:
    """진로선택과목 성취도 분석"""
    result = []
    for semester, subjects in grades.items():
        for subj in subjects:
            if subj.get('과목유형') == '진로선택':
                result.append({
                    '학기': semester,
                    '과목명': subj['과목명'],
                    '성취도': subj.get('성취도', ''),
                    '성취도별분포비율': subj.get('성취도별분포비율', ''),
                    '수강자수': subj.get('수강자수', 0),
                })
    return result


def analyze_grade_context(grades: dict) -> list:
    """원점수/평균/성취도분포 기반 실질 위치 분석"""
    result = []
    for semester, subjects in grades.items():
        for subj in subjects:
            raw = subj.get('원점수', 0)
            avg = subj.get('과목평균', 0)
            grade = subj.get('성취도', '')
            ratio_str = subj.get('성취도별분포비율', '')

            if not grade or not ratio_str:
                continue

            gap = raw - avg
            ratios = _parse_ratio(ratio_str)
            my_grade_ratio = ratios.get(grade, 0)

            # 분석 텍스트 생성
            if gap >= 15:
                position = '과목 내 최상위권'
            elif gap >= 5:
                position = '과목 내 상위권'
            elif gap >= -5:
                position = '과목 내 중위권'
            else:
                position = '과목 내 하위권'

            if my_grade_ratio <= 10:
                density = f'{grade}등급 비율 {my_grade_ratio}%로 경쟁력 높음'
            elif my_grade_ratio <= 25:
                density = f'{grade}등급 비율 {my_grade_ratio}%로 보통'
            else:
                density = f'{grade}등급 비율 {my_grade_ratio}%로 변별력 낮음'

            result.append({
                '학기': semester,
                '과목명': subj['과목명'],
                '성취도': grade,
                '원점수': raw,
                '과목평균': avg,
                '원점수_평균차': round(gap, 1),
                '실질위치': position,
                '분포분석': density,
            })
    return result


def _parse_ratio(ratio_str: str) -> dict:
    """성취도별분포비율 문자열 파싱. "15.2/25.3/30.1/20.0/9.4" → {'A':15.2, 'B':25.3, ...}"""
    grades = ['A', 'B', 'C', 'D', 'E']
    parts = str(ratio_str).split('/')
    result = {}
    for i, g in enumerate(grades):
        if i < len(parts):
            try:
                result[g] = float(parts[i])
            except ValueError:
                result[g] = 0
        else:
            result[g] = 0
    return result


def load_course_requirements(xlsx_path: str = None,
                              university: str = None,
                              department: str = None) -> dict | None:
    """교과 이수 기준표 로드. 대학+학과에 맞는 핵심/권장 과목 반환.

    ※ 2026-04 스키마 정규화:
        - 단일 시트 "권장과목" (Wide + 쉼표 구분)
        - 열: 대학 | 모집단위 | 핵심과목 | 권장과목 | 비고
        - school-record-analyzer 와 ipsilounge 가 동일 파일 공유

    Args:
        xlsx_path: 파일 경로 (기본 data/course_requirements.xlsx)
        university: 대학명 (예: "서울대")
        department: 모집단위명 (예: "컴퓨터공학부")

    Returns:
        {'핵심': [...], '권장': [...], '대학': university_or_"종합"}
        또는 None (파일 없음/매칭 없음)
    """
    if xlsx_path is None:
        xlsx_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                 'data', 'course_requirements.xlsx')

    if not os.path.exists(xlsx_path):
        return None

    rows = _load_normalized_rows(xlsx_path)
    if not rows:
        return None

    if university and department:
        # 우선순위 1: 대학+모집단위 지정
        row = _lookup_row(rows, university, department)
        if row:
            return {
                '핵심': list(row['핵심과목']),
                '권장': list(row['권장과목']),
                '대학': row['대학'],
            }
        return None
    elif department:
        # 우선순위 2: 모집단위만 지정 → 모든 대학 대상 매칭 후 과반수 종합
        matched = [r for r in rows if _major_match(r['모집단위'], department)]
        if not matched:
            return None
        universal_style = [
            {'핵심': r['핵심과목'], '권장': r['권장과목'], '대학': r['대학']}
            for r in matched
        ]
        return _merge_requirements(universal_style)
    else:
        return None


def _load_normalized_rows(xlsx_path: str) -> list:
    """정규화된 권장과목 시트를 읽어 행 딕셔너리 리스트로 반환."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = '권장과목' if '권장과목' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = []
    for i, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row_cells or len(row_cells) < 2:
            continue
        univ = row_cells[0]
        major = row_cells[1] if len(row_cells) > 1 else None
        if not univ or not major:
            continue
        univ_s = str(univ).strip()
        major_s = str(major).strip()
        # 주석 행(※) 스킵
        if univ_s.startswith('※'):
            continue
        core = _split_courses(row_cells[2] if len(row_cells) > 2 else None)
        recommended = _split_courses(row_cells[3] if len(row_cells) > 3 else None)
        rows.append({
            '대학': univ_s,
            '모집단위': major_s,
            '핵심과목': core,
            '권장과목': recommended,
        })

    wb.close()
    return rows


def _split_courses(cell_value) -> list:
    """쉼표 구분 셀 → 과목 리스트."""
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text or text == '-':
        return []
    return [p.strip() for p in text.split(',') if p.strip()]


def _normalize_key(name: str) -> str:
    """매칭용 키 정규화 (공백 제거 + 소문자)."""
    if not name:
        return ''
    return ''.join(name.split()).lower()


def _lookup_row(rows: list, university: str, department: str) -> dict | None:
    """대학+모집단위로 행 조회. 정확 매칭 → 부분 매칭 순."""
    univ_n = _normalize_key(university)
    major_n = _normalize_key(department)

    # 정확 매칭
    for r in rows:
        if _normalize_key(r['대학']) == univ_n and _normalize_key(r['모집단위']) == major_n:
            return r

    # 대학 정확 + 모집단위 부분 매칭
    for r in rows:
        if _normalize_key(r['대학']) == univ_n:
            r_major = _normalize_key(r['모집단위'])
            if major_n in r_major or r_major in major_n:
                return r

    # 대학 부분 매칭까지 허용
    for r in rows:
        r_univ = _normalize_key(r['대학'])
        if univ_n in r_univ or r_univ in univ_n:
            r_major = _normalize_key(r['모집단위'])
            if major_n in r_major or r_major in major_n:
                return r

    return None


def _major_match(db_major: str, student_major: str) -> bool:
    """모집단위 부분 매칭 (학과만 지정한 경우 전체 대학 스캔용)."""
    a = _normalize_key(db_major)
    b = _normalize_key(student_major)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def _merge_requirements(requirements_list: list) -> dict:
    """여러 대학의 이수 기준을 종합 (과반수 핵심 기준)"""
    from collections import Counter
    core_counter = Counter()
    rec_counter = Counter()
    total = len(requirements_list)

    for req in requirements_list:
        for s in req.get('핵심', []):
            core_counter[s] += 1
        for s in req.get('권장', []):
            rec_counter[s] += 1

    # 과반수 이상 등장한 과목을 핵심으로
    threshold = total / 2
    core = [s for s, c in core_counter.items() if c >= threshold]
    recommended = [s for s, c in rec_counter.items() if c >= threshold and s not in core]

    return {'핵심': core, '권장': recommended, '대학': '종합'}


def evaluate_course_fulfillment(student_subjects: list,
                                 requirements: dict | None) -> dict:
    """교과 이수 충실도 평가 (1~5점)"""
    if not requirements:
        return {
            'score': 0,
            '핵심_이수': [],
            '핵심_미이수': [],
            '권장_이수': [],
            '권장_미이수': [],
            'detail': '교과 이수 기준 데이터 없음 (미평가)'
        }

    core_required = requirements.get('핵심', [])
    recommended = requirements.get('권장', [])

    # 학생 이수 과목 이름 리스트
    student_names = [s.get('과목명', '') for s in student_subjects]

    def is_taken(req_name):
        return any(req_name in sn or sn in req_name for sn in student_names)

    core_taken = [s for s in core_required if is_taken(s)]
    core_missing = [s for s in core_required if not is_taken(s)]
    rec_taken = [s for s in recommended if is_taken(s)]
    rec_missing = [s for s in recommended if not is_taken(s)]

    # 점수 산출
    if not core_required:
        score = 3  # 기준 없으면 중간
        detail = '핵심과목 기준 없음'
    elif len(core_missing) == 0:
        if len(rec_taken) >= 2:
            score = 5
            detail = f'핵심과목 전부 이수 + 권장과목 {len(rec_taken)}개 이수'
        elif len(rec_taken) == 1:
            score = 4
            detail = '핵심과목 전부 이수 + 권장과목 1개 이수'
        else:
            score = 3
            detail = '핵심과목 전부 이수, 권장과목 미이수'
    elif len(core_missing) <= len(core_required) / 2:
        score = 2
        detail = f'핵심과목 일부 미이수: {", ".join(core_missing)}'
    else:
        score = 1
        detail = f'핵심과목 대부분 미이수: {", ".join(core_missing)}'

    return {
        'score': score,
        '핵심_이수': core_taken,
        '핵심_미이수': core_missing,
        '권장_이수': rec_taken,
        '권장_미이수': rec_missing,
        'detail': detail
    }


def calc_attendance_score(attendance: dict, config: dict = None) -> float:
    """출결 데이터를 100점 만점 점수로 환산"""
    if config is None:
        config = load_config()
    scoring = config.get('attendance_scoring', {})
    base = scoring.get('base', 100)

    total_deduction = 0
    for year, data in attendance.items():
        if not isinstance(data, dict):
            continue
        absence = data.get('결석', {})
        total_deduction += absence.get('미인정', 0) * abs(scoring.get('미인정결석_per_day', -5))
        total_deduction += absence.get('질병', 0) * abs(scoring.get('질병결석_per_day', -1))
        total_deduction += absence.get('기타', 0) * abs(scoring.get('질병결석_per_day', -1))

        for category in ['지각', '조퇴', '결과']:
            cat_data = data.get(category, {})
            if isinstance(cat_data, dict):
                for _, count in cat_data.items():
                    total_deduction += count * abs(scoring.get(f'{category}_per_count', -0.5))

    score = max(scoring.get('min_score', 0), base - total_deduction)
    return round(score, 1)


def calc_grade_score_100(overall_avg: float) -> float:
    """등급 평균(1.0~5.0)을 100점 만점으로 환산. 1.0→100, 5.0→0"""
    if overall_avg <= 0:
        return 0
    score = max(0, min(100, (5.0 - overall_avg) / 4.0 * 100))
    return round(score, 1)


def run_grade_analysis(extracted_data: dict, config: dict = None,
                        university: str = None, department: str = None) -> dict:
    """성적 분석 메인 함수. Step 3 전체 실행."""
    if config is None:
        config = load_config()

    grades = extracted_data.get('grades', {})
    attendance = extracted_data.get('attendance', {})

    # 계열 추정
    track = estimate_track(grades)

    # 학기별 평균
    semester_averages = {}
    for semester, subjects in grades.items():
        if subjects:
            semester_averages[semester] = calc_semester_average(subjects, track)

    # 학년별 평균
    yearly_averages = calc_yearly_averages(grades, track)

    # 전체 평균
    all_subjects = []
    for subjects in grades.values():
        all_subjects.extend(subjects)
    overall_average = calc_semester_average(all_subjects, track) if all_subjects else {'전교과': 0, '주요교과': 0}

    # 5→9등급 환산
    conversion_table = load_grade_conversion()
    grade_converted_9 = convert_5to9(overall_average['전교과'], conversion_table)

    # 성적 추이
    threshold = config.get('trend_threshold', 0.3)
    trend_pattern, trend_detail = analyze_trend(yearly_averages, threshold)

    # 소인수 과목
    small_threshold = config.get('scoring', {}).get('small_class_threshold', 13)
    small_class_subjects = find_small_class_subjects(grades, small_threshold)

    # 진로선택과목
    jinro_subjects = analyze_jinro_subjects(grades)

    # 원점수/분포 맥락 분석
    context_analysis = analyze_grade_context(grades)

    # 교과 이수 충실도
    requirements = load_course_requirements(university=university, department=department)
    course_fulfillment = evaluate_course_fulfillment(all_subjects, requirements)

    # 점수 환산
    grade_score_100 = calc_grade_score_100(overall_average['전교과'])
    attendance_score_100 = calc_attendance_score(attendance, config)

    return {
        'track': track,
        'semester_averages': semester_averages,
        'yearly_averages': yearly_averages,
        'overall_average': overall_average,
        'grade_converted_9': grade_converted_9,
        'trend_pattern': trend_pattern,
        'trend_detail': trend_detail,
        'small_class_subjects': small_class_subjects,
        'jinro_subjects': jinro_subjects,
        'context_analysis': context_analysis,
        'course_fulfillment': course_fulfillment,
        'grade_score_100': grade_score_100,
        'attendance_score_100': attendance_score_100,
    }


# ============================================================================
# 대학별 내신 산출 (average_school_grade_db.xlsx 활용)
# ============================================================================
#
# CLAUDE.md §3-1 "전형별 반영 내신 산출" 구현. 대학·전형·전형유형별로
# 정의된 학년 가중치 / 교과 가중치 / 반영학기 / 진로선택 처리 방식 등을
# 적용해 학생의 환산 내신을 산출.
#
# MVP 범위 (1차):
#   - 반영학기 필터링 (예: "1-1~3-1")
#   - 학년 가중치 (모두 0이면 동일 비중)
#   - 교과 가중치 (모두 0이면 학점 가중만)
#   - 진로선택 기본 분기: 석차등급 / 성취도 / 등급or성취도(상위) / 미반영 / 정성평가
#   - 1등급환산점 ↔ 최하등급환산점 사이 선형 환산
#
# Out of scope (향후 작업):
#   - 매트릭스환산 (시트 04 환산점수표 참조 필요)
#   - 변환형 / 가산형 (시트 05 진로융합처리 참조 필요)
#   - 정시 별도식 / 졸업자 별도식
# ============================================================================

# 진로선택 성취도 → 등급 환산 (지표_진로선택='성취도' 케이스용)
_ACHIEVEMENT_TO_GRADE = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5}

# 가중치 컬럼 키 → 학생 데이터의 교과군 키 매핑
_SUBJECT_WEIGHT_MAP = {
    '교과_국어가중치': '국어',
    '교과_수학가중치': '수학',
    '교과_영어가중치': '영어',
    '교과_사회가중치': '사회',
    '교과_과학가중치': '과학',
}

_GRADING_DB_SHEET = '01_플랫통합'
_GRADING_DB_FILENAME = 'average_school_grade_db.xlsx'

_ALL_SEMESTERS = ['1-1', '1-2', '2-1', '2-2', '3-1', '3-2']

# "1-1~3-1" 형태가 아닌, 전체 학기 적용을 의미하는 값들
_FULL_RANGE_TOKENS = {'미명시', '전학기', '전체', '전체학기', '정성평가', ''}

# 반영교과군 키워드 → 학생 데이터의 교과군 풀네임 매핑
# xlsx 의 '반영교과군' 셀은 "국·영·수·과" 같이 한글자 약어로 표기됨
_REFLECT_GROUP_MAP = {
    '국': '국어', '영': '영어', '수': '수학',
    '사': '사회', '과': '과학', '한': '한국사',
}

# track(계열) → 매칭할 '기본' 룰의 반영교과군 패턴
_DEFAULT_TRACK_TO_REFLECT = {
    '인문': '국·영·수·사',
    '자연': '국·영·수·과',
    '종합': '국·영·수·사·과',
    '공통': '국·영·수·사·과',  # estimate_track 기본값
}


def load_university_grading(
    university: str | None = None,
    admission_type: str | None = None,
    admission_category: str | None = None,
    default_track: str = '종합',
    xlsx_path: str | None = None,
) -> dict | None:
    """대학별 내신 산출 룰 로드.

    매칭 우선순위:
      1. 대학명 + 전형명 + 전형유형 정확 매칭
      2. 대학명 + 전형명 매칭 (전형유형 무시)
      3. 대학명 + 전형유형 매칭 (전형명 무시)
      4. 대학명만 매칭 → 첫 번째 행 (대표 룰로 사용)
      5. 매칭 실패 또는 university 미지정 → '기본' 룰 fallback
         · default_track 에 따라 적절한 기본 행 선택
         · '인문' → "기본 (국·영·수·사)"
         · '자연' → "기본 (국·영·수·과)"
         · '종합'(기본값) → "기본 (국·영·수·사·과)"

    학생이 지망 대학·전형을 지정하지 않은 경우엔 university=None 으로 호출하면
    자동으로 '기본' 룰이 적용된다. estimate_track() 으로 계열을 추정한 뒤
    그 결과를 default_track 으로 넘기면 학생 계열에 맞는 기본 룰이 픽된다.

    Args:
        university: 대학명 (예: "서울대학교"). None 이면 '기본' 룰 fallback.
        admission_type: 전형명 (예: "지역균형", "학교장추천").
        admission_category: 전형유형 (예: "학생부교과", "학생부종합", "논술").
        default_track: '기본' 룰 fallback 시 사용할 계열. '인문'/'자연'/'종합'/'공통'.
        xlsx_path: 파일 경로. 기본은 data/average_school_grade_db.xlsx.

    Returns:
        dict: 룰 데이터 (구조는 _row_to_rule 참조). 파일 없으면 None.
    """
    if xlsx_path is None:
        xlsx_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'data', _GRADING_DB_FILENAME,
        )
    if not os.path.exists(xlsx_path):
        return None

    rows = _load_grading_rows(xlsx_path)
    if not rows:
        return None

    matched = _match_grading_row(rows, university, admission_type, admission_category)
    if matched is None:
        # 매칭 실패 또는 university 미지정 → '기본' 룰 fallback
        matched = _select_default_rule(rows, default_track)
    if matched is None:
        return None
    return _row_to_rule(matched)


def _select_default_rule(rows: list, track: str = '종합') -> dict | None:
    """학생 계열에 맞는 '기본' 행을 선택.

    xlsx 에 '기본' 행이 3개 있으며 반영교과군이 다름:
      - "기본 (국·영·수·사)"   ← 인문계용
      - "기본 (국·영·수·과)"   ← 자연계용
      - "기본 (국·영·수·사·과)" ← 종합 (전 교과)

    Args:
        track: 학생 계열. '인문'/'자연'/'종합'/'공통'.
    """
    target_pattern = _DEFAULT_TRACK_TO_REFLECT.get(track, '국·영·수·사·과')
    base_rows = [r for r in rows if str(r.get('대학명', '')).strip() == '기본']

    # 정확 매칭: 전형명에 target_pattern 포함
    for r in base_rows:
        type_str = str(r.get('전형명', '')).strip()
        if target_pattern in type_str:
            return r

    # fallback — 첫 번째 '기본' 행
    return base_rows[0] if base_rows else None


def _load_grading_rows(xlsx_path: str) -> list:
    """xlsx 의 01_플랫통합 시트를 dict 리스트로 로드."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = _GRADING_DB_SHEET if _GRADING_DB_SHEET in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    header = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or v == '' for v in row):
            continue
        d = {h: v for h, v in zip(header, row, strict=False) if h is not None}
        rows.append(d)
    wb.close()
    return rows


def _normalize_for_match(s) -> str:
    """매칭용 정규화 — 공백 제거 + 소문자."""
    if s is None:
        return ''
    return ''.join(str(s).split()).lower()


def _match_grading_row(
    rows: list,
    university: str | None,
    admission_type: str | None,
    admission_category: str | None,
) -> dict | None:
    """우선순위에 따라 최적 매칭 행을 찾아 반환."""
    if not university:
        return None
    u = _normalize_for_match(university)
    a = _normalize_for_match(admission_type) if admission_type else None
    c = _normalize_for_match(admission_category) if admission_category else None

    # 우선순위 1: 대학+전형+전형유형
    if a and c:
        for r in rows:
            if (_normalize_for_match(r.get('대학명')) == u
                    and _normalize_for_match(r.get('전형명')) == a
                    and _normalize_for_match(r.get('전형유형')) == c):
                return r
    # 우선순위 2: 대학+전형
    if a:
        for r in rows:
            if (_normalize_for_match(r.get('대학명')) == u
                    and _normalize_for_match(r.get('전형명')) == a):
                return r
    # 우선순위 3: 대학+전형유형
    if c:
        for r in rows:
            if (_normalize_for_match(r.get('대학명')) == u
                    and _normalize_for_match(r.get('전형유형')) == c):
                return r
    # 우선순위 4: 대학만
    for r in rows:
        if _normalize_for_match(r.get('대학명')) == u:
            return r
    return None


def _safe_float(v, default: float = 0.0) -> float:
    """문자열/None/숫자 혼재 셀을 안전하게 float 로 변환."""
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _row_to_rule(row: dict) -> dict:
    """xlsx 행 dict 를 산출 룰 구조로 변환."""
    year_w = {
        1: _safe_float(row.get('학년_1가중치')),
        2: _safe_float(row.get('학년_2가중치')),
        3: _safe_float(row.get('학년_3가중치')),
    }
    subject_w = {
        kor_key: _safe_float(row.get(col_key))
        for col_key, kor_key in _SUBJECT_WEIGHT_MAP.items()
    }
    type_w = {
        '공통': _safe_float(row.get('유형_공통가중치')),
        '일반선택': _safe_float(row.get('유형_일반선택가중치')),
        '진로선택': _safe_float(row.get('유형_진로선택가중치')),
        '융합선택': _safe_float(row.get('유형_융합선택가중치')),
    }
    return {
        '대학명': row.get('대학명'),
        '전형명': row.get('전형명'),
        '전형유형': row.get('전형유형'),
        '반영학기': str(row.get('반영학기') or '1-1~3-1').strip(),
        '학년_가중치': year_w,
        '교과_가중치': subject_w,
        '유형_가중치': type_w,
        '지표_공통': str(row.get('지표_공통') or '석차등급').strip(),
        '지표_일반선택': str(row.get('지표_일반선택') or '석차등급').strip(),
        '지표_진로선택': str(row.get('지표_진로선택') or '석차등급').strip(),
        '지표_융합_국수영': str(row.get('지표_융합_국수영') or '석차등급').strip(),
        '지표_융합_사과': str(row.get('지표_융합_사과') or '석차등급').strip(),
        '반영교과군': _parse_reflect_groups(row.get('반영교과군')),
        '반영교과군_원문': row.get('반영교과군'),
        '1등급환산점': _safe_float(row.get('1등급환산점'), 1.0),
        '최하등급환산점': _safe_float(row.get('최하등급환산점'), 5.0),
        '매트릭스환산_여부': str(row.get('매트릭스환산_여부') or 'N').strip().upper() == 'Y',
        '변환형_여부': str(row.get('변환형_여부') or 'N').strip().upper() == 'Y',
        '가산형_여부': str(row.get('가산형_여부') or 'N').strip().upper() == 'Y',
        '_raw': row,  # 디버깅/향후 확장용
    }


# 반영교과군 셀에서 "전 과목 반영(필터 없음)" 으로 해석할 키워드
_NO_FILTER_KEYWORDS = (
    '미반영', '미명시', '해당없음', '없음',
    '전교과', '전 교과', '전과목', '전 과목', '전체',
    '정성', '종합평가', '학생부', '교과미반영', '비교과', '서류평가',
    '석차등급 산출 전 교과목',
)


def _parse_reflect_groups(reflect_str) -> set:
    """반영교과군 문자열을 학생 데이터 교과군 set 으로 파싱.

    예) "국·영·수·과"     → {'국어', '영어', '수학', '과학'}
        "국·영·수·사·과"  → {'국어', '영어', '수학', '사회', '과학'}
        "전교과(체·예·교양 제외)" → set()  (= 필터 없음, 전 과목 반영)
        "전교과(정성)"    → set()  (= 필터 없음, 정성평가는 notes 로 별도 안내)
        "미명시"/"-"/None → set()  (= 필터 없음)

    복잡한 표현 (예: "국·수·영·사(인문)/국·수·영·과(자연)") 은 괄호 제거 후
    구분자 split 으로 합집합 처리. 정확한 계열별 분기는 MVP 미지원.
    """
    if reflect_str is None:
        return set()
    s = str(reflect_str).strip()
    if not s or s == '-':
        return set()

    # "전 과목 반영" 패턴 키워드 탐지 (괄호 제거 전에 체크 — "전교과(정성)" 같은 케이스 잡기 위해)
    for kw in _NO_FILTER_KEYWORDS:
        if kw in s:
            return set()

    # 괄호 안 부가 설명 제거 ("국·영(인문)/국·과(자연)" → "국·영/국·과")
    import re
    s_main = re.sub(r'\([^)]*\)', '', s).strip()

    # 다양한 구분자 정규화 (·, ・, /, ,, 가운뎃점 변형들)
    for sep in ('・', ',', '，', '/'):
        s_main = s_main.replace(sep, '·')
    parts = [p.strip() for p in s_main.split('·') if p.strip()]
    result = set()
    for p in parts:
        if p in _REFLECT_GROUP_MAP:
            result.add(_REFLECT_GROUP_MAP[p])
        elif p in {'한국사'}:
            result.add('한국사')
        else:
            # 이미 풀네임("국어"/"영어" 등)이거나 인식 못 하는 토큰
            result.add(p)
    return result


def _parse_semester_range(spec: str) -> list:
    """반영학기 문자열을 학기 코드 리스트로 변환.

    예) "1-1~3-1" → ['1-1', '1-2', '2-1', '2-2', '3-1']
        "1-1~2-2" → ['1-1', '1-2', '2-1', '2-2']
        "1-1"     → ['1-1']
        "미명시"/"전학기"/"" → 전체 6학기 (정성평가 등 학기 무관 케이스)
    """
    spec = (spec or '').strip()
    if spec in _FULL_RANGE_TOKENS:
        return list(_ALL_SEMESTERS)
    if '~' not in spec:
        # "1-1" 같은 단일 학기 패턴만 인정. 그 외 알 수 없는 값은 전체 학기 fallback.
        if '-' in spec and len(spec.split('-')) == 2:
            return [spec]
        return list(_ALL_SEMESTERS)
    start, end = [s.strip() for s in spec.split('~', 1)]
    try:
        sy, ss = (int(x) for x in start.split('-'))
        ey, es = (int(x) for x in end.split('-'))
    except ValueError:
        return list(_ALL_SEMESTERS)
    result = []
    for y in range(sy, ey + 1):
        for s in (1, 2):
            if y == sy and s < ss:
                continue
            if y == ey and s > es:
                continue
            result.append(f'{y}-{s}')
    return result


def _resolve_grade_by_indicator(subject: dict, indicator: str) -> float | None:
    """지표 종류에 따라 과목 등급 값을 산출.

    indicator 예: "석차등급", "성취도", "등급or성취도(상위)", "미반영", "정성평가"
    반환: float 등급 값 또는 None (미반영 / 정성평가 / 산출 불가)
    """
    ind = (indicator or '').strip()
    if '미반영' in ind or '정성평가' in ind:
        # 정성평가는 수치 산출이 의미 없으므로 산출 대상에서 제외
        return None

    seokcha = subject.get('석차등급')
    achieve = subject.get('성취도')
    seokcha_val = float(seokcha) if seokcha is not None else None
    achieve_val = _ACHIEVEMENT_TO_GRADE.get(achieve) if achieve else None

    if ind == '석차등급':
        return seokcha_val if seokcha_val is not None else achieve_val
    if ind == '성취도':
        return achieve_val if achieve_val is not None else seokcha_val
    if '상위' in ind or 'or' in ind.lower():
        # 등급or성취도(상위) — 둘 중 더 좋은 것 (작은 값)
        candidates = [v for v in (seokcha_val, achieve_val) if v is not None]
        return min(candidates) if candidates else None
    # 기타 (변환형/매트릭스 등 향후 처리) — 일단 석차등급 fallback
    return seokcha_val if seokcha_val is not None else achieve_val


def calc_university_specific_average(grades: dict, grading_rule: dict) -> dict:
    """대학별 룰을 적용해 환산 내신 평균 산출.

    Args:
        grades: extracted_data['grades'] — {학기코드: [과목 dict, ...]} 형태.
            과목 dict 키: 과목명, 학점, 석차등급, 성취도, 교과군, 과목유형 등.
        grading_rule: load_university_grading() 결과.

    Returns:
        dict: {
            'university', 'admission_type', 'admission_category',
            '반영학기_적용': [...],     # 실제로 적용된 학기 코드 리스트
            '평균등급': float,           # 가중치 적용된 평균 등급
            '환산점수': float,           # 1등급환산점 ↔ 최하등급환산점 사이 선형 환산
            '적용_과목수': int,
            '제외_과목수': int,           # 미반영/지표 산출 불가
            'breakdown': {
                'by_year': {1: avg, 2: avg, 3: avg},
                'by_category': {국어: avg, 수학: avg, ...},
            },
            'notes': [...],              # 적용된 룰 설명 / 경고
        }
    """
    if grading_rule is None:
        return {'error': 'grading_rule is None — load_university_grading() 결과 확인 필요'}

    semesters_spec = grading_rule.get('반영학기', '1-1~3-1')
    target_semesters = _parse_semester_range(semesters_spec)
    year_w = grading_rule.get('학년_가중치', {})
    subj_w = grading_rule.get('교과_가중치', {})
    reflect_groups = grading_rule.get('반영교과군') or set()
    notes: list = []

    # 미명시/정성평가 안내
    if str(semesters_spec).strip() in _FULL_RANGE_TOKENS:
        notes.append(f'반영학기 "{semesters_spec}" → 전체 학기 fallback 으로 산출 (정성평가 위주 전형)')
    if '정성평가' in str(grading_rule.get('지표_공통', '')):
        notes.append('지표_공통 정성평가 — 수치 산출 결과는 참고용')
    if '정성평가' in str(grading_rule.get('지표_진로선택', '')):
        notes.append('지표_진로선택 정성평가 — 진로선택과목은 산출 제외')
    if reflect_groups:
        notes.append(f'반영교과군 필터: {grading_rule.get("반영교과군_원문")} (이외 교과군은 제외)')

    # MVP 미지원 룰 안내
    if grading_rule.get('매트릭스환산_여부'):
        notes.append('매트릭스환산 룰은 MVP 미구현 — 기본 산출 적용')
    if grading_rule.get('변환형_여부'):
        notes.append('변환형 룰은 MVP 미구현 — 기본 산출 적용')
    if grading_rule.get('가산형_여부'):
        notes.append('가산형 룰은 MVP 미구현 — 기본 산출 적용')

    # 가중치 정규화 정책: 합계 0 이면 가중치 없음(전부 1)
    has_year_weight = sum(year_w.values()) > 0
    has_subj_weight = sum(subj_w.values()) > 0

    weighted_sum = 0.0
    weight_sum = 0.0
    applied_count = 0
    skipped_count = 0
    by_year_acc: dict = {1: [0.0, 0.0], 2: [0.0, 0.0], 3: [0.0, 0.0]}  # [w_sum, w]
    by_cat_acc: dict = {}

    for sem, subjects in grades.items():
        if sem not in target_semesters:
            continue
        try:
            year = int(str(sem).split('-')[0])
        except ValueError:
            year = None

        for subj in subjects or []:
            credit = _safe_float(subj.get('학점'))
            if credit <= 0:
                continue

            # 반영교과군 필터: rule 에 명시돼 있으면 해당 교과군만 반영
            cat = (subj.get('교과군') or '').strip()
            if reflect_groups and cat not in reflect_groups:
                skipped_count += 1
                continue

            # 과목 유형에 따른 지표 결정
            sub_type = (subj.get('과목유형') or '').strip()
            if '진로선택' in sub_type:
                indicator = grading_rule.get('지표_진로선택', '석차등급')
            elif '융합선택' in sub_type:
                indicator = grading_rule.get('지표_융합_국수영', '석차등급')
            elif '일반선택' in sub_type:
                indicator = grading_rule.get('지표_일반선택', '석차등급')
            else:
                indicator = grading_rule.get('지표_공통', '석차등급')

            grade_val = _resolve_grade_by_indicator(subj, indicator)
            if grade_val is None:
                skipped_count += 1
                continue

            # 가중치 산출: 학점 × 학년 가중치 × 교과 가중치
            w = credit
            if has_year_weight and year in year_w:
                w *= max(year_w[year], 0.0)
                if w == 0:
                    skipped_count += 1
                    continue
            if has_subj_weight:
                cat = (subj.get('교과군') or '').strip()
                cat_weight = subj_w.get(cat, 0.0)
                # 가중치 매트릭스에 없는 교과군은 0 처리 → 제외
                if cat_weight == 0:
                    skipped_count += 1
                    continue
                w *= cat_weight

            weighted_sum += grade_val * w
            weight_sum += w
            applied_count += 1

            if year in by_year_acc:
                by_year_acc[year][0] += grade_val * w
                by_year_acc[year][1] += w
            cat = (subj.get('교과군') or '기타').strip()
            by_cat_acc.setdefault(cat, [0.0, 0.0])
            by_cat_acc[cat][0] += grade_val * w
            by_cat_acc[cat][1] += w

    avg_grade = round(weighted_sum / weight_sum, 3) if weight_sum > 0 else 0.0

    # 1등급환산점 ↔ 최하등급환산점 사이 선형 환산
    g1 = grading_rule.get('1등급환산점', 1.0)
    g_low = grading_rule.get('최하등급환산점', 5.0)
    # 등급제(1~5)는 등급 그대로, 점수제(100~0)는 (5-등급)/4*(g1-g_low)+g_low
    if g1 == 1 and g_low == 5:
        score = avg_grade
    else:
        # 등급 1.0 → g1 점, 등급 5.0 → g_low 점
        if avg_grade <= 0:
            score = 0.0
        else:
            ratio = (avg_grade - 1.0) / 4.0
            ratio = max(0.0, min(1.0, ratio))
            score = round(g1 + (g_low - g1) * ratio, 2)

    by_year_avg = {
        y: round(s / w, 3) if w > 0 else 0.0
        for y, (s, w) in by_year_acc.items()
    }
    by_cat_avg = {
        c: round(s / w, 3) if w > 0 else 0.0
        for c, (s, w) in by_cat_acc.items()
    }

    return {
        'university': grading_rule.get('대학명'),
        'admission_type': grading_rule.get('전형명'),
        'admission_category': grading_rule.get('전형유형'),
        '반영학기_적용': target_semesters,
        '평균등급': avg_grade,
        '환산점수': score,
        '적용_과목수': applied_count,
        '제외_과목수': skipped_count,
        'breakdown': {
            'by_year': by_year_avg,
            'by_category': by_cat_avg,
        },
        'notes': notes,
    }


# 기본 룰 3개의 표준 라벨/계열 매핑
_BASELINE_TRACKS = [
    ('자연', '국·영·수·과'),
    ('인문', '국·영·수·사'),
    ('종합', '국·영·수·사·과'),
]


def calc_all_grading(
    grades: dict,
    university: str | None = None,
    admission_type: str | None = None,
    admission_category: str | None = None,
    xlsx_path: str | None = None,
) -> dict:
    """기본 룰 3개 + (지정 시) 대학별 룰을 모두 산출.

    학생 지망 대학·전형 지정 여부와 무관하게 항상 3개 기본 룰
    (자연/인문/종합) 결과가 산출된다. 대학·전형이 지정되고 DB 에서
    매칭되면 그 대학 룰 결과도 추가된다.

    리포트에서 학생을 다각도로 보여주거나, 지망 대학 미정 학생도
    내신 산출이 가능하도록 하는 용도.

    Args:
        grades: extracted_data['grades'] — {학기코드: [과목 dict, ...]}
        university: 지망 대학명 (선택). 미지정이면 기본 3개만 산출.
        admission_type: 지망 전형명 (선택).
        admission_category: 지망 전형유형 (선택).
        xlsx_path: DB 파일 경로 (선택).

    Returns:
        {
            'baseline': [
                {'track': '자연', 'label': '기본 (국·영·수·과)', 'rule': {...}, 'result': {...}},
                {'track': '인문', 'label': '기본 (국·영·수·사)', 'rule': {...}, 'result': {...}},
                {'track': '종합', 'label': '기본 (국·영·수·사·과)', 'rule': {...}, 'result': {...}},
            ],
            'university': {
                'requested': '서울대학교',           # 사용자 입력 그대로
                'admission_type': '...',
                'admission_category': '...',
                'matched': True/False,              # DB 매칭 성공 여부
                'label': '서울대학교 / 지역균형',     # matched=True 시
                'rule': {...},                      # matched=True 시
                'result': {...},                    # matched=True 시
                'message': '...',                   # matched=False 시 안내
            } or None,  # university 미지정 시 None
        }
    """
    baseline: list = []
    for track_name, pattern in _BASELINE_TRACKS:
        rule = load_university_grading(default_track=track_name, xlsx_path=xlsx_path)
        if rule is None:
            continue
        result = calc_university_specific_average(grades, rule)
        baseline.append({
            'track': track_name,
            'label': rule.get('전형명') or f'기본 ({pattern})',
            'rule': rule,
            'result': result,
        })

    university_block = None
    if university:
        rule = load_university_grading(
            university=university,
            admission_type=admission_type,
            admission_category=admission_category,
            xlsx_path=xlsx_path,
        )
        matched = rule is not None and str(rule.get('대학명') or '').strip() != '기본'
        if matched:
            result = calc_university_specific_average(grades, rule)
            label_parts = [str(rule.get('대학명') or university)]
            if rule.get('전형명'):
                label_parts.append(str(rule['전형명']))
            university_block = {
                'requested': university,
                'admission_type': admission_type,
                'admission_category': admission_category,
                'matched': True,
                'label': ' / '.join(label_parts),
                'rule': rule,
                'result': result,
            }
        else:
            # 사용자가 요청한 대학·전형이 DB 에 없음 — 기본 룰은 이미 baseline 에 포함됨
            req_str = university
            if admission_type:
                req_str += f' / {admission_type}'
            university_block = {
                'requested': university,
                'admission_type': admission_type,
                'admission_category': admission_category,
                'matched': False,
                'message': f'요청 "{req_str}" 이(가) DB 에 없어 기본 룰만 산출됨 (baseline 참조)',
            }

    return {
        'baseline': baseline,
        'university': university_block,
    }
