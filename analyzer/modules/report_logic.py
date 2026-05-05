"""
report_logic.py
- 학생부 리포트 생성 본체 (create_excel, create_pdf)
- 학생 데이터는 sd 모듈 객체로 받음
- 공통 상수는 report_constants에서 import
"""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Flowable,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.platypus import Image as RLImage

from .report_constants import (
    CHANGCHE_ITEMS,
    HAENGTUK_ITEMS,
    is_major_mode,
    register_fonts,
    resolve_setuek_items,
    setuek_score_slice_end,
)

# 폰트 등록 (모듈 로드 시 1회)
register_fonts()

# ── Grade thresholds (1-10) ──
def score_to_grade(score):
    if score >= 8.5: return "S"
    if score >= 7.0: return "A"
    if score >= 5.0: return "B"
    if score >= 3.5: return "C"
    return "D"

# ── Colors ──
HEADER_FILL_HEX = "4472C4"
GRADE_COLORS = {
    "S": ("1F4E79", "FFFFFF"),
    "A": ("2E75B6", "FFFFFF"),
    "B": ("9DC3E6", "000000"),
    "C": ("F4B183", "000000"),
    "D": ("FF0000", "FFFFFF"),
}


# ===================================================================
# EXCEL GENERATION
# ===================================================================

def create_excel(sd, xlsx_path, mode_config=None):
    """Excel 리포트 생성. sd: 학생 데이터 모듈, xlsx_path: 출력 파일 경로.

    mode_config (modules.mode_config.ModeConfig, optional):
      - None → 전체 시트 출력 (기본 full 모드와 동일, 하위호환).
      - partial 모드 → 미선택 영역 시트 스킵.
      - no-grade 모드 → 내신/입결 시트 스킵 (현재 미구현이라 NOOP).
    """
    # 2026-04-19: mode_config 기본값 처리 (기본=full, 하위호환)
    if mode_config is None:
        from .mode_config import default_config
        mode_config = default_config()
    inc_setuek   = mode_config.include_setuek
    inc_changche = mode_config.include_changche
    inc_haengtuk = mode_config.include_haengtuk

    setuek_data       = sd.setuek_data
    setuek_comments   = sd.setuek_comments
    comment_keys      = sd.comment_keys
    good_sentences    = sd.good_sentences
    changche_data     = sd.changche_data
    changche_comments = sd.changche_comments
    haengtuk_data     = sd.haengtuk_data
    haengtuk_comments = sd.haengtuk_comments
    linkage_data      = sd.linkage_data
    eval_data         = sd.eval_data
    fix_data          = sd.fix_data
    summary_data      = sd.summary_data
    STUDENT           = sd.STUDENT
    SCHOOL            = sd.SCHOOL
    TODAY             = sd.TODAY

    # 공통 상수 (학생 무관) - report_constants에서 import한 값을 v2 변수명으로 매핑
    changche_items = CHANGCHE_ITEMS
    haengtuk_items = HAENGTUK_ITEMS

    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
    cell_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    def style_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def style_cell(ws, row, col, value, align=center):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = cell_font
        cell.alignment = align
        cell.border = border
        return cell

    def apply_grade_color(cell, grade):
        if grade in GRADE_COLORS:
            bg, fg = GRADE_COLORS[grade]
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.font = Font(name="Arial", size=10, bold=True, color=fg)

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: 종합요약 ──
    ws = wb.active
    ws.title = "종합요약"
    headers = ["항목", "내용"]
    style_header(ws, headers)
    rows = [
        ("학생명", summary_data["학생"]),
        ("학교명", summary_data["학교명"]),
        ("비교과 종합", summary_data["비교과종합"]),
        ("비교과 종합점수", summary_data["비교과종합점수"]),
        ("핵심강점1", summary_data["핵심강점1"]),
        ("핵심강점2", summary_data["핵심강점2"]),
        ("핵심강점3", summary_data["핵심강점3"]),
        ("보완영역1", summary_data["보완영역1"]),
        ("보완영역2", summary_data["보완영역2"]),
        ("보완영역3", summary_data["보완영역3"]),
        ("성장스토리", summary_data["성장스토리"]),
    ]
    for r, (k, v) in enumerate(rows, 2):
        style_cell(ws, r, 1, k)
        style_cell(ws, r, 2, v, left_align)
    set_col_widths(ws, [18, 90])
    ws.auto_filter.ref = f"A1:B{len(rows)+1}"
    ws.freeze_panes = "A2"

    # ── Sheet 2~4: 세특 영역 (mode_config 로 스킵 가능, 2026-04-19) ──
    # TARGET_MAJOR 유무에 따라 6항목/7항목 동적 분기
    setuek_items_local = resolve_setuek_items(sd)
    n_items = len(setuek_items_local)           # 6 또는 7
    score_end = setuek_score_slice_end(sd)      # 8(미지정) 또는 9(지정)

    if inc_setuek:
        # ── Sheet 2: 세특분석 ──
        ws = wb.create_sheet("세특분석")
        headers = ["학년", "과목"] + setuek_items_local + ["가중합산", "등급"]
        style_header(ws, headers)
        for r, d in enumerate(setuek_data, 2):
            yr, subj = d[0], d[1]
            scores = d[2:score_end]
            wavg, grade = d[score_end], d[score_end + 1]
            style_cell(ws, r, 1, yr)
            style_cell(ws, r, 2, subj)
            for ci, s in enumerate(scores):
                style_cell(ws, r, 3 + ci, s)
            style_cell(ws, r, 3 + n_items, wavg)
            gc = style_cell(ws, r, 4 + n_items, grade)
            apply_grade_color(gc, grade)
        # 컬럼 폭: [학년, 과목, 각 항목..., 가중합산, 등급]
        col_widths = [6, 18] + [10] * n_items + [10, 8]
        set_col_widths(ws, col_widths)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

        # ── Sheet 3: 세특코멘트 ──
        ws = wb.create_sheet("세특코멘트")
        headers = ["학년", "과목", "강점", "보완점"]
        style_header(ws, headers)
        for r, key in enumerate(comment_keys, 2):
            # parse year from key
            yr = key[-2]  # "1", "2", or "3"
            subj = key[:-3] if key[-3] == "(" else key  # remove (N)
            # Better parsing
            parts = key.rsplit("(", 1)
            subj_name = parts[0]
            yr_num = parts[1].rstrip(")") if len(parts) > 1 else ""
            strength, weakness = setuek_comments[key]
            style_cell(ws, r, 1, yr_num)
            style_cell(ws, r, 2, subj_name)
            style_cell(ws, r, 3, strength, left_align)
            style_cell(ws, r, 4, weakness, left_align)
        set_col_widths(ws, [6, 18, 70, 70])
        ws.auto_filter.ref = "A1:D1"
        ws.freeze_panes = "A2"

        # ── Sheet 4: 핵심평가문장 ──
        ws = wb.create_sheet("핵심평가문장")
        headers = ["과목", "핵심문장", "이유", "표현역량"]
        style_header(ws, headers)
        for r, (subj, sent, reason, comp) in enumerate(good_sentences, 2):
            style_cell(ws, r, 1, subj)
            style_cell(ws, r, 2, sent, left_align)
            style_cell(ws, r, 3, reason, left_align)
            style_cell(ws, r, 4, comp, left_align)
        set_col_widths(ws, [16, 60, 55, 40])
        ws.auto_filter.ref = "A1:D1"
        ws.freeze_panes = "A2"

    # ── Sheet 5: 창체분석 (mode_config 로 스킵 가능) ──
    if inc_changche:
        ws = wb.create_sheet("창체분석")
        headers = ["학년", "영역", "시간"] + changche_items + ["합계(/50)", "환산(100)", "등급", "분량", "비율", "활용"]
        style_header(ws, headers)
        for r, d in enumerate(changche_data, 2):
            yr, area, hours = d[0], d[1], d[2]
            scores = d[3]
            total, scaled = d[4], d[5]
            vol, pct, usage = d[6], d[7], d[8]
            grade = score_to_grade(scaled / 10)
            style_cell(ws, r, 1, yr)
            style_cell(ws, r, 2, area)
            style_cell(ws, r, 3, hours)
            for ci, s in enumerate(scores):
                style_cell(ws, r, 4+ci, s)
            style_cell(ws, r, 9, f"{total}/50")
            style_cell(ws, r, 10, scaled)
            gc = style_cell(ws, r, 11, grade)
            apply_grade_color(gc, grade)
            style_cell(ws, r, 12, vol)
            style_cell(ws, r, 13, pct)
            style_cell(ws, r, 14, usage)
        set_col_widths(ws, [6, 8, 6, 14, 14, 14, 14, 12, 10, 10, 8, 12, 8, 10])
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

    # ── Sheet 6: 행특분석 (mode_config 로 스킵 가능) ──
    if inc_haengtuk:
        ws = wb.create_sheet("행특분석")
        headers = ["학년", "강점", "보완점"]
        style_header(ws, headers)
        haengtuk_years_sorted = sorted(haengtuk_comments.keys())
        for r, yr in enumerate(haengtuk_years_sorted, 2):
            vals = haengtuk_comments[yr]
            strength, weakness = vals[0], vals[1]
            style_cell(ws, r, 1, yr)
            style_cell(ws, r, 2, strength, left_align)
            style_cell(ws, r, 3, weakness, left_align)
        set_col_widths(ws, [6, 75, 75])
        ws.auto_filter.ref = "A1:C1"
        ws.freeze_panes = "A2"

    # ── Sheet 7: 연계성분석 ──
    ws = wb.create_sheet("연계성분석")
    headers = ["분석영역", "연계도", "상세내용"]
    style_header(ws, headers)
    for r, (area, level, detail) in enumerate(linkage_data, 2):
        style_cell(ws, r, 1, area)
        style_cell(ws, r, 2, level)
        style_cell(ws, r, 3, detail, left_align)
    set_col_widths(ws, [18, 10, 90])
    ws.auto_filter.ref = "A1:C1"
    ws.freeze_panes = "A2"

    # ── Sheet 8: 대학평가요소 ──
    ws = wb.create_sheet("대학평가요소")
    headers = ["대분류", "세부항목", "등급", "근거"]
    style_header(ws, headers)
    for r, (cat, item, grade, reason) in enumerate(eval_data, 2):
        style_cell(ws, r, 1, cat)
        style_cell(ws, r, 2, item)
        gc = style_cell(ws, r, 3, grade)
        apply_grade_color(gc, grade)
        style_cell(ws, r, 4, reason, left_align)
    set_col_widths(ws, [14, 16, 8, 60])
    ws.auto_filter.ref = "A1:D1"
    ws.freeze_panes = "A2"

    # ── Sheet 9: 역량별보완법 ──
    ws = wb.create_sheet("역량별보완법")
    headers = ["역량", "항목", "현재등급", "진단", "보완활동", "중요도"]
    style_header(ws, headers)
    for r, (cap, item, cur, diag, action, pri) in enumerate(fix_data, 2):
        style_cell(ws, r, 1, cap)
        style_cell(ws, r, 2, item)
        style_cell(ws, r, 3, cur)
        style_cell(ws, r, 4, diag, left_align)
        style_cell(ws, r, 5, action, left_align)
        style_cell(ws, r, 6, pri)
    set_col_widths(ws, [10, 14, 10, 30, 80, 8])
    ws.auto_filter.ref = "A1:F1"
    ws.freeze_panes = "A2"

    # ── Sheet 10: 키워드분석 (G5 / CLAUDE.md § Step 8-4) ──
    # raw_texts 가 비어있으면 이 시트 스킵.
    try:
        from .keyword_extractor import has_raw_texts
        if has_raw_texts(sd):
            _write_keyword_sheet(wb, sd, style_header, style_cell, set_col_widths,
                                  center, left_align, get_column_letter)
    except Exception as e:
        print(f"[WARN] 키워드분석 시트 생성 실패 (스킵): {type(e).__name__}: {e}")

    # ── Sheet 11: 출결·봉사 (G7 / CLAUDE.md § 4, §12) ──
    # attendance_data / volunteer_data 모두 비어있으면 스킵.
    try:
        from .attendance_calculator import has_attendance_or_volunteer
        if has_attendance_or_volunteer(sd):
            _write_attendance_sheet(wb, sd, style_header, style_cell, set_col_widths,
                                     center, left_align, get_column_letter)
    except Exception as e:
        print(f"[WARN] 출결·봉사 시트 생성 실패 (스킵): {type(e).__name__}: {e}")

    # ── Sheet 12: 이전분석대비변화 (G3+G4 / CLAUDE.md § Step 0-4, §12) ──
    # compare_data 비어있으면 스킵 (최초 분석).
    try:
        from .compare_generator import has_compare_data
        if has_compare_data(sd):
            _write_compare_sheet(wb, sd, style_header, style_cell, set_col_widths,
                                  center, left_align, get_column_letter)
    except Exception as e:
        print(f"[WARN] 이전분석대비변화 시트 생성 실패 (스킵): {type(e).__name__}: {e}")

    # ── Sheet 13: 대학별내신 (2026-05-05 신규, average_school_grade_db.xlsx) ──
    # grade_data 비어있으면 스킵.
    try:
        if _has_grade_data(sd):
            _write_grading_sheet(wb, sd, style_header, style_cell, set_col_widths,
                                  center, left_align)
    except Exception as e:
        print(f"[WARN] 대학별내신 시트 생성 실패 (스킵): {type(e).__name__}: {e}")

    wb.save(str(xlsx_path))
    print(f"Excel saved: {xlsx_path}")


def _has_grade_data(sd) -> bool:
    """학생 데이터 파일에 grade_data 가 입력되어 있는지 확인."""
    grade_data = getattr(sd, "grade_data", None) or {}
    if not grade_data:
        return False
    # 한 학기라도 과목이 있으면 True
    return any(subjects for subjects in grade_data.values() if subjects)


def _write_grading_sheet(wb, sd, style_header, style_cell, set_col_widths,
                          center, left_align):
    """대학별 내신 산출 결과 시트 작성.

    baseline 3개(자연/인문/종합) + 지망 대학 룰(있으면) 의 평균등급/환산점수를
    표로 표시. 학년별·교과별 breakdown 도 함께.
    """
    from .grade_analyzer import calc_all_grading

    grade_data = getattr(sd, "grade_data", {}) or {}
    target_univ      = (getattr(sd, "TARGET_UNIV", "") or "").strip()
    target_admission = (getattr(sd, "TARGET_ADMISSION_TYPE", "") or "").strip()
    target_category  = (getattr(sd, "TARGET_ADMISSION_CATEGORY", "") or "").strip()

    out = calc_all_grading(
        grade_data,
        university=target_univ or None,
        admission_type=target_admission or None,
        admission_category=target_category or None,
    )

    ws = wb.create_sheet("대학별내신")
    row = 1

    # ── 표 1: 산출 룰별 요약 ──
    ws.cell(row=row, column=1, value="■ 산출 룰별 평균등급 / 환산점수").font = \
        __import__('openpyxl').styles.Font(name="Arial", bold=True, size=11)
    row += 1

    headers = ["구분", "라벨", "평균등급", "환산점수", "적용 과목수", "제외 과목수"]
    style_header(ws, headers, row=row)
    row += 1

    for b in out.get('baseline') or []:
        r = b.get('result') or {}
        style_cell(ws, row, 1, f"baseline ({b.get('track', '')})")
        style_cell(ws, row, 2, b.get('label', ''), left_align)
        style_cell(ws, row, 3, r.get('평균등급', 0))
        style_cell(ws, row, 4, r.get('환산점수', 0))
        style_cell(ws, row, 5, r.get('적용_과목수', 0))
        style_cell(ws, row, 6, r.get('제외_과목수', 0))
        row += 1

    u = out.get('university')
    if u:
        if u.get('matched'):
            r = u.get('result') or {}
            style_cell(ws, row, 1, "지망 대학")
            style_cell(ws, row, 2, u.get('label', ''), left_align)
            style_cell(ws, row, 3, r.get('평균등급', 0))
            style_cell(ws, row, 4, r.get('환산점수', 0))
            style_cell(ws, row, 5, r.get('적용_과목수', 0))
            style_cell(ws, row, 6, r.get('제외_과목수', 0))
            row += 1
        else:
            style_cell(ws, row, 1, "지망 대학")
            style_cell(ws, row, 2, u.get('message', '미매칭'), left_align)
            row += 1

    # ── 표 2: 학년별 breakdown (지망 대학 우선, 없으면 baseline 종합) ──
    row += 1
    ws.cell(row=row, column=1, value="■ 학년별 평균등급 (대표 룰 기준)").font = \
        __import__('openpyxl').styles.Font(name="Arial", bold=True, size=11)
    row += 1

    repr_block = u if (u and u.get('matched')) else (
        next((b for b in (out.get('baseline') or []) if b.get('track') == '종합'), None)
    )
    repr_label = repr_block.get('label', '-') if repr_block else '-'
    repr_result = (repr_block or {}).get('result') or {}
    by_year = repr_result.get('breakdown', {}).get('by_year', {})

    style_header(ws, ["룰", "1학년", "2학년", "3학년"], row=row)
    row += 1
    style_cell(ws, row, 1, repr_label, left_align)
    style_cell(ws, row, 2, by_year.get(1, 0))
    style_cell(ws, row, 3, by_year.get(2, 0))
    style_cell(ws, row, 4, by_year.get(3, 0))
    row += 1

    # ── 표 3: 교과별 breakdown ──
    row += 1
    ws.cell(row=row, column=1, value="■ 교과별 평균등급 (대표 룰 기준)").font = \
        __import__('openpyxl').styles.Font(name="Arial", bold=True, size=11)
    row += 1

    by_cat = repr_result.get('breakdown', {}).get('by_category', {})
    if by_cat:
        cats = list(by_cat.keys())
        style_header(ws, ["룰"] + cats, row=row)
        row += 1
        style_cell(ws, row, 1, repr_label, left_align)
        for i, c in enumerate(cats, 2):
            style_cell(ws, row, i, by_cat[c])
        row += 1

    # ── 표 4: 적용된 룰 메모 (notes) ──
    notes_collected = []
    for b in out.get('baseline') or []:
        for n in (b.get('result') or {}).get('notes') or []:
            notes_collected.append(f"[{b.get('track', '')}] {n}")
    if u and u.get('matched'):
        for n in (u.get('result') or {}).get('notes') or []:
            notes_collected.append(f"[지망] {n}")

    if notes_collected:
        row += 1
        ws.cell(row=row, column=1, value="■ 산출 메모").font = \
            __import__('openpyxl').styles.Font(name="Arial", bold=True, size=11)
        row += 1
        for note in notes_collected:
            style_cell(ws, row, 1, note, left_align)
            row += 1

    set_col_widths(ws, [22, 40, 12, 12, 12, 12])
    ws.freeze_panes = "A3"


def _write_compare_sheet(wb, sd, style_header, style_cell, set_col_widths,
                          center, left_align, get_column_letter_fn):
    """Excel 이전분석대비변화 시트 작성. G3+G4 (2026-04-17)."""
    from openpyxl.styles import Font, PatternFill

    cd = sd.compare_data or {}
    ws = wb.create_sheet("이전분석대비변화")

    row = 1
    # ── 1. 이전 리포트 정보 ──
    ws.cell(row=row, column=1, value="이전 리포트 정보").font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1
    ws.cell(row=row, column=1, value="회차").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=2, value=f"이전 {cd.get('previous_round', '?')}회차 → 현재")
    row += 1
    ws.cell(row=row, column=1, value="이전 리포트 날짜").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=2, value=cd.get("previous_date", "-"))
    row += 1
    ws.cell(row=row, column=1, value="이전 버전").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=2, value=f"v{cd.get('previous_version', 1)}")
    row += 2

    # ── 2. 등급 변화 추이 ──
    ws.cell(row=row, column=1, value="등급 변화 추이").font = Font(name="Arial", bold=True, size=11)
    row += 1
    headers = ["영역", "이전", "현재", "변화"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = center
    row += 1
    for gc in cd.get("grade_changes", []) or []:
        style_cell(ws, row, 1, gc.get("영역", "-"))
        style_cell(ws, row, 2, gc.get("이전", "-"))
        style_cell(ws, row, 3, gc.get("현재", "-"))
        arrow = gc.get("변화", "-")
        arrow_cell = style_cell(ws, row, 4, arrow)
        if arrow == "↑":
            arrow_cell.font = Font(name="Arial", bold=True, color="2E75B6")
        elif arrow == "↓":
            arrow_cell.font = Font(name="Arial", bold=True, color="C00000")
        row += 1
    row += 1

    # ── 3. 이전 핵심강점 추적 ──
    ws.cell(row=row, column=1, value="이전 핵심강점 추적 (현재 상태)").font = Font(name="Arial", bold=True, size=11)
    row += 1
    headers = ["이전 강점", "출처", "현재 상태", "근거"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = center
    row += 1
    for s in cd.get("strengths_tracking", []) or []:
        style_cell(ws, row, 1, s.get("이전강점", ""), left_align)
        style_cell(ws, row, 2, s.get("출처", ""))
        state = s.get("현재상태", "")
        state_cell = style_cell(ws, row, 3, state)
        if state == "강화됨":
            state_cell.font = Font(name="Arial", bold=True, color="2E75B6")
        elif state == "약화됨":
            state_cell.font = Font(name="Arial", bold=True, color="C00000")
        style_cell(ws, row, 4, s.get("근거", ""), left_align)
        row += 1
    row += 1

    # ── 4. 이전 보완점 반영도 ──
    ws.cell(row=row, column=1, value="이전 보완점 반영도").font = Font(name="Arial", bold=True, size=11)
    row += 1
    headers = ["이전 보완점", "출처", "상태", "근거"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = center
    row += 1
    for s in cd.get("issues_tracking", []) or []:
        style_cell(ws, row, 1, s.get("이전보완점", ""), left_align)
        style_cell(ws, row, 2, s.get("출처", ""))
        state = s.get("상태", "")
        state_cell = style_cell(ws, row, 3, state)
        if state == "반영됨":
            state_cell.font = Font(name="Arial", bold=True, color="2E75B6")
        elif state == "미반영":
            state_cell.font = Font(name="Arial", bold=True, color="C00000")
        elif state == "부분반영":
            state_cell.font = Font(name="Arial", bold=True, color="ED7D31")
        style_cell(ws, row, 4, s.get("근거", ""), left_align)
        row += 1
    row += 1

    # ── 5. 새 강점 / 새 보완점 ──
    ws.cell(row=row, column=1, value="새로 발견된 강점").font = Font(name="Arial", bold=True, size=11)
    row += 1
    for i, s in enumerate(cd.get("new_strengths", []) or [], 1):
        style_cell(ws, row, 1, f"{i}. {s}", left_align)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="새로 발견된 보완점").font = Font(name="Arial", bold=True, size=11)
    row += 1
    for i, s in enumerate(cd.get("new_issues", []) or [], 1):
        style_cell(ws, row, 1, f"{i}. {s}", left_align)
        row += 1
    row += 1

    # ── 6. 성장 코멘트 ──
    ws.cell(row=row, column=1, value="성장 코멘트 (200자 이상)").font = Font(name="Arial", bold=True, size=11)
    row += 1
    style_cell(ws, row, 1, cd.get("growth_comment", ""), left_align)
    growth = str(cd.get("growth_comment", "") or "")
    if growth and len(growth) < 200:
        ws.cell(row=row, column=1).font = Font(name="Arial", color="C00000")

    set_col_widths(ws, [35, 24, 12, 70])
    ws.freeze_panes = "A2"


def _write_attendance_sheet(wb, sd, style_header, style_cell, set_col_widths,
                             center, left_align, get_column_letter_fn):
    """Excel 출결·봉사 시트 작성. G7 (2026-04-17)."""
    from openpyxl.styles import Font, PatternFill

    from .attendance_calculator import calculate_attendance_score, summarize_volunteer

    att_data = getattr(sd, "attendance_data", {}) or {}
    vol_data = getattr(sd, "volunteer_data", {}) or {}
    att_report = calculate_attendance_score(att_data)
    vol_summary = summarize_volunteer(vol_data)

    ws = wb.create_sheet("출결·봉사")

    # ── 1부: 출결 학년별 상세 (4종×3사유) ──
    ws.cell(row=1, column=1, value="출결 현황 (학년별)").font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=1, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    headers = ["학년",
               "결석_질병", "결석_미인정", "결석_기타",
               "지각_질병", "지각_미인정", "지각_기타",
               "조퇴_질병", "조퇴_미인정", "조퇴_기타",
               "결과_질병", "결과_미인정", "결과_기타",
               "특기사항"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = center

    row_end = 2
    if att_report.has_data:
        for i, (yr, entry) in enumerate(sorted(att_report.by_year.items()), 1):
            r = 2 + i
            row_end = r
            style_cell(ws, r, 1, f"{yr}학년")
            col = 2
            for t in ("결석", "지각", "조퇴", "결과"):
                sub = entry[t]
                for reason in ("질병", "미인정", "기타"):
                    style_cell(ws, r, col, sub[reason])
                    col += 1
            style_cell(ws, r, col, entry.get("특기사항", ""), left_align)
    else:
        style_cell(ws, 3, 1, "출결 데이터 없음 (만점 100점 가정)", left_align)
        row_end = 3

    # ── 2부: 출결 점수 요약 ──
    summary_row = row_end + 2
    ws.cell(row=summary_row, column=1, value="출결 점수 (100점 만점)").font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=summary_row + 1, column=1, value="기본점수").font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row + 1, column=2, value=att_report.base)
    ws.cell(row=summary_row + 2, column=1, value="감점 합계").font = Font(name="Arial", bold=True)
    total_ded = sum(att_report.deductions.values())
    ws.cell(row=summary_row + 2, column=2, value=round(total_ded, 1))
    ws.cell(row=summary_row + 3, column=1, value="최종점수").font = Font(name="Arial", bold=True)
    final_cell = ws.cell(row=summary_row + 3, column=2, value=att_report.score)
    final_cell.font = Font(name="Arial", bold=True,
                            color="2E75B6" if att_report.score >= 95 else "F4B183" if att_report.score >= 80 else "FF0000")

    # 감점 내역 (있을 때만)
    if att_report.deductions:
        detail_row = summary_row + 5
        ws.cell(row=detail_row, column=1, value="감점 내역").font = Font(name="Arial", bold=True)
        for i, (label, val) in enumerate(att_report.deductions.items(), 1):
            ws.cell(row=detail_row + i, column=1, value=label)
            ws.cell(row=detail_row + i, column=2, value=round(val, 1))

    # ── 3부: 봉사활동 ──
    vol_row = summary_row + 7 + len(att_report.deductions)
    ws.cell(row=vol_row, column=1, value="봉사활동 시수").font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=vol_row + 1, column=1, value="총 봉사시간").font = Font(name="Arial", bold=True)
    ws.cell(row=vol_row + 1, column=2, value=f"{vol_summary.total_hours}시간" if vol_summary.has_data else "-")

    if vol_summary.has_data:
        ws.cell(row=vol_row + 2, column=1, value="학년별 봉사시간 / 주요 활동").font = Font(name="Arial", bold=True)
        for i, (yr, entry) in enumerate(sorted(vol_summary.by_year.items()), 1):
            style_cell(ws, vol_row + 2 + i, 1, f"{yr}학년 ({entry['hours']}h)")
            acts = entry.get("activities") or []
            style_cell(ws, vol_row + 2 + i, 2, ", ".join(acts) if acts else "-", left_align)

    # 컬럼 폭
    col_widths = [14, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 40]
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "A3"


def _write_keyword_sheet(wb, sd, style_header, style_cell, set_col_widths,
                          center, left_align, get_column_letter_fn):
    """Excel 키워드분석 시트 작성. raw_texts 존재 시에만 호출됨."""
    from .keyword_extractor import extract_keywords
    report = extract_keywords(sd.raw_texts, top_n=50, min_frequency=2)
    if not report.keywords:
        return  # 키워드 0개면 시트 생성 안 함

    # 학년 목록 (빈도 기록된 학년만)
    all_years = sorted({yr for e in report.keywords for yr in e.year_frequencies.keys()})

    ws = wb.create_sheet("키워드분석")

    # 1부: 빈도 상위 키워드 테이블
    headers = ["키워드", "총빈도", "역량 카테고리", "출현 영역"] + [f"{yr}학년" for yr in all_years]
    style_header(ws, headers)
    for r, e in enumerate(report.keywords, 2):
        style_cell(ws, r, 1, e.word)
        style_cell(ws, r, 2, e.frequency)
        style_cell(ws, r, 3, e.category)
        style_cell(ws, r, 4, " / ".join(e.areas) if e.areas else "-")
        for i, yr in enumerate(all_years):
            style_cell(ws, r, 5 + i, e.year_frequencies.get(yr, 0))
    col_widths = [14, 8, 14, 18] + [8] * len(all_years)
    set_col_widths(ws, col_widths)
    ws.auto_filter.ref = f"A1:{get_column_letter_fn(len(headers))}1"
    ws.freeze_panes = "A2"

    # 하단에 여백 후 2부: 학년별 변화 요약 (신규 등장 / 사라진 키워드)
    change_start_row = len(report.keywords) + 4
    ws.cell(row=change_start_row - 1, column=1, value="학년별 키워드 변화 추이").font = \
        __import__("openpyxl").styles.Font(name="Arial", bold=True, size=11)

    ws.cell(row=change_start_row, column=1, value="학년").font = \
        __import__("openpyxl").styles.Font(name="Arial", bold=True, size=10)
    ws.cell(row=change_start_row, column=2, value="신규 등장 키워드").font = \
        __import__("openpyxl").styles.Font(name="Arial", bold=True, size=10)
    ws.cell(row=change_start_row, column=3, value="사라진 키워드").font = \
        __import__("openpyxl").styles.Font(name="Arial", bold=True, size=10)

    for i, (yr, ch) in enumerate(sorted(report.yearly_changes.items()), 1):
        ws.cell(row=change_start_row + i, column=1, value=f"{yr}학년")
        ws.cell(row=change_start_row + i, column=2, value=", ".join(ch.get("new", [])[:15]) or "-")
        ws.cell(row=change_start_row + i, column=3, value=", ".join(ch.get("disappeared", [])[:15]) or "-")


# ===================================================================
# PDF GENERATION
# ===================================================================

def create_pdf(sd, pdf_path, mode_config=None):
    """PDF 리포트 생성. sd: 학생 데이터 모듈, pdf_path: 출력 파일 경로.

    mode_config (modules.mode_config.ModeConfig, optional):
      - None → 전체 섹션 출력 (기본 full 모드와 동일, 하위호환).
      - partial → 미선택 영역 섹션 스킵 + 섹션 번호 동적 재할당.
    """
    # 2026-04-19: mode_config 기본값 처리 (기본=full, 하위호환)
    if mode_config is None:
        from .mode_config import default_config
        mode_config = default_config()
    inc_setuek   = mode_config.include_setuek
    inc_changche = mode_config.include_changche
    inc_haengtuk = mode_config.include_haengtuk

    setuek_data       = sd.setuek_data
    setuek_comments   = sd.setuek_comments
    comment_keys      = sd.comment_keys
    good_sentences    = sd.good_sentences
    changche_data     = sd.changche_data
    changche_comments = sd.changche_comments
    haengtuk_data     = sd.haengtuk_data
    haengtuk_comments = sd.haengtuk_comments
    linkage_data      = sd.linkage_data
    eval_data         = sd.eval_data
    fix_data          = sd.fix_data
    summary_data      = sd.summary_data
    STUDENT           = sd.STUDENT
    SCHOOL            = sd.SCHOOL
    TODAY             = sd.TODAY

    LOGO_PATH = os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png")

    def _footer_cover_page(canvas, doc):
        """표지: 로고 없음, 페이지 번호 없음"""
        pass

    def _footer_content_pages(canvas, doc):
        """내용 페이지: 왼쪽 하단 로고 + 오른쪽 하단 페이지 번호"""
        canvas.saveState()
        # 로고 (왼쪽 하단) - 높이 17pt (24pt의 70% 축소)
        if os.path.exists(LOGO_PATH):
            from PIL import Image as PILImage
            pil_img = PILImage.open(LOGO_PATH)
            img_w, img_h = pil_img.size
            logo_h = 17
            logo_w = logo_h * (img_w / img_h)
            canvas.drawImage(LOGO_PATH, 10*mm, 4*mm, width=logo_w, height=logo_h, mask='auto')
        # 페이지 번호 (오른쪽 하단)
        page_num = doc.page  # 전체 페이지 번호
        canvas.setFont("Malgun", 8)
        canvas.setFillColor(colors.HexColor("#999999"))
        canvas.drawRightString(A4[0] - 10*mm, 8*mm, f"{page_num}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    # ── 제목 색상: #1E3A5F ──
    TITLE_COLOR = colors.HexColor("#1E3A5F")

    # Styles
    style_title = ParagraphStyle("Title", fontName="MalgunBd", fontSize=14, leading=18, alignment=1, spaceAfter=6, textColor=TITLE_COLOR)
    style_subtitle = ParagraphStyle("Subtitle", fontName="MalgunBd", fontSize=11, leading=14, spaceAfter=4, spaceBefore=10, keepWithNext=True, textColor=TITLE_COLOR)
    style_body = ParagraphStyle("Body", fontName="Malgun", fontSize=9, leading=12)
    style_body_small = ParagraphStyle("BodySmall", fontName="Malgun", fontSize=9, leading=12)
    style_header_p = ParagraphStyle("HeaderP", fontName="MalgunBd", fontSize=9, leading=12, textColor=colors.white, alignment=1)
    style_cell_center = ParagraphStyle("CellCenter", fontName="Malgun", fontSize=9, leading=12, alignment=1)
    style_cell_left = ParagraphStyle("CellLeft", fontName="Malgun", fontSize=9, leading=12)

    BLUE = colors.HexColor("#0E2841")  # RGB(14,40,65)
    LIGHT_BLUE = colors.HexColor("#D6E4F0")
    WHITE = colors.white

    # ── Admiral Slate 등급 컬러 체계 (이미지1 기반) ──
    GRADE_COLORS = {
        # ELITE (S급) - 짙은 남색 계열
        "S+": "#001F3F", "S": "#1E3A5F", "S-": "#3A5A8F",
        # HIGH (A급) - 틸/청록 계열
        "A+": "#008080", "A": "#20B2AA", "A-": "#40E0D0",
        # MID (B급) - 금색/노란 계열
        "B+": "#FFB800", "B": "#CC9600", "B-": "#B88A00",
        # LOW (C급) - 주황 계열
        "C+": "#FF6600", "C": "#FF8533", "C-": "#FFA366",
        # WARN (D급) - 빨간 계열
        "D+": "#E60000", "D": "#FF4D4D", "D-": "#FF8080",
    }

    # 텍스트용 (가독성 보장 - 밝은 색은 어두운 톤 사용)
    GRADE_TEXT_COLORS = {
        "S+": "#001F3F", "S": "#1E3A5F", "S-": "#3A5A8F",
        "A+": "#006666", "A": "#008080", "A-": "#2E8B8B",
        "B+": "#CC9600", "B": "#B88A00", "B-": "#A07800",
        "C+": "#E55C00", "C": "#CC6600", "C-": "#B35900",
        "D+": "#CC0000", "D": "#E60000", "D-": "#CC3333",
    }

    def get_grade_color(grade, for_text=False):
        """등급에 해당하는 색상 반환"""
        g = str(grade).strip()
        source = GRADE_TEXT_COLORS if for_text else GRADE_COLORS
        if g in source:
            return colors.HexColor(source[g])
        # +/- 없는 기본 등급으로 폴백
        base = g.rstrip("+-")
        if base in source:
            return colors.HexColor(source[base])
        return colors.black

    def grade_cell_color(grade):
        """Always return WHITE - no background coloring for grade cells."""
        return colors.white

    def grade_font_color(grade):
        """Return colored text for grade cells."""
        return get_grade_color(grade, for_text=True)

    def grade_bar_color(grade):
        """Return bar chart color for grade."""
        return get_grade_color(grade, for_text=False)

    def grade_paragraph(grade_text):
        """Create a bold, colored Paragraph for a grade value."""
        g = str(grade_text)
        c = get_grade_color(g, for_text=True)
        gs = ParagraphStyle("GradeStyle", fontName="MalgunBd", fontSize=9, leading=12, alignment=1, textColor=c)
        return Paragraph(g, gs)

    # ── 역량 텍스트 컬러링 ──
    COMP_COLORS = {"학업": "#006666", "진로": "#CC9600", "공동체": "#7B2D8E"}

    def colorize_competency(text):
        """핵심역량 텍스트에서 역량명에만 컬러를 적용한 HTML 반환.
        예: '학업역량(탐구력) -- 설명' → '<font color><b>학업역량(탐구력)</b></font><br/>설명'
        """
        import re
        t = str(text).replace("->", " \u2192 ")
        # " -- " 를 기준으로 역량명과 코멘트를 분리
        t = t.replace(" -- ", "<br/>")
        for keyword, color in COMP_COLORS.items():
            # 역량명(괄호 포함)만 매칭: "학업역량" 또는 "학업역량(탐구력)" 또는 "학업역량(협업·소통, 리더십)"
            pattern = f'({keyword}역량(?:\\([^)]*\\))?)'
            replacement = f'<font color="{color}"><b>\\1</b></font>'
            t = re.sub(pattern, replacement, t)
        return t

    def PComp(text):
        """핵심역량 칼럼용 Paragraph (역량별 컬러 적용)"""
        html = colorize_competency(text)
        s = ParagraphStyle("CompStyle", fontName="Malgun", fontSize=9, leading=12)
        return Paragraph(html, s)

    # ── 제목 번호 원형 배지 ──
    class NumberedTitle(Flowable):
        """숫자를 원형 배지 안에 넣고 제목 텍스트를 옆에 배치하는 Flowable.
        is_sub=True이면 세부 제목 (테두리만, 바탕색 없음)"""
        def __init__(self, number_text, title_text, is_sub=False, circle_r=9, font_size=11):
            Flowable.__init__(self)
            self.number = number_text
            self.title = title_text
            self.is_sub = is_sub
            self.r = circle_r
            self.fs = font_size
            self.width = 500
            self.height = circle_r * 2 + 4

        def draw(self):
            c = self.canv
            r = self.r
            cx, cy = r + 1, self.height / 2
            title_color = "#1E3A5F"

            if self.is_sub:
                # 세부 제목: 테두리만 (바탕색 없음)
                c.setStrokeColor(colors.HexColor(title_color))
                c.setLineWidth(1.2)
                c.setFillColor(colors.white)
                c.circle(cx, cy, r, fill=1, stroke=1)
                # 숫자 (#1E3A5F)
                c.setFillColor(colors.HexColor(title_color))
                c.setFont("MalgunBd", 7)
                tw = c.stringWidth(self.number, "MalgunBd", 7)
                c.drawString(cx - tw/2, cy - 3, self.number)
            else:
                # 메인 제목: 바탕색 채움
                c.setFillColor(colors.HexColor(title_color))
                c.circle(cx, cy, r, fill=1, stroke=0)
                # 숫자 (흰색)
                c.setFillColor(colors.white)
                c.setFont("MalgunBd", 8)
                tw = c.stringWidth(self.number, "MalgunBd", 8)
                c.drawString(cx - tw/2, cy - 3, self.number)

            # 제목 텍스트 (오른쪽, #1E3A5F)
            c.setFillColor(colors.HexColor(title_color))
            c.setFont("MalgunBd", self.fs)
            c.drawString(cx + r + 6, cy - self.fs/2 + 1, self.title)

    def section_title(text):
        """'1. 종합의견' → NumberedTitle 변환. '-' 포함 시 세부 제목"""
        import re
        m = re.match(r'^(\d[\d\-]*)\.\s*(.+)$', text.strip())
        if m:
            num = m.group(1)
            is_sub = '-' in num  # 2-1, 2-2, 3-1 등은 세부 제목
            return KeepTogether([Spacer(1, 6*mm), NumberedTitle(num, m.group(2), is_sub=is_sub), Spacer(1, 6*mm)])
        # ■ 로 시작하는 서브타이틀은 그대로
        return Paragraph(text, style_subtitle)

    def P(text, style=style_body):
        return Paragraph(str(text), style)

    def PH(text):
        return Paragraph(str(text), style_header_p)

    def PC(text):
        return Paragraph(str(text).replace("->", " \u2192 "), style_cell_center)

    def PL(text):
        return Paragraph(str(text).replace("->", " \u2192 "), style_cell_left)

    def make_table(data, col_widths=None, has_header=True, repeat_rows=None):
        if repeat_rows is not None:
            rr = repeat_rows
        else:
            rr = 1 if has_header else 0
        t = Table(data, colWidths=col_widths, repeatRows=rr)
        style_cmds = [
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ]
        if has_header:
            style_cmds.append(("BACKGROUND", (0,0), (-1,0), BLUE))
            style_cmds.append(("BACKGROUND", (0,1), (-1,-1), WHITE))
        t.setStyle(TableStyle(style_cmds))
        return t

    def apply_grade_styles(table, data, grade_col, start_row=1):
        """Apply grade coloring to table cells - white bg, colored text."""
        for r in range(start_row, len(data)):
            grade_val = data[r][grade_col]
            # Extract text from Paragraph if needed
            if hasattr(grade_val, 'text'):
                g = grade_val.text
            else:
                g = str(grade_val)
            if g in GRADE_COLORS:
                table.setStyle(TableStyle([
                    ("BACKGROUND", (grade_col, r), (grade_col, r), colors.white),
                    ("TEXTCOLOR", (grade_col, r), (grade_col, r), grade_font_color(g)),
                ]))

    elements = []
    page_w = A4[0] - 20*mm  # 좌우 10mm씩

    # ═══════════════════════════════════════
    # 표지 페이지
    # ═══════════════════════════════════════
    class CoverPage(Flowable):
        """표지 페이지 전체를 그리는 Flowable"""
        def __init__(self, logo_path, student, school, record_range="1~3학년"):
            Flowable.__init__(self)
            self.logo_path = logo_path
            self.student = student
            self.school = school
            self.record_range = record_range  # 예: "1학년 1학기", "1~3학년"
            self.width = 520   # 프레임 내부에 맞춤 (526보다 작게)
            self.height = 755  # 프레임 내부에 맞춤 (761보다 작게)

        def draw(self):
            c = self.canv
            w = self.width
            h = self.height

            # 왼쪽 정렬 기준선 (모든 요소 통일)
            left_x = 0

            # 로고 (왼쪽 상단) - 실제 이미지 비율 사용
            if os.path.exists(self.logo_path):
                from PIL import Image as PILImage
                pil_img = PILImage.open(self.logo_path)
                img_w, img_h = pil_img.size
                logo_h = 54  # 90 → 54 (60% 축소)
                logo_w = logo_h * (img_w / img_h)
                c.drawImage(self.logo_path, left_x, h - logo_h - 10, width=logo_w, height=logo_h, mask='auto')

            # 메인 타이틀: "학생부라운지"
            c.setFillColor(colors.HexColor("#1E3A5F"))
            c.setFont("MalgunBd", 43)  # 36 → 43 (120% 확대)
            c.drawString(left_x, h - 220, "학생부라운지")

            # 하단 정보 박스
            box_y = 30
            box_h = 70
            c.setStrokeColor(colors.HexColor("#DDDDDD"))
            c.setLineWidth(1)
            c.setFillColor(colors.HexColor("#FAFAFA"))
            c.roundRect(left_x, box_y, w - left_x, box_h, 5, fill=1, stroke=1)

            # 학생 이름 (굵게) + 학교명
            c.setFillColor(colors.HexColor("#333333"))
            c.setFont("MalgunBd", 12)
            c.drawString(left_x + 15, box_y + box_h - 22, f"{self.student}")
            c.setFont("Malgun", 12)
            name_w = c.stringWidth(self.student, "MalgunBd", 12)
            c.drawString(left_x + 15 + name_w + 8, box_y + box_h - 22, f"{self.school}")

            # 학교생활기록부 기록 학년 (데이터로부터 자동 생성)
            c.setFont("Malgun", 9)
            c.setFillColor(colors.HexColor("#888888"))
            c.drawString(15, box_y + box_h - 42, f"학교생활기록부 기록: {self.record_range}")

            # 분석일
            c.drawString(15, box_y + box_h - 56, f"분석일: {date.today().strftime('%Y-%m-%d')}")

    # 학년 범위 자동 계산 (setuek_data + changche_data + haengtuk_data 모두 고려)
    def _calc_record_range():
        all_years = set()
        for d in setuek_data:
            all_years.add(d[0])
        for d in changche_data:
            all_years.add(d[0])
        for d in haengtuk_data:
            all_years.add(d[0])
        if not all_years:
            return "-"
        years_sorted = sorted(all_years)
        mn, mx = years_sorted[0], years_sorted[-1]
        if mn == mx:
            return f"{mn}학년"
        return f"{mn}~{mx}학년"

    record_range = _calc_record_range()
    elements.append(CoverPage(LOGO_PATH, STUDENT, SCHOOL, record_range))
    elements.append(PageBreak())

    # ═══════════════════════════════════════
    # 목차 페이지
    # ═══════════════════════════════════════
    indent = 20  # 2칸 들여쓰기
    style_toc_title = ParagraphStyle("TocTitle", fontName="MalgunBd", fontSize=22, leading=28, textColor=TITLE_COLOR, spaceAfter=8, leftIndent=indent)
    style_toc_main = ParagraphStyle("TocMain", fontName="MalgunBd", fontSize=20, leading=26, spaceBefore=14, textColor=colors.HexColor("#333333"), leftIndent=indent)
    style_toc_sub = ParagraphStyle("TocSub", fontName="Malgun", fontSize=15, leading=24, textColor=colors.HexColor("#666666"), leftIndent=indent + 25)

    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph("목차", style_toc_title))
    elements.append(Spacer(1, 10*mm))

    toc_items = [
        (False, "1. 종합의견"),
        (False, "2. 세부능력 및 특기사항"),
        (True, "2-1. 과목 별 분석"),
        (True, "2-2. 과목 별 세부내용"),
        (True, "2-3. 핵심평가문장"),
        (False, "3. 창의적 체험활동"),
        (True, "3-1. 자율활동"),
        (True, "3-2. 동아리활동"),
        (True, "3-3. 진로활동"),
        (False, "4. 행동특성 및 종합의견"),
        (False, "5. 연계성분석"),
        (False, "6. 대학평가요소"),
        (False, "7. 역량별보완법"),
    ]
    for is_sub, item in toc_items:
        if is_sub:
            elements.append(Paragraph(item, style_toc_sub))
        else:
            elements.append(Paragraph(item, style_toc_main))

    elements.append(PageBreak())

    # ═══════════════════════════════════════
    # 평가 기준 안내 페이지
    # ═══════════════════════════════════════
    style_guide_title = ParagraphStyle("GuideTitle", fontName="MalgunBd", fontSize=14, leading=18, textColor=TITLE_COLOR, spaceAfter=6)
    style_guide_h2 = ParagraphStyle("GuideH2", fontName="MalgunBd", fontSize=11, leading=15, textColor=TITLE_COLOR, spaceBefore=8, spaceAfter=4)
    style_guide_body = ParagraphStyle("GuideBody", fontName="Malgun", fontSize=9, leading=13, spaceAfter=2)
    style_guide_item = ParagraphStyle("GuideItem", fontName="Malgun", fontSize=9, leading=13, leftIndent=10, spaceAfter=1)

    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph("평가 기준 안내", style_guide_title))
    elements.append(Spacer(1, 4*mm))

    # 1) 세특 평가 기준
    elements.append(Paragraph("1) 세부능력 및 특기사항 평가 항목", style_guide_h2))
    guide_setuek = [
        ["항목", "설명"],
        ["교과연계성", "해당 과목의 핵심 개념·단원에서 탐구가 시작되었는가"],
        ["탐구동기", "왜 이 주제에 관심을 가졌는지 구체적으로 드러나는가"],
        ["탐구과정", "어떤 방법으로, 어떤 자료를 활용해 탐구했는가"],
        ["결과성찰", "탐구를 통해 무엇을 알게 되었고, 어떤 변화가 있었는가"],
        ["차별성", "다른 학생 대비 독창적이고 심화된 수준인가"],
        ["학업태도", "수업 참여도, 질문/발표 적극성, 자기주도 학습 의지"],
    ]
    tg = Table([[Paragraph(c, style_header_p if r == 0 else style_guide_body) for c in row]
                for r, row in enumerate(guide_setuek)],
               colWidths=[80, page_w - 80])
    tg.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), BLUE),
        ("BACKGROUND", (0,1), (-1,-1), colors.white),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(tg)
    elements.append(Spacer(1, 4*mm))

    # 2) 창체 평가 기준
    elements.append(Paragraph("2) 창의적 체험활동 평가 항목", style_guide_h2))
    guide_changche = [
        ["영역", "평가 항목"],
        ["자율활동", "참여주도성, 활동구체성, 활동완결성, 성장변화, 전공적합성"],
        ["동아리활동", "전문성, 개인기여도, 활동완결성, 지속성발전, 전공적합성"],
        ["진로활동", "탐색구체성, 목표일관성, 활동완결성, 자기주도성, 전공적합성"],
    ]
    tg2 = Table([[Paragraph(c, style_header_p if r == 0 else style_guide_body) for c in row]
                 for r, row in enumerate(guide_changche)],
                colWidths=[80, page_w - 80])
    tg2.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), BLUE),
        ("BACKGROUND", (0,1), (-1,-1), colors.white),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(tg2)
    elements.append(Spacer(1, 4*mm))

    # 3) 행특 평가 기준
    elements.append(Paragraph("3) 행동특성 및 종합의견 평가 항목", style_guide_h2))
    guide_haengtuk = [
        ["항목", "설명"],
        ["활동구체성", "특정 상황·행동·결과가 에피소드로 서술되었는가"],
        ["동기-과정-결과", "세 요소가 구체적으로 연결되고 성찰이 포함되었는가"],
        ["성장변화", "1→2→3학년에 걸친 역량 심화·확장이 드러나는가"],
        ["인성·공동체", "배려, 협력, 리더십이 구체적 에피소드로 서술되었는가"],
        ["분량·밀도", "충분한 분량에 다양한 활동이 구체적으로 기술되었는가"],
    ]
    tg3 = Table([[Paragraph(c, style_header_p if r == 0 else style_guide_body) for c in row]
                 for r, row in enumerate(guide_haengtuk)],
                colWidths=[80, page_w - 80])
    tg3.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), BLUE),
        ("BACKGROUND", (0,1), (-1,-1), colors.white),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(tg3)
    elements.append(Spacer(1, 4*mm))

    # 4) 등급 기준
    elements.append(Paragraph("4) 등급 기준 (1~10점 척도)", style_guide_h2))
    guide_grades = [
        ["등급", "점수 범위", "의미"],
        ["S", "8.50 ~ 10.00", "최상위권 학종 경쟁력"],
        ["A", "7.00 ~ 8.49", "상위권 경쟁력"],
        ["B", "5.00 ~ 6.99", "평균 수준"],
        ["C", "3.50 ~ 4.99", "보완 필요"],
        ["D", "1.00 ~ 3.49", "경쟁력 재고 필요"],
    ]
    tg4 = Table([[Paragraph(c, style_header_p if r == 0 else style_guide_body) for c in row]
                 for r, row in enumerate(guide_grades)],
                colWidths=[40, 80, page_w - 120])
    tg4.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), BLUE),
        ("BACKGROUND", (0,1), (-1,-1), colors.white),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("ALIGN", (0,1), (1,-1), "CENTER"),
    ]))
    # 등급별 텍스트 컬러 적용
    for r in range(1, len(guide_grades)):
        g = guide_grades[r][0]
        tg4.setStyle(TableStyle([
            ("TEXTCOLOR", (0, r), (0, r), grade_font_color(g)),
        ]))
    elements.append(tg4)
    elements.append(Spacer(1, 4*mm))

    # 5) 핵심역량 분류
    elements.append(Paragraph("5) 핵심역량 분류", style_guide_h2))
    guide_comp = [
        ["역량", "색상", "세부 항목"],
        ["학업역량", "틸(청록)", "학업성취도, 학업태도, 탐구력"],
        ["진로역량", "금색", "교과이수 노력, 교과성취도, 진로탐색"],
        ["공동체역량", "보라", "협업·소통, 나눔·배려, 성실성·규칙준수, 리더십"],
    ]
    tg5 = Table([[Paragraph(c, style_header_p if r == 0 else style_guide_body) for c in row]
                 for r, row in enumerate(guide_comp)],
                colWidths=[70, 60, page_w - 130])
    tg5.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), BLUE),
        ("BACKGROUND", (0,1), (-1,-1), colors.white),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    # 역량별 색상 적용
    tg5.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor("#006666")),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor("#006666")),
        ("TEXTCOLOR", (0, 2), (0, 2), colors.HexColor("#CC9600")),
        ("TEXTCOLOR", (1, 2), (1, 2), colors.HexColor("#CC9600")),
        ("TEXTCOLOR", (0, 3), (0, 3), colors.HexColor("#7B2D8E")),
        ("TEXTCOLOR", (1, 3), (1, 3), colors.HexColor("#7B2D8E")),
    ]))
    elements.append(tg5)

    elements.append(PageBreak())

    # ═══════════════════════════════════════
    # 분석 내용 시작
    # ═══════════════════════════════════════
    # ── Title ──
    elements.append(P(f"{STUDENT} ({SCHOOL}) 학생부 분석 리포트", style_title))
    elements.append(Spacer(1, 2*mm))

    # ── 1. 종합의견 ──
    elements.append(section_title("1. 종합의견"))
    tdata = [[PH("항목"), PH("내용")]]
    for k, v in [
        ("학생명", summary_data["학생"]),
        ("학교명", summary_data["학교명"]),
        ("비교과 종합", summary_data["비교과종합"]),
        ("비교과 종합점수", summary_data["비교과종합점수"]),
        ("핵심강점1", summary_data["핵심강점1"]),
        ("핵심강점2", summary_data["핵심강점2"]),
        ("핵심강점3", summary_data["핵심강점3"]),
        ("보완영역1", summary_data["보완영역1"]),
        ("보완영역2", summary_data["보완영역2"]),
        ("보완영역3", summary_data["보완영역3"]),
        ("성장스토리", summary_data["성장스토리"]),
    ]:
        tdata.append([PC(k), PL(v)])
    t = make_table(tdata, col_widths=[80, page_w - 80])
    # Change 6: Add minimum row padding for 2-line fitting
    t.setStyle(TableStyle([
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 3*mm))

    # ── Change 7: 비교과 종합 Bar Graph (2x2 grid) ──
    # 학생 데이터에서 자동 계산 (세특 평점 평균, 창체/행특 환산점수 → 100점 스케일)
    def _to_100(v):
        """1~10점 척도를 0~100 스케일로 변환"""
        return v * 10

    # 세특 평점 평균 (이미 1~10 스케일). 튜플의 가중합산은 score_end 위치.
    _pdf_score_end = setuek_score_slice_end(sd)
    setuek_avg = sum(d[_pdf_score_end] for d in setuek_data) / len(setuek_data) if setuek_data else 0
    setuek_100 = _to_100(setuek_avg)

    # 창체 환산점수 평균 (이미 100 스케일)
    changche_avg = sum(d[5] for d in changche_data) / len(changche_data) if changche_data else 0

    # 행특 환산점수 평균 (이미 100 스케일)
    haengtuk_avg = sum(d[3] for d in haengtuk_data) / len(haengtuk_data) if haengtuk_data else 0

    # 비교과 종합 = 가중 평균 (세특:창체:행특 = 30:25:10 → 정규화)
    comp_total = (setuek_100 * 0.30 + changche_avg * 0.25 + haengtuk_avg * 0.10) / 0.65

    def _grade_100(v):
        """100점 척도를 등급으로 변환 (1~10 변환 후 score_to_grade)"""
        g = score_to_grade(v / 10)
        # +/- 세분화
        v10 = v / 10
        if g == "S":
            if v10 >= 9.5: return "S+"
            if v10 < 8.7: return "S-"
            return "S"
        if g == "A":
            if v10 >= 8.0: return "A+"
            if v10 < 7.3: return "A-"
            return "A"
        if g == "B":
            if v10 >= 6.3: return "B+"
            if v10 < 5.7: return "B-"
            return "B"
        if g == "C":
            if v10 >= 4.5: return "C+"
            if v10 < 4.0: return "C-"
            return "C"
        return "D"

    bar_items = [
        ("세부능력 및 특기사항", round(setuek_100, 2), _grade_100(setuek_100)),
        ("창의적 체험활동", round(changche_avg, 2), _grade_100(changche_avg)),
        ("행동특성 및 종합의견", round(haengtuk_avg, 2), _grade_100(haengtuk_avg)),
        ("비교과 종합", round(comp_total, 2), _grade_100(comp_total)),
    ]

    cell_w = page_w / 2 - 4  # 한 칸 너비
    cell_h = 42              # 한 칸 높이 (라벨+막대)
    draw_h = cell_h * 2 + 8  # 2행 + 간격
    d = Drawing(page_w, draw_h)

    for idx, (name, score, grade) in enumerate(bar_items):
        col = idx % 2
        row = idx // 2
        ox = col * (cell_w + 8)
        oy = draw_h - (row + 1) * cell_h - row * 4

        # 등급별 컬러 (Admiral Slate 기반)
        bar_color = grade_bar_color(grade)
        text_color = get_grade_color(grade, for_text=True)

        # 라벨 (상단)
        d.add(String(ox + 4, oy + cell_h - 12, name, fontName="MalgunBd", fontSize=8, fillColor=colors.HexColor("#333333")))

        # 막대 배경 (100% 회색)
        bar_y = oy + 4
        bar_max_w = cell_w - 70
        bar_h = 16
        d.add(Rect(ox + 4, bar_y, bar_max_w, bar_h, fillColor=colors.HexColor("#EDEDED"), strokeColor=None))

        # 막대 (등급별 컬러)
        bar_w = (score / 100.0) * bar_max_w
        d.add(Rect(ox + 4, bar_y, bar_w, bar_h, fillColor=bar_color, strokeColor=None))

        # 점수 + 등급 (등급 컬러 텍스트)
        d.add(String(ox + 4 + bar_max_w + 6, bar_y + bar_h / 2 - 4, f"{score:.2f}  {grade}",
                      fontName="MalgunBd", fontSize=9, fillColor=text_color))

    elements.append(d)
    elements.append(Spacer(1, 2*mm))

    # ── Change 8: 종합 코멘트 styled box ──
    comment_text = summary_data["종합코멘트"].replace("\n\n", "<br/><br/>").replace("->", " \u2192 ")
    style_comment_title = ParagraphStyle("CommentTitle", fontName="MalgunBd", fontSize=10, leading=14, textColor=colors.HexColor("#0E2841"), spaceAfter=4)
    style_comment_body = ParagraphStyle("CommentBody", fontName="Malgun", fontSize=9, leading=14)
    comment_title_p = Paragraph("종합 코멘트", style_comment_title)
    comment_body_p = Paragraph(comment_text, style_comment_body)
    comment_tdata = [[comment_title_p], [comment_body_p]]
    comment_table = Table(comment_tdata, colWidths=[page_w - 14])
    comment_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8F9FA")),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LINEBEFORESTARTOFFSET", (0, 0), (-1, -1), 0),
        ("LINEBEFORE", (0, 0), (0, -1), 4, colors.HexColor("#0E2841")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
    ]))
    elements.append(comment_table)
    elements.append(PageBreak())

    # ── 섹션 번호 동적 할당 ──
    # 1. 종합의견 (고정)
    # 2. 이전 분석 대비 변화 (선택, compare_data)     ← G3+G4
    # 3. 출결 및 봉사활동 (선택, attendance/volunteer) ← G7
    # 이후 세특/창체/행특/연계성/대학평가요소/역량별보완법/키워드분석 섹션 번호 동적 rename
    _has_compare_section = False
    _has_attendance_section = False
    _has_grading_section = False
    try:
        from .compare_generator import has_compare_data
        _has_compare_section = has_compare_data(sd)
    except Exception as e:
        print(f"[WARN] compare 판정 예외: {type(e).__name__}: {e}")
    try:
        from .attendance_calculator import has_attendance_or_volunteer
        _has_attendance_section = has_attendance_or_volunteer(sd)
    except Exception as e:
        print(f"[WARN] attendance 판정 예외: {type(e).__name__}: {e}")
    try:
        _has_grading_section = _has_grade_data(sd)
    except Exception as e:
        print(f"[WARN] grading 판정 예외: {type(e).__name__}: {e}")

    _n = 1  # 1. 종합의견
    _sec_compare = _sec_attendance = _sec_grading = None
    _sec_setuek = _sec_changche = _sec_haengtuk = None
    if _has_compare_section:
        _n += 1; _sec_compare = _n
    if _has_attendance_section:
        _n += 1; _sec_attendance = _n
    if _has_grading_section:
        _n += 1; _sec_grading = _n
    # 2026-04-19: mode_config 에 따라 세특/창체/행특 섹션 스킵
    if inc_setuek:
        _n += 1; _sec_setuek   = _n
    if inc_changche:
        _n += 1; _sec_changche = _n
    if inc_haengtuk:
        _n += 1; _sec_haengtuk = _n
    _n += 1; _sec_linkage  = _n
    _n += 1; _sec_eval     = _n
    _n += 1; _sec_fix      = _n
    _n += 1; _sec_keyword  = _n

    # ── 이전 분석 대비 변화 섹션 (compare_data 있을 때) ──
    if _has_compare_section:
        try:
            _append_pdf_compare_section(elements, sd, section_title,
                                         P, PC, PL, PH, make_table,
                                         style_subtitle, page_w,
                                         section_number=_sec_compare)
            elements.append(PageBreak())
        except Exception as e:
            print(f"[WARN] PDF 이전분석대비변화 섹션 생성 실패 (스킵): {type(e).__name__}: {e}")

    # ── 출결 및 봉사활동 섹션 (attendance/volunteer 있을 때) ──
    if _has_attendance_section:
        try:
            _append_pdf_attendance_section(elements, sd, section_title,
                                            P, PC, PL, PH, make_table,
                                            style_subtitle, page_w,
                                            section_number=_sec_attendance)
            elements.append(PageBreak())
        except Exception as e:
            print(f"[WARN] PDF 출결·봉사 섹션 생성 실패 (스킵): {type(e).__name__}: {e}")

    # ── 대학별 내신 섹션 (grade_data 있을 때, 2026-05-05) ──
    if _has_grading_section:
        try:
            _append_pdf_grading_section(elements, sd, section_title,
                                         P, PC, PL, PH, make_table,
                                         style_subtitle, page_w,
                                         section_number=_sec_grading)
            elements.append(PageBreak())
        except Exception as e:
            print(f"[WARN] PDF 대학별내신 섹션 생성 실패 (스킵): {type(e).__name__}: {e}")

    # ── 2. 세부능력 및 특기사항 (mode_config 로 스킵 가능, 2026-04-19) ──
    # setuek 미포함 시 elements 임시 리다이렉트 (코드 재들여쓰기 회피).
    _setuek_saved_elements = elements
    if not inc_setuek:
        elements = []  # 모든 세특 append 를 버리는 더미 리스트
    elements.append(section_title(f"{_sec_setuek}. 세부능력 및 특기사항"))
    elements.append(Spacer(1, 2*mm))
    elements.append(section_title(f"{_sec_setuek}-1. 과목 별 분석"))

    # ── 교과 중심 분석 테이블 ──
    subject_to_category = {
        "국어": "국어", "독서(국어)": "국어",
        "수학": "수학", "수학II": "수학", "고급수학I": "수학",
        "영어": "영어",
        "한국사": "사회", "통합사회": "사회", "사회탐구방법론": "사회",
        "통합과학": "과학", "과학탐구실험": "과학", "물리학I": "과학",
        "생명과학I": "과학", "화학I": "과학", "생명과학실험": "과학",
        "생명과학심화": "과학", "화학심화": "과학", "생명과학실험심화": "과학",
    }
    category_order = ["국어", "수학", "영어", "사회", "과학"]

    cat_rows = []
    for cat in category_order:
        cat_subjects = [d for d in setuek_data if subject_to_category.get(d[1], "기타") == cat]
        if not cat_subjects:
            continue
        subj_names = ", ".join([d[1] for d in cat_subjects])
        n = len(cat_subjects)
        avgs = [round(sum(d[i] for d in cat_subjects) / n, 1) for i in range(2, _pdf_score_end)]
        weighted_avg = round(sum(d[_pdf_score_end] for d in cat_subjects) / n, 2)
        grade = score_to_grade(weighted_avg)
        cat_rows.append((cat, subj_names, avgs, weighted_avg, grade))

    style_ref_note = ParagraphStyle("RefNote", fontName="Malgun", fontSize=8, leading=11, textColor=colors.HexColor("#999999"), spaceBefore=2, spaceAfter=4)
    elements.append(Paragraph("※ 각 항목은 1~10점 척도로 평가됩니다. 상세 평가 기준은 '평가 기준 안내' 페이지를 참고하세요.", style_ref_note))

    # ── 세특 PDF 테이블 헤더/SPAN 스펙 빌더 (지정/미지정 모드 공용) ──
    _major = is_major_mode(sd)

    def _setuek_pdf_spec(col1_label, col1_w, col2_label, col2_w):
        """세특 PDF 테이블의 2-row 헤더, 컬럼폭, SPAN 리스트, 등급 컬럼 인덱스 생성.
        지정 모드(7항목)면 '전공적합성' 컬럼이 교과연계성 뒤에 추가됨.
        """
        if _major:
            # 11열: col1 | col2 | 교과연계 | 전공적합 | 탐구활동(3) | 차별성 | 학업태도 | 평점 | 등급
            hdr0 = [PH(col1_label), PH(col2_label), PH("교과\n연계성"), PH("전공\n적합성"),
                    PH("탐구활동"), PH(""), PH(""),
                    PH("차별성"), PH("학업태도"), PH("평점"), PH("등급")]
            hdr1 = [PH("")] * 4 + [PH("동기"), PH("과정"), PH("결과")] + [PH("")] * 4
            n_dyn = 7  # 교과연계+전공적합+탐구3+차별+학업
            fixed_w = col1_w + col2_w + 50 + 34
            each_w = (page_w - fixed_w) / n_dyn
            cw_out = [col1_w, col2_w] + [each_w] * n_dyn + [50, 34]
            spans = [
                ("SPAN", (4, 0), (6, 0)),   # 탐구활동 (3열)
                ("SPAN", (0, 0), (0, 1)),   # col1
                ("SPAN", (1, 0), (1, 1)),   # col2
                ("SPAN", (2, 0), (2, 1)),   # 교과연계성
                ("SPAN", (3, 0), (3, 1)),   # 전공적합성
                ("SPAN", (7, 0), (7, 1)),   # 차별성
                ("SPAN", (8, 0), (8, 1)),   # 학업태도
                ("SPAN", (9, 0), (9, 1)),   # 평점
                ("SPAN", (10, 0), (10, 1)), # 등급
            ]
            grade_col = 10
        else:
            # 10열 (기존): col1 | col2 | 교과연계 | 탐구활동(3) | 차별성 | 학업태도 | 평점 | 등급
            hdr0 = [PH(col1_label), PH(col2_label), PH("교과\n연계성"),
                    PH("탐구활동"), PH(""), PH(""),
                    PH("차별성"), PH("학업태도"), PH("평점"), PH("등급")]
            hdr1 = [PH("")] * 3 + [PH("동기"), PH("과정"), PH("결과")] + [PH("")] * 4
            n_dyn = 6
            fixed_w = col1_w + col2_w + 50 + 34
            each_w = (page_w - fixed_w) / n_dyn
            cw_out = [col1_w, col2_w] + [each_w] * n_dyn + [50, 34]
            spans = [
                ("SPAN", (3, 0), (5, 0)),
                ("SPAN", (0, 0), (0, 1)),
                ("SPAN", (1, 0), (1, 1)),
                ("SPAN", (2, 0), (2, 1)),
                ("SPAN", (6, 0), (6, 1)),
                ("SPAN", (7, 0), (7, 1)),
                ("SPAN", (8, 0), (8, 1)),
                ("SPAN", (9, 0), (9, 1)),
            ]
            grade_col = 9
        return hdr0, hdr1, cw_out, spans, grade_col

    elements.append(P("■ 교과 중심 분석", style_subtitle))
    hdr0, hdr1, cw_cat, spans_cat, grade_col_cat = _setuek_pdf_spec("교과", 40, "이수과목", 120)
    tdata_cat = [hdr0, hdr1]
    for cat, subj_names, avgs, wavg, grade in cat_rows:
        row = [PC(cat), PL(subj_names)]
        row += [PC(str(a)) for a in avgs]
        row += [PC(f"{wavg:.2f}"), grade_paragraph(grade)]
        tdata_cat.append(row)
    t_cat = make_table(tdata_cat, col_widths=cw_cat, has_header=False, repeat_rows=2)
    # Apply 2-row header styling
    t_cat.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 1), BLUE),
        *spans_cat,
        ("BACKGROUND", (0, 2), (-1, -1), WHITE),
    ]))
    for r in range(2, len(tdata_cat)):
        g = cat_rows[r-2][4]
        t_cat.setStyle(TableStyle([
            ("BACKGROUND", (grade_col_cat, r), (grade_col_cat, r), colors.white),
            ("TEXTCOLOR", (grade_col_cat, r), (grade_col_cat, r), grade_font_color(g)),
        ]))
    elements.append(t_cat)
    elements.append(Spacer(1, 6*mm))

    # ── 학년별/과목별 상세 분석 테이블 (학년 셀병합) ──
    elements.append(P("■ 학년별 · 과목별 분석", style_subtitle))
    hdr0, hdr1, cw, spans_yr, grade_col_yr = _setuek_pdf_spec("학년", 42, "과목", 90)
    tdata = [hdr0, hdr1]

    # 학년별로 그룹핑 (데이터에 존재하는 학년만)
    setuek_years_in_data = sorted(set(d[0] for d in setuek_data))
    for yr in setuek_years_in_data:
        yr_subjects = [d for d in setuek_data if d[0] == yr]
        for idx, d in enumerate(yr_subjects):
            yr_label = f"{yr}학년" if idx == 0 else ""
            scores_pd = d[2:_pdf_score_end]
            wavg_pd = d[_pdf_score_end]
            grade_pd = d[_pdf_score_end + 1]
            row = [PC(yr_label), PL(d[1])] + [PC(str(s)) for s in scores_pd] + [PC(f"{wavg_pd:.2f}"), grade_paragraph(grade_pd)]
            tdata.append(row)

    t = make_table(tdata, col_widths=cw, has_header=False, repeat_rows=2)

    # Apply 2-row header styling
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 1), BLUE),
        *spans_yr,
        ("BACKGROUND", (0, 2), (-1, -1), WHITE),
    ]))

    # 학년 셀병합 (SPAN) - offset by 2 for 2 header rows
    row_idx = 2  # 2 header rows
    for yr in setuek_years_in_data:
        yr_count = len([d for d in setuek_data if d[0] == yr])
        if yr_count > 1:
            t.setStyle(TableStyle([
                ("SPAN", (0, row_idx), (0, row_idx + yr_count - 1)),
                ("VALIGN", (0, row_idx), (0, row_idx + yr_count - 1), "MIDDLE"),
            ]))
        row_idx += yr_count

    # Grade coloring - offset by 2 for 2 header rows
    for r in range(2, len(tdata)):
        g = setuek_data[r-2][_pdf_score_end + 1]
        t.setStyle(TableStyle([
            ("BACKGROUND", (grade_col_yr, r), (grade_col_yr, r), colors.white),
            ("TEXTCOLOR", (grade_col_yr, r), (grade_col_yr, r), grade_font_color(g)),
        ]))
    elements.append(t)
    elements.append(PageBreak())

    # ── 2-2. 세부능력 및 특기사항 과목 별 세부내용 ──
    elements.append(section_title(f"{_sec_setuek}-2. 과목 별 세부내용"))
    # 과목명 칸 넓이: 9pt 맑은고딕 기준 한글 1자 ~9pt, 영문/숫자/괄호 ~5pt
    max_subj_w = 0
    for d in setuek_data:
        w = sum(9 if ord(ch) > 127 else 5 for ch in d[1])
        max_subj_w = max(max_subj_w, w)
    subj_col_w = max(80, max_subj_w + 16)
    cw = [42, subj_col_w, (page_w - 42 - subj_col_w)/2, (page_w - 42 - subj_col_w)/2]

    # comment_keys를 학년별로 그룹핑
    from collections import OrderedDict
    year_groups = OrderedDict()
    for key in comment_keys:
        parts = key.rsplit("(", 1)
        subj_name = parts[0]
        yr_num = parts[1].rstrip(")") if len(parts) > 1 else ""
        year_groups.setdefault(yr_num, []).append((subj_name, setuek_comments[key]))

    # 학년별로 별도 테이블 생성 (SPAN 없음 → 페이지 경계 분할 가능)
    for yr_num, items in year_groups.items():
        # 학년 라벨을 소제목으로 표시
        elements.append(P(f"■ {yr_num}학년", style_subtitle))
        hdr = [PH("과목"), PH("강점"), PH("보완점")]
        tdata = [hdr]
        cw_yr = [subj_col_w, (page_w - subj_col_w) / 2, (page_w - subj_col_w) / 2]
        for subj_name, (strength, weakness) in items:
            tdata.append([PL(subj_name), PL(strength), PL(weakness)])

        t = make_table(tdata, col_widths=cw_yr)
        elements.append(t)
        elements.append(Spacer(1, 3*mm))

    elements.append(PageBreak())

    # ── 2-3. 세부능력 및 특기사항 핵심평가문장 ──
    elements.append(section_title(f"{_sec_setuek}-3. 핵심평가문장"))
    # Change 7: 표현역량 -> 핵심역량
    hdr = [PH("과목"), PH("핵심문장"), PH("이유"), PH("핵심역량")]
    tdata = [hdr]
    # 과목 column: 가장 긴 과목명(학년 포함) 기준 동적 계산
    max_gs_w = 0
    for subj, sent, reason, comp in good_sentences:
        w = sum(9 if ord(ch) > 127 else 5 for ch in subj)
        max_gs_w = max(max_gs_w, w)
    gs_subj_w = max(80, max_gs_w + 16)
    cw = [gs_subj_w, page_w*0.35, page_w*0.28, page_w - gs_subj_w - page_w*0.35 - page_w*0.28]
    for subj, sent, reason, comp in good_sentences:
        # 핵심역량 칼럼: 역량별 컬러 적용
        tdata.append([PL(subj), PL(sent), PL(reason), PComp(comp)])
    t = make_table(tdata, col_widths=cw)

    # Change 5: Merge same subject cells using SPAN
    prev_subj = None
    span_start = None
    span_ranges = []
    for r in range(1, len(tdata)):
        subj_val = good_sentences[r-1][0]
        if subj_val == prev_subj:
            # Continue the span
            pass
        else:
            if prev_subj is not None and span_start is not None and r - 1 > span_start:
                span_ranges.append((span_start, r - 1))
            span_start = r
            prev_subj = subj_val
    # Handle last group
    if prev_subj is not None and span_start is not None and len(tdata) - 1 > span_start:
        span_ranges.append((span_start, len(tdata) - 1))

    for start_r, end_r in span_ranges:
        t.setStyle(TableStyle([
            ("SPAN", (0, start_r), (0, end_r)),
            ("VALIGN", (0, start_r), (0, end_r), "MIDDLE"),
        ]))

    elements.append(t)
    elements.append(Spacer(1, 6*mm))

    # 역량 정리표 - competency summary table
    elements.append(P("■ 핵심문장 역량 매핑", style_subtitle))
    style_dot_academic = ParagraphStyle("DotA", fontName="MalgunBd", fontSize=10, leading=12, alignment=1, textColor=colors.HexColor("#006666"))
    style_dot_career = ParagraphStyle("DotC", fontName="MalgunBd", fontSize=10, leading=12, alignment=1, textColor=colors.HexColor("#CC9600"))
    style_dot_community = ParagraphStyle("DotCo", fontName="MalgunBd", fontSize=10, leading=12, alignment=1, textColor=colors.HexColor("#7B2D8E"))
    hdr_comp = [PH("과목"), PH("핵심문장(요약)"), PH("학업역량"), PH("진로역량"), PH("공동체역량")]
    tdata_comp = [hdr_comp]
    cw_comp = [gs_subj_w, page_w - gs_subj_w - 60 - 60 - 60, 60, 60, 60]
    comp_subj_list = []  # 셀병합용
    for subj, sent, reason, comp in good_sentences:
        if sent == "해당 없음":
            continue
        sent_short = sent[:20] + "..." if len(sent) > 20 else sent
        comp_str = str(comp)
        has_academic = "학업" in comp_str
        has_career = "진로" in comp_str
        has_community = "공동체" in comp_str
        dot_a = Paragraph("\u25cf", style_dot_academic) if has_academic else PC("")
        dot_c = Paragraph("\u25cf", style_dot_career) if has_career else PC("")
        dot_co = Paragraph("\u25cf", style_dot_community) if has_community else PC("")
        tdata_comp.append([PL(subj), PL(sent_short), dot_a, dot_c, dot_co])
        comp_subj_list.append(subj)
    t_comp = make_table(tdata_comp, col_widths=cw_comp)

    # 같은 과목 셀병합
    prev_subj = None
    span_start = None
    span_ranges = []
    for r, subj in enumerate(comp_subj_list):
        row = r + 1  # 헤더 offset
        if subj == prev_subj:
            pass
        else:
            if prev_subj is not None and span_start is not None and row - 1 > span_start:
                span_ranges.append((span_start, row - 1))
            span_start = row
            prev_subj = subj
    if prev_subj is not None and span_start is not None and len(comp_subj_list) > span_start:
        span_ranges.append((span_start, len(comp_subj_list)))
    for sr, er in span_ranges:
        t_comp.setStyle(TableStyle([
            ("SPAN", (0, sr), (0, er)),
            ("VALIGN", (0, sr), (0, er), "MIDDLE"),
        ]))

    elements.append(t_comp)
    elements.append(PageBreak())
    # ── 세특 블록 종료: elements 복원 (2026-04-19) ──
    elements = _setuek_saved_elements

    # ── 3. 창의적 체험활동 (mode_config 로 스킵 가능) ──
    if inc_changche:
        elements.append(section_title(f"{_sec_changche}. 창의적 체험활동"))
        elements.append(Paragraph("※ 각 항목은 1~10점 척도로 평가됩니다. 상세 평가 기준은 '평가 기준 안내' 페이지를 참고하세요.", style_ref_note))
        elements.append(Spacer(1, 2*mm))

        # 창체 영역별 항목명
        ch_items_map = {
            "자율": ["참여\n주도성", "활동\n구체성", "활동\n완결성", "성장변화", "전공\n적합성"],
            "동아리": ["전문성", "개인\n기여도", "활동\n완결성", "지속성\n발전", "전공\n적합성"],
            "진로": ["탐색\n구체성", "목표\n일관성", "활동\n완결성", "자기\n주도성", "전공\n적합성"],
        }

        # 창체 테이블 colWidths (학년, 5항목, 평점, 등급, 분량, 비율, 활용)
        _c2_fixed = 42 + 50 + 34 + 50 + 40 + 48  # = 264
        _c2_each = (page_w - _c2_fixed) / 5
        cw_ch = [42] + [_c2_each]*5 + [50, 34, 50, 40, 48]

        for area_kr, sub_num in [("자율활동", f"{_sec_changche}-1"), ("동아리활동", f"{_sec_changche}-2"), ("진로활동", f"{_sec_changche}-3")]:
            area_short = area_kr.replace("활동", "")
            elements.append(section_title(f"{sub_num}. {area_kr}"))
            items = ch_items_map[area_short]
            hdr = [PH("학년")] + [PH(x) for x in items] + [PH("평점"), PH("등급"), PH("분량"), PH("비율"), PH("활용")]
            tdata = [hdr]
            for d in changche_data:
                if d[1] == area_short:
                    yr = d[0]
                    scores = d[3]
                    scaled = d[5]
                    avg_score = scaled / 10  # 환산점수를 10으로 나눠 평점으로 변환
                    vol, pct, usage = d[6], d[7], d[8]
                    grade = score_to_grade(avg_score)
                    row = [PC(f"{yr}학년")]
                    row += [PC(str(s)) for s in scores]
                    row += [PC(f"{avg_score:.2f}"), grade_paragraph(grade), PC(vol), PC(pct), PC(usage)]
                    tdata.append(row)
            t = make_table(tdata, col_widths=cw_ch)
            for r in range(1, len(tdata)):
                area_rows = [d for d in changche_data if d[1] == area_short]
                scaled = area_rows[r-1][5]
                g = score_to_grade(scaled / 10)
                t.setStyle(TableStyle([
                    ("BACKGROUND", (7, r), (7, r), colors.white),
                    ("TEXTCOLOR", (7, r), (7, r), grade_font_color(g)),
                ]))
            elements.append(t)
            elements.append(Spacer(1, 2*mm))

            # 창체 코멘트 테이블 (학년별 강점/보완점/역량)
            style_badge_academic = ParagraphStyle("BA", fontName="MalgunBd", fontSize=7, leading=9, textColor=colors.HexColor("#006666"), alignment=1)
            style_badge_career = ParagraphStyle("BC", fontName="MalgunBd", fontSize=7, leading=9, textColor=colors.HexColor("#CC9600"), alignment=1)
            style_badge_community = ParagraphStyle("BCC", fontName="MalgunBd", fontSize=7, leading=9, textColor=colors.HexColor("#7B2D8E"), alignment=1)

            hdr_c = [PH("학년"), PH("강점"), PH("보완점"), PH("핵심역량")]
            tdata_c = [hdr_c]
            comp_col_w = page_w * 0.28  # 핵심역량 칼럼 (코멘트 포함)
            remain = page_w - 42 - comp_col_w
            cw_c = [42, remain/2, remain/2, comp_col_w]

            for d in changche_data:
                if d[1] == area_short:
                    yr = d[0]
                    key = (area_short, yr)
                    if key in changche_comments:
                        vals = changche_comments[key]
                        strength = vals[0]
                        weakness = vals[1]
                        comp_comment = vals[3] if len(vals) > 3 else ""

                        # 핵심역량 칼럼: 세특과 동일한 PComp 함수로 컬러 적용
                        comp_p = PComp(comp_comment) if comp_comment else PL("-")

                        tdata_c.append([PC(f"{yr}학년"), PL(strength), PL(weakness), comp_p])

            if len(tdata_c) > 1:
                t_c = make_table(tdata_c, col_widths=cw_c)
                elements.append(t_c)

            elements.append(Spacer(1, 6*mm))

        # ── 창체 역량 매핑 정리표 ──
        elements.append(P("■ 창의적 체험활동 역량 매핑", style_subtitle))
        hdr_ch_comp = [PH("영역"), PH("학년"), PH("학업역량"), PH("진로역량"), PH("공동체역량")]
        tdata_ch_comp = [hdr_ch_comp]
        cw_ch_comp = [70, 42, (page_w - 112) / 3, (page_w - 112) / 3, (page_w - 112) / 3]
        ch_comp_areas = []
        for key, vals in changche_comments.items():
            area_short, yr = key
            competencies = vals[2] if len(vals) > 2 else []
            has_a = "학업" in competencies
            has_c = "진로" in competencies
            has_co = "공동체" in competencies
            area_kr = {"자율": "자율활동", "동아리": "동아리활동", "진로": "진로활동"}.get(area_short, area_short)
            dot_a = Paragraph("\u25cf", style_dot_academic) if has_a else PC("")
            dot_c = Paragraph("\u25cf", style_dot_career) if has_c else PC("")
            dot_co = Paragraph("\u25cf", style_dot_community) if has_co else PC("")
            tdata_ch_comp.append([PL(area_kr), PC(f"{yr}학년"), dot_a, dot_c, dot_co])
            ch_comp_areas.append(area_kr)

        t_ch_comp = make_table(tdata_ch_comp, col_widths=cw_ch_comp)

        # 같은 영역 셀병합
        prev_area = None
        span_start = None
        for r, area in enumerate(ch_comp_areas):
            row = r + 1
            if area == prev_area:
                pass
            else:
                if prev_area is not None and span_start is not None and row - 1 > span_start:
                    t_ch_comp.setStyle(TableStyle([
                        ("SPAN", (0, span_start), (0, row - 1)),
                        ("VALIGN", (0, span_start), (0, row - 1), "MIDDLE"),
                    ]))
                span_start = row
                prev_area = area
        if prev_area is not None and span_start is not None and len(ch_comp_areas) > span_start:
            t_ch_comp.setStyle(TableStyle([
                ("SPAN", (0, span_start), (0, len(ch_comp_areas))),
                ("VALIGN", (0, span_start), (0, len(ch_comp_areas)), "MIDDLE"),
            ]))

        elements.append(t_ch_comp)
        elements.append(PageBreak())

    # ── 4. 행동특성 및 종합의견 (mode_config 로 스킵 가능) ──
    if inc_haengtuk:
        elements.append(section_title(f"{_sec_haengtuk}. 행동특성 및 종합의견"))
        elements.append(Paragraph("※ 각 항목은 1~10점 척도로 평가됩니다. 상세 평가 기준은 '평가 기준 안내' 페이지를 참고하세요.", style_ref_note))
        hdr = [PH("학년"), PH("강점"), PH("보완점"), PH("핵심역량")]
        tdata = [hdr]
        h_comp_col_w = page_w * 0.28
        h_remain = page_w - 42 - h_comp_col_w
        cw = [42, h_remain/2, h_remain/2, h_comp_col_w]
        for yr in sorted(haengtuk_comments.keys()):
            vals = haengtuk_comments[yr]
            strength = vals[0]
            weakness = vals[1]
            comp_comment = vals[3] if len(vals) > 3 else ""
            comp_p = PComp(comp_comment) if comp_comment else PL("-")
            tdata.append([PC(f"{yr}학년"), PL(strength), PL(weakness), comp_p])
        t = make_table(tdata, col_widths=cw)
        elements.append(t)
        elements.append(PageBreak())

    # ── 5. 연계성분석 ──
    elements.append(section_title(f"{_sec_linkage}. 연계성분석"))
    hdr = [PH("분석영역"), PH("연계도"), PH("상세내용")]
    tdata = [hdr]
    cw = [90, 36, page_w - 126]
    for area, level, detail in linkage_data:
        tdata.append([PL(area), PC(level), PL(detail)])
    t = make_table(tdata, col_widths=cw)
    elements.append(t)
    elements.append(Spacer(1, 10*mm))

    # ── 6. 대학평가요소 ──
    elements.append(section_title(f"{_sec_eval}. 대학평가요소"))
    hdr = [PH("대분류"), PH("세부항목"), PH("등급"), PH("근거")]
    tdata = [hdr]
    cw = [60, 80, 28, page_w - 168]
    for cat, item, grade, reason in eval_data:
        tdata.append([PL(cat), PL(item), grade_paragraph(grade), PL(reason)])
    t = make_table(tdata, col_widths=cw)

    # 등급 컬러
    for r in range(1, len(tdata)):
        g = eval_data[r-1][2]
        t.setStyle(TableStyle([
            ("BACKGROUND", (2, r), (2, r), colors.white),
            ("TEXTCOLOR", (2, r), (2, r), grade_font_color(g)),
        ]))

    # 대분류 셀병합
    prev_cat = None
    span_start = None
    for r in range(len(eval_data)):
        row = r + 1  # 헤더 offset
        cat = eval_data[r][0]
        if cat == prev_cat:
            pass
        else:
            if prev_cat is not None and span_start is not None and row - 1 > span_start:
                t.setStyle(TableStyle([
                    ("SPAN", (0, span_start), (0, row - 1)),
                    ("VALIGN", (0, span_start), (0, row - 1), "MIDDLE"),
                ]))
            span_start = row
            prev_cat = cat
    # 마지막 그룹
    if prev_cat is not None and span_start is not None and len(eval_data) > span_start:
        t.setStyle(TableStyle([
            ("SPAN", (0, span_start), (0, len(eval_data))),
            ("VALIGN", (0, span_start), (0, len(eval_data)), "MIDDLE"),
        ]))

    elements.append(t)
    elements.append(PageBreak())

    # ── 7. 역량별보완법 ──
    elements.append(section_title(f"{_sec_fix}. 역량별보완법"))
    hdr = [PH("역량"), PH("항목"), PH("현재"), PH("진단"), PH("보완활동"), PH("중요도")]
    tdata = [hdr]
    # 역량40(전체/학업/진로/공동체 한줄), 항목60(3학년비교과 등), 현재32(B~C 한줄), 진단90, 중요도38(중 한줄)
    _fix_fixed = 40 + 60 + 32 + 90 + 38  # = 260
    cw = [40, 60, 32, 90, page_w - _fix_fixed, 38]
    for cap, item, cur, diag, action, pri in fix_data:
        tdata.append([PL(cap), PL(item), PC(cur), PL(diag), PL(action), PC(pri)])
    t = make_table(tdata, col_widths=cw)
    elements.append(t)

    # ── 키워드분석 섹션 (G5 / CLAUDE.md § Step 8-4) ──
    # raw_texts 가 비어있으면 이 섹션 스킵. 섹션 번호는 _sec_keyword 변수로 동적.
    try:
        from .keyword_extractor import has_raw_texts
        if has_raw_texts(sd):
            elements.append(PageBreak())
            _append_pdf_keyword_section(elements, sd, pdf_path, section_title,
                                         P, PC, PL, PH, make_table,
                                         style_subtitle, page_w,
                                         section_number=_sec_keyword)
    except Exception as e:
        print(f"[WARN] PDF 키워드분석 섹션 생성 실패 (스킵): {type(e).__name__}: {e}")

    doc.build(elements, onFirstPage=_footer_cover_page, onLaterPages=_footer_content_pages)
    print(f"PDF saved: {pdf_path}")


def _append_pdf_keyword_section(elements, sd, pdf_path, section_title,
                                 P, PC, PL, PH, make_table,
                                 style_subtitle, page_w,
                                 section_number: int = 8):
    """PDF 키워드분석 섹션 추가. 워드클라우드 이미지 + 카테고리 요약 + 학년별 변화.
    section_number: 동적 섹션 번호 (기본 8, 출결 섹션 있을 때 9 등).
    """
    from pathlib import Path

    from .keyword_extractor import extract_keywords, generate_wordcloud_image

    report = extract_keywords(sd.raw_texts, top_n=50, min_frequency=2)
    if not report.keywords:
        return

    elements.append(section_title(f"{section_number}. 키워드분석"))

    # ── 워드클라우드 이미지 생성 & 삽입 ──
    pdf_dir = Path(pdf_path).resolve().parent
    wc_png = pdf_dir / f"{sd.STUDENT}_wordcloud_{sd.TODAY}.png"
    try:
        generated = generate_wordcloud_image(report, wc_png)
    except Exception as e:
        print(f"[WARN] 워드클라우드 이미지 생성 실패: {type(e).__name__}: {e}")
        generated = None

    if generated and generated.exists():
        elements.append(P("■ 핵심 키워드 워드클라우드", style_subtitle))
        # 이미지 크기: 페이지 너비 90%
        img_w = page_w * 0.95
        img_h = img_w * 0.5  # 2:1 비율
        try:
            img = RLImage(str(generated), width=img_w, height=img_h)
            elements.append(img)
            elements.append(Spacer(1, 4*mm))
        except Exception as e:
            print(f"[WARN] 워드클라우드 PDF 삽입 실패: {type(e).__name__}: {e}")

    # ── 카테고리별 상위 키워드 테이블 ──
    elements.append(P("■ 역량 카테고리별 상위 키워드", style_subtitle))
    CATS = ["학업역량", "진로역량", "공동체역량", "일반"]
    by_cat = {c: [] for c in CATS}
    for e in report.keywords:
        if e.category in by_cat:
            by_cat[e.category].append(e)

    hdr = [PH("카테고리"), PH("상위 키워드 (빈도)")]
    tdata = [hdr]
    cw = [80, page_w - 80]
    for cat in CATS:
        entries = by_cat.get(cat, [])[:10]
        if not entries:
            tdata.append([PL(cat), PL("해당 없음")])
        else:
            txt = ", ".join(f"{e.word}({e.frequency})" for e in entries)
            tdata.append([PL(cat), PL(txt)])
    elements.append(make_table(tdata, col_widths=cw))
    elements.append(Spacer(1, 4*mm))

    # ── 학년별 변화 추이 ──
    elements.append(P("■ 학년별 키워드 변화 추이", style_subtitle))
    hdr = [PH("학년"), PH("신규 등장"), PH("사라진 키워드")]
    tdata = [hdr]
    cw = [40, (page_w - 40) / 2, (page_w - 40) / 2]
    for yr, ch in sorted(report.yearly_changes.items()):
        new_list = ", ".join(ch.get("new", [])[:10]) or "-"
        gone_list = ", ".join(ch.get("disappeared", [])[:10]) or "-"
        tdata.append([PC(f"{yr}학년"), PL(new_list), PL(gone_list)])
    elements.append(make_table(tdata, col_widths=cw))


def _append_pdf_compare_section(elements, sd, section_title,
                                  P, PC, PL, PH, make_table,
                                  style_subtitle, page_w,
                                  section_number: int = 2):
    """PDF 이전 분석 대비 변화 섹션 (G3+G4 / CLAUDE.md § Step 0-4).
    compare_data 있을 때만 호출됨.
    """
    from reportlab.platypus import Spacer as _RLSpacer

    cd = sd.compare_data or {}
    elements.append(section_title(f"{section_number}. 이전 분석 대비 변화"))
    elements.append(_RLSpacer(1, 2*mm))

    # ── 이전 리포트 정보 ──
    prev_round = cd.get("previous_round", "?")
    prev_date = cd.get("previous_date", "-")
    prev_version = cd.get("previous_version", 1)
    elements.append(P(
        f"<b>이전 리포트</b>: {prev_round}회차 (v{prev_version}, {prev_date}) → 현재",
        style_subtitle))
    elements.append(_RLSpacer(1, 2*mm))

    # ── 등급 변화 추이 ──
    elements.append(P("■ 등급 변화 추이", style_subtitle))
    hdr = [PH("영역"), PH("이전"), PH("현재"), PH("변화")]
    tdata = [hdr]
    cw = [80, 50, 50, 50]
    for gc in cd.get("grade_changes", []) or []:
        arrow = gc.get("변화", "-")
        arrow_str = arrow
        if arrow == "↑":
            arrow_str = '<font color="#2E75B6"><b>↑</b></font>'
        elif arrow == "↓":
            arrow_str = '<font color="#C00000"><b>↓</b></font>'
        tdata.append([PC(gc.get("영역", "-")), PC(gc.get("이전", "-")),
                       PC(gc.get("현재", "-")), P(arrow_str, style_subtitle)])
    if len(tdata) > 1:
        elements.append(make_table(tdata, col_widths=cw))
    else:
        elements.append(P("등급 변화 데이터 없음", style_subtitle))
    elements.append(_RLSpacer(1, 4*mm))

    # ── 이전 핵심강점 추적 ──
    elements.append(P("■ 이전 핵심강점 추적", style_subtitle))
    hdr = [PH("이전 강점"), PH("현재 상태"), PH("근거")]
    tdata = [hdr]
    cw = [page_w * 0.30, 60, page_w - 60 - page_w * 0.30]
    for s in cd.get("strengths_tracking", []) or []:
        state = s.get("현재상태", "")
        state_str = state
        if state == "강화됨":
            state_str = f'<font color="#2E75B6"><b>{state}</b></font>'
        elif state == "약화됨":
            state_str = f'<font color="#C00000"><b>{state}</b></font>'
        tdata.append([PL(s.get("이전강점", "")), P(state_str, style_subtitle), PL(s.get("근거", ""))])
    if len(tdata) > 1:
        elements.append(make_table(tdata, col_widths=cw))
    else:
        elements.append(P("추적 데이터 없음", style_subtitle))
    elements.append(_RLSpacer(1, 4*mm))

    # ── 이전 보완점 반영도 ──
    elements.append(P("■ 이전 보완점 반영도", style_subtitle))
    hdr = [PH("이전 보완점"), PH("상태"), PH("근거")]
    tdata = [hdr]
    cw = [page_w * 0.30, 60, page_w - 60 - page_w * 0.30]
    for s in cd.get("issues_tracking", []) or []:
        state = s.get("상태", "")
        state_str = state
        if state == "반영됨":
            state_str = f'<font color="#2E75B6"><b>{state}</b></font>'
        elif state == "미반영":
            state_str = f'<font color="#C00000"><b>{state}</b></font>'
        elif state == "부분반영":
            state_str = f'<font color="#ED7D31"><b>{state}</b></font>'
        tdata.append([PL(s.get("이전보완점", "")), P(state_str, style_subtitle), PL(s.get("근거", ""))])
    if len(tdata) > 1:
        elements.append(make_table(tdata, col_widths=cw))
    else:
        elements.append(P("추적 데이터 없음", style_subtitle))
    elements.append(_RLSpacer(1, 4*mm))

    # ── 새 강점 / 새 보완점 ──
    new_s = cd.get("new_strengths", []) or []
    new_i = cd.get("new_issues", []) or []
    if new_s:
        elements.append(P("■ 새로 발견된 강점", style_subtitle))
        for i, s in enumerate(new_s, 1):
            elements.append(P(f"{i}. {s}", style_subtitle))
        elements.append(_RLSpacer(1, 3*mm))
    if new_i:
        elements.append(P("■ 새로 발견된 보완점", style_subtitle))
        for i, s in enumerate(new_i, 1):
            elements.append(P(f"{i}. {s}", style_subtitle))
        elements.append(_RLSpacer(1, 3*mm))

    # ── 성장 코멘트 ──
    growth = str(cd.get("growth_comment", "") or "")
    if growth:
        elements.append(P("■ 성장 코멘트", style_subtitle))
        elements.append(P(growth, style_subtitle))


def _append_pdf_attendance_section(elements, sd, section_title,
                                    P, PC, PL, PH, make_table,
                                    style_subtitle, page_w,
                                    section_number: int = 2):
    """PDF "출결 및 봉사활동" 섹션 추가 (G7 / CLAUDE.md § 4).
    attendance_data / volunteer_data 있을 때만 호출됨.
    """
    from reportlab.lib import colors as _rl_colors
    from reportlab.platypus import Spacer as _RLSpacer
    from reportlab.platypus import TableStyle as _RLTableStyle

    from .attendance_calculator import calculate_attendance_score, summarize_volunteer

    att_report = calculate_attendance_score(getattr(sd, "attendance_data", {}) or {})
    vol_summary = summarize_volunteer(getattr(sd, "volunteer_data", {}) or {})

    elements.append(section_title(f"{section_number}. 출결 및 봉사활동"))
    elements.append(_RLSpacer(1, 2*mm))

    # ── 출결 학년별 매트릭스 테이블 ──
    if att_report.has_data:
        elements.append(P("■ 출결 현황 (학년별)", style_subtitle))
        # 2-row 헤더: Row0 "결석/지각/조퇴/결과" merge, Row1 "질병/미인정/기타"
        hdr0 = [PH("학년"), PH("결석"), PH(""), PH(""), PH("지각"), PH(""), PH(""),
                PH("조퇴"), PH(""), PH(""), PH("결과"), PH(""), PH("")]
        hdr1 = [PH("")] + [PH("질"), PH("미"), PH("기")] * 4
        tdata = [hdr0, hdr1]
        _yr_col = 36
        _each = (page_w - _yr_col) / 12
        cw = [_yr_col] + [_each] * 12
        for yr, entry in sorted(att_report.by_year.items()):
            row = [PC(f"{yr}학년")]
            for t in ("결석", "지각", "조퇴", "결과"):
                for r in ("질병", "미인정", "기타"):
                    v = entry[t][r]
                    # 미인정 은 붉은색 강조
                    if r == "미인정" and v > 0:
                        row.append(P(f"<font color='#C00000'><b>{v}</b></font>", style_subtitle))
                    else:
                        row.append(PC(str(v)))
            tdata.append(row)

        t = make_table(tdata, col_widths=cw, has_header=False, repeat_rows=2)
        t.setStyle(_RLTableStyle([
            ("BACKGROUND", (0, 0), (-1, 1), _rl_colors.HexColor("#0E2841")),
            ("TEXTCOLOR", (0, 0), (-1, 1), _rl_colors.white),
            ("SPAN", (1, 0), (3, 0)),   # 결석 merge
            ("SPAN", (4, 0), (6, 0)),   # 지각 merge
            ("SPAN", (7, 0), (9, 0)),   # 조퇴 merge
            ("SPAN", (10, 0), (12, 0)), # 결과 merge
            ("SPAN", (0, 0), (0, 1)),   # 학년 merge
            ("BACKGROUND", (0, 2), (-1, -1), _rl_colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "MalgunBd"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, _rl_colors.HexColor("#D0D0D0")),
        ]))
        elements.append(t)
        elements.append(_RLSpacer(1, 4*mm))

        # 출결 점수 요약
        elements.append(P("■ 출결 점수 (100점 만점, 미인정만 감점)", style_subtitle))
        score_hdr = [PH("기본"), PH("감점 합계"), PH("최종점수")]
        total_ded = sum(att_report.deductions.values())
        score_row = [PC(f"{att_report.base:.0f}"),
                     PC(f"{total_ded:+.1f}"),
                     P(f"<b>{att_report.score:.1f}</b>", style_subtitle)]
        elements.append(make_table([score_hdr, score_row],
                                    col_widths=[page_w/3]*3))

        # 감점 내역 (있을 때)
        if att_report.deductions:
            elements.append(_RLSpacer(1, 2*mm))
            det_lines = " · ".join(f"{k}: {v:+.1f}" for k, v in att_report.deductions.items())
            elements.append(P(f"감점 내역 — {det_lines}", style_subtitle))
    else:
        elements.append(P("■ 출결 데이터 미입력 (만점 100점으로 가정)", style_subtitle))

    elements.append(_RLSpacer(1, 5*mm))

    # ── 봉사활동 ──
    elements.append(P("■ 봉사활동 시수", style_subtitle))
    if vol_summary.has_data:
        hdr = [PH("학년"), PH("시간"), PH("주요 활동")]
        tdata = [hdr]
        cw = [50, 50, page_w - 100]
        for yr, entry in sorted(vol_summary.by_year.items()):
            acts = ", ".join(entry.get("activities") or []) or "-"
            tdata.append([PC(f"{yr}학년"), PC(f"{entry['hours']}h"), PL(acts)])
        tdata.append([PC("합계"), PC(f"{vol_summary.total_hours}h"), PL("")])
        elements.append(make_table(tdata, col_widths=cw))
    else:
        elements.append(P("봉사활동 시수 데이터 미입력", style_subtitle))


def _append_pdf_grading_section(elements, sd, section_title,
                                 P, PC, PL, PH, make_table,
                                 style_subtitle, page_w,
                                 section_number: int = 4):
    """PDF "대학별 내신" 섹션 추가 (2026-05-05).

    grade_data 있을 때만 호출됨. baseline 3개(자연/인문/종합) +
    지망 대학 룰(있으면) 의 환산 결과를 표로 표시.
    """
    from reportlab.platypus import Spacer as _RLSpacer

    from .grade_analyzer import calc_all_grading

    grade_data = getattr(sd, "grade_data", {}) or {}
    target_univ      = (getattr(sd, "TARGET_UNIV", "") or "").strip()
    target_admission = (getattr(sd, "TARGET_ADMISSION_TYPE", "") or "").strip()
    target_category  = (getattr(sd, "TARGET_ADMISSION_CATEGORY", "") or "").strip()

    out = calc_all_grading(
        grade_data,
        university=target_univ or None,
        admission_type=target_admission or None,
        admission_category=target_category or None,
    )

    elements.append(section_title(f"{section_number}. 대학별 내신 산출"))
    elements.append(_RLSpacer(1, 2*mm))

    # ── 표 1: 산출 룰별 평균등급 / 환산점수 ──
    elements.append(P("■ 산출 룰별 평균등급 / 환산점수", style_subtitle))
    hdr = [PH("구분"), PH("라벨"), PH("평균등급"), PH("환산점수"),
           PH("적용\n과목수"), PH("제외\n과목수")]
    tdata = [hdr]

    for b in out.get('baseline') or []:
        r = b.get('result') or {}
        tdata.append([
            PC(f"baseline\n({b.get('track', '')})"),
            PL(b.get('label', '')),
            PC(f"{r.get('평균등급', 0):.3f}"),
            PC(f"{r.get('환산점수', 0):.2f}"),
            PC(str(r.get('적용_과목수', 0))),
            PC(str(r.get('제외_과목수', 0))),
        ])

    u = out.get('university')
    if u:
        if u.get('matched'):
            r = u.get('result') or {}
            tdata.append([
                PC("지망 대학"),
                PL(u.get('label', '')),
                PC(f"{r.get('평균등급', 0):.3f}"),
                PC(f"{r.get('환산점수', 0):.2f}"),
                PC(str(r.get('적용_과목수', 0))),
                PC(str(r.get('제외_과목수', 0))),
            ])
        else:
            tdata.append([
                PC("지망 대학"),
                PL(u.get('message', '미매칭')),
                PC("-"), PC("-"), PC("-"), PC("-"),
            ])

    cw = [70, page_w - 70 - 60 - 60 - 50 - 50, 60, 60, 50, 50]
    elements.append(make_table(tdata, col_widths=cw))
    elements.append(_RLSpacer(1, 5*mm))

    # 대표 룰 결정 (지망 대학 매칭 우선, 없으면 baseline 종합)
    repr_block = u if (u and u.get('matched')) else (
        next((b for b in (out.get('baseline') or []) if b.get('track') == '종합'), None)
    )
    repr_label = repr_block.get('label', '-') if repr_block else '-'
    repr_result = (repr_block or {}).get('result') or {}

    # ── 표 2: 학년별 평균 ──
    by_year = repr_result.get('breakdown', {}).get('by_year', {})
    if any(by_year.values()):
        elements.append(P(f"■ 학년별 평균등급 (대표 룰: {repr_label})", style_subtitle))
        yr_hdr = [PH("1학년"), PH("2학년"), PH("3학년")]
        yr_row = [
            PC(f"{by_year.get(1, 0):.3f}"),
            PC(f"{by_year.get(2, 0):.3f}"),
            PC(f"{by_year.get(3, 0):.3f}"),
        ]
        elements.append(make_table([yr_hdr, yr_row], col_widths=[page_w/3]*3))
        elements.append(_RLSpacer(1, 5*mm))

    # ── 표 3: 교과별 평균 ──
    by_cat = repr_result.get('breakdown', {}).get('by_category', {})
    if by_cat:
        elements.append(P(f"■ 교과별 평균등급 (대표 룰: {repr_label})", style_subtitle))
        cats = list(by_cat.keys())
        cat_hdr = [PH(c) for c in cats]
        cat_row = [PC(f"{by_cat[c]:.3f}") for c in cats]
        col_w = page_w / len(cats) if cats else page_w
        elements.append(make_table([cat_hdr, cat_row], col_widths=[col_w]*len(cats)))
        elements.append(_RLSpacer(1, 5*mm))

    # ── 표 4: 산출 메모 (notes) ──
    notes_collected = []
    for b in out.get('baseline') or []:
        for n in (b.get('result') or {}).get('notes') or []:
            notes_collected.append(f"[{b.get('track', '')}] {n}")
    if u and u.get('matched'):
        for n in (u.get('result') or {}).get('notes') or []:
            notes_collected.append(f"[지망] {n}")

    if notes_collected:
        elements.append(P("■ 산출 메모", style_subtitle))
        for note in notes_collected:
            elements.append(P(f"• {note}", style_subtitle))
