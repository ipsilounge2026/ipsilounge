# -*- coding: utf-8 -*-
"""
compare_generator.py
- G3+G4 (2026-04-17): 2회차 이상 분석 시 이전 리포트와 비교
- 이전 Excel 파싱 → 핵심강점/보완영역/fix_data 등 자동 추출
- 현재 학생 데이터와 diff → grade_changes 자동 생성

CLAUDE.md § Step 0 (기존 리포트 확인) + § 0-4 (이전 대비 변화 분석) 구현체.

사용 워크플로우:
    1. output/ 폴더에 이전 리포트(xlsx) 존재 확인
    2. find_previous_reports(student_name) → 리포트 경로 + 회차
    3. extract_previous_info(prev_xlsx) → PreviousReport dataclass
    4. compute_grade_changes(prev, current_sd) → grade_changes 리스트
    5. build_tracking_targets(prev) → strengths/issues 판정 대상 리스트

이 결과를 Claude 대화에서 참조하여 compare_data 를 작성하면,
QA 검증(P1-G) 이 판정 대상 전수 tracking 여부를 자동 체크함.

CLI 사용:
    python -m modules.compare_generator <학생명>
    → 이전 리포트 자동 탐색 + 파싱 결과 콘솔 출력
"""

from __future__ import annotations
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple


# ═══════════════════════════════════════════════════════
# 데이터 클래스
# ═══════════════════════════════════════════════════════

@dataclass
class PreviousReport:
    """이전 리포트에서 자동 추출한 정보."""
    path: Path
    round_num: int                         # 회차 (1 = 최초)
    date_str: str                          # 리포트 날짜 (YYYYMMDD)

    # 종합요약 시트에서 추출
    core_strengths: List[str] = field(default_factory=list)   # 핵심강점 1~3
    core_issues: List[str] = field(default_factory=list)      # 보완영역 1~3
    growth_story: str = ""

    # 역량별보완법 시트에서 추출 (역량, 항목, 보완활동)
    fix_items: List[Dict[str, str]] = field(default_factory=list)

    # 각 영역 평균 등급 (계산)
    setuek_grade: str = "-"
    changche_grade: str = "-"
    haengtuk_grade: str = "-"

    # 비교과 종합 (종합요약에서 raw 문자열)
    overall_summary_text: str = ""


@dataclass
class TrackingTargets:
    """Claude 가 compare_data 작성 시 반드시 판정해야 할 대상들."""
    strengths_to_track: List[Dict[str, str]]  # [{"이전강점": "...", "출처": "종합요약-핵심강점1"}]
    issues_to_track: List[Dict[str, str]]     # [{"이전보완점": "...", "출처": "종합요약-보완영역1"}]


# ═══════════════════════════════════════════════════════
# 이전 리포트 탐색
# ═══════════════════════════════════════════════════════

_REPORT_PATTERN = re.compile(r"^(?P<name>.+?)_학생부분석_(?P<date>\d{8})(?:_v(?P<version>\d+))?\.xlsx$")


def find_previous_reports(student_name: str, output_dir: Optional[Path] = None) -> List[Tuple[Path, int, str, int]]:
    """학생명에 해당하는 기존 리포트 파일 탐색.
    반환: [(Path, version_num, date_str, round_num), ...] - 날짜·버전 오름차순 정렬.
      - version_num: _v{N} 의 N (없으면 1)
      - round_num: 전체 목록에서의 순서 (1, 2, 3, ...)
    """
    if output_dir is None:
        output_dir = Path(__file__).resolve().parent.parent / "output"
    if not output_dir.exists():
        return []

    found = []
    for p in output_dir.glob("*.xlsx"):
        m = _REPORT_PATTERN.match(p.name)
        if not m:
            continue
        if m.group("name") != student_name:
            continue
        version = int(m.group("version")) if m.group("version") else 1
        found.append((p, version, m.group("date")))

    # 정렬: 날짜 → 버전 오름차순
    found.sort(key=lambda x: (x[2], x[1]))
    # round_num 부여
    return [(p, v, d, i + 1) for i, (p, v, d) in enumerate(found)]


def get_next_version_number(student_name: str, today: str,
                             output_dir: Optional[Path] = None) -> Optional[int]:
    """다음 리포트 파일명에 부여할 버전 번호.
    - 기존 리포트 없음 → None (접미사 없이 기본 파일명 사용)
    - 기존 1개 이상 → 최댓값 + 1
    - 같은 날짜로 동일 version 이 이미 있으면 해당 version + 1 (덮어쓰기 방지)
    """
    existing = find_previous_reports(student_name, output_dir)
    if not existing:
        return None
    # 모든 (date, version) 페어 중 최댓값 기준
    max_round = max(v for _, v, _, _ in existing)
    # 같은 날짜의 최고 버전 + 1
    same_date = [v for _, v, d, _ in existing if d == today]
    if same_date:
        return max(max_date := max(same_date), max_round) + 1
    return max_round + 1


# ═══════════════════════════════════════════════════════
# 이전 Excel 파싱
# ═══════════════════════════════════════════════════════

def _grade_from_scores(grades: List[str]) -> str:
    """과목별 등급 리스트 → 대표 등급 (최빈값, 동점이면 상위 우선)."""
    if not grades:
        return "-"
    order = ["S", "A", "B", "C", "D"]
    counts = {g: grades.count(g) for g in order}
    max_count = max(counts.values())
    for g in order:
        if counts[g] == max_count and max_count > 0:
            return g
    return "-"


def extract_previous_info(xlsx_path: Path, round_num: int = 1) -> PreviousReport:
    """이전 리포트 Excel 파싱. report_logic.create_excel() 에서 생성한 구조를 전제.
    """
    from openpyxl import load_workbook

    m = _REPORT_PATTERN.match(xlsx_path.name)
    date_str = m.group("date") if m else ""

    prev = PreviousReport(path=xlsx_path, round_num=round_num, date_str=date_str)

    wb = load_workbook(xlsx_path, data_only=True)

    # ── 1. 종합요약 ──
    if "종합요약" in wb.sheetnames:
        ws = wb["종합요약"]
        rows = {}
        for r in range(2, ws.max_row + 1):
            key = ws.cell(r, 1).value
            val = ws.cell(r, 2).value
            if key:
                rows[str(key).strip()] = "" if val is None else str(val)
        for i in (1, 2, 3):
            v = rows.get(f"핵심강점{i}", "").strip()
            if v:
                prev.core_strengths.append(v)
            v = rows.get(f"보완영역{i}", "").strip()
            if v:
                prev.core_issues.append(v)
        prev.growth_story = rows.get("성장스토리", "")
        prev.overall_summary_text = rows.get("비교과 종합", rows.get("비교과종합", ""))

    # ── 2. 역량별보완법 → fix_items ──
    if "역량별보완법" in wb.sheetnames:
        ws = wb["역량별보완법"]
        # 헤더: 역량 / 항목 / 현재등급 / 진단 / 보완활동 / 중요도
        for r in range(2, ws.max_row + 1):
            cap = ws.cell(r, 1).value
            item = ws.cell(r, 2).value
            action = ws.cell(r, 5).value
            pri = ws.cell(r, 6).value
            if not cap and not item:
                continue
            prev.fix_items.append({
                "역량": str(cap or "").strip(),
                "항목": str(item or "").strip(),
                "보완활동": str(action or "").strip(),
                "중요도": str(pri or "").strip(),
            })

    # ── 3. 각 영역 대표 등급 ──
    def _collect_grades(sheet_name: str, grade_col: int) -> List[str]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        grades = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, grade_col).value
            if v and str(v).strip() in ("S", "A", "B", "C", "D"):
                grades.append(str(v).strip())
        return grades

    # 세특분석: 등급은 마지막 컬럼. 미지정 모드 10열 / 지정 모드 11열이므로 동적.
    if "세특분석" in wb.sheetnames:
        ws = wb["세특분석"]
        grade_col = ws.max_column  # 마지막 컬럼 = 등급
        prev.setuek_grade = _grade_from_scores(_collect_grades("세특분석", grade_col))

    # 창체분석: "등급" 컬럼 위치 변동 있을 수 있음. 헤더에서 "등급" 찾기
    if "창체분석" in wb.sheetnames:
        ws = wb["창체분석"]
        grade_col_idx = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(1, c).value == "등급":
                grade_col_idx = c
                break
        if grade_col_idx:
            prev.changche_grade = _grade_from_scores(_collect_grades("창체분석", grade_col_idx))

    # 행특분석: 등급 컬럼 없음(강점/보완점만). 대신 점수 데이터가 없으므로 "-" 유지
    # 만약 향후 등급 컬럼 추가되면 동일 방식.

    return prev


# ═══════════════════════════════════════════════════════
# Diff / 판정 대상 생성
# ═══════════════════════════════════════════════════════

def _current_area_grade(sd, area: str) -> str:
    """현재 학생 데이터(sd) 에서 영역별 대표 등급 계산.
    area: 'setuek' | 'changche' | 'haengtuk'
    """
    if area == "setuek":
        setuek_data = getattr(sd, "setuek_data", [])
        if not setuek_data:
            return "-"
        grades = [str(d[-1]) for d in setuek_data if len(d) >= 2]
        return _grade_from_scores(grades)
    if area == "changche":
        cd = getattr(sd, "changche_data", [])
        # changche_data 구조: (학년, 영역, 시간, [점수], 합, 환산, 분량, 비율, 활용) → 마지막에 등급 별도 아님
        # report_logic.create_excel 에서는 score_to_grade(scaled / 10) 로 동적 계산
        from .report_constants import CHANGCHE_ITEMS  # noqa (side effect 없음)
        grades = []
        for d in cd:
            try:
                scaled = d[5]
                from .report_logic import score_to_grade  # 순환 임포트 주의, 지연 임포트
                grades.append(score_to_grade(scaled / 10))
            except Exception:
                continue
        return _grade_from_scores(grades)
    if area == "haengtuk":
        hd = getattr(sd, "haengtuk_data", [])
        grades = [str(d[-1]) for d in hd if d and len(d) >= 5]
        return _grade_from_scores(grades)
    return "-"


def compute_grade_changes(prev: PreviousReport, sd) -> List[Dict[str, str]]:
    """이전 vs 현재 영역별 등급 비교 → grade_changes 리스트 자동 생성."""
    result = []
    _ORDER = {"S": 5, "A": 4, "B": 3, "C": 2, "D": 1, "-": 0}

    def _arrow(prev_g: str, cur_g: str) -> str:
        pv = _ORDER.get(prev_g, 0)
        cv = _ORDER.get(cur_g, 0)
        if pv == 0 or cv == 0:
            return "-"
        if cv > pv:
            return "↑"
        if cv < pv:
            return "↓"
        return "="

    for area_key, area_label in [("setuek", "세특"), ("changche", "창체"), ("haengtuk", "행특")]:
        prev_g = getattr(prev, f"{area_key}_grade", "-")
        cur_g = _current_area_grade(sd, area_key)
        result.append({
            "영역": area_label,
            "이전": prev_g,
            "현재": cur_g,
            "변화": _arrow(prev_g, cur_g),
        })
    return result


def build_tracking_targets(prev: PreviousReport) -> TrackingTargets:
    """Claude 가 compare_data 작성 시 반드시 판정해야 할 대상 리스트.
    - strengths: 종합요약 핵심강점 3개 (출처: 종합요약-핵심강점N)
    - issues: 종합요약 보완영역 3개 + 역량별보완법 fix_items 전체
    """
    strengths = []
    for i, s in enumerate(prev.core_strengths, 1):
        strengths.append({"이전강점": s, "출처": f"종합요약-핵심강점{i}"})

    issues = []
    for i, s in enumerate(prev.core_issues, 1):
        issues.append({"이전보완점": s, "출처": f"종합요약-보완영역{i}"})

    for item in prev.fix_items:
        cap = item.get("역량", "")
        it = item.get("항목", "")
        act = item.get("보완활동", "")
        # "보완점" 은 항목 + 보완활동 요약
        label = f"[{cap}] {it}" if cap and it else (it or cap or "?")
        issues.append({
            "이전보완점": f"{label} — {act[:80]}" if act else label,
            "출처": f"역량별보완법-{cap}" if cap else "역량별보완법",
        })

    return TrackingTargets(strengths_to_track=strengths, issues_to_track=issues)


# ═══════════════════════════════════════════════════════
# 편의 함수
# ═══════════════════════════════════════════════════════

def has_compare_data(sd) -> bool:
    """학생 데이터 모듈에 compare_data 가 있고 비어있지 않은지."""
    cd = getattr(sd, "compare_data", None)
    if not isinstance(cd, dict) or not cd:
        return False
    # 핵심 필드 1개라도 있으면 True
    for key in ("grade_changes", "strengths_tracking", "issues_tracking",
                "new_strengths", "new_issues", "growth_comment"):
        v = cd.get(key)
        if v:
            return True
    return False


# ═══════════════════════════════════════════════════════
# CLI (독립 실행)
# ═══════════════════════════════════════════════════════

def _print_analysis(student_name: str):
    """CLI: 이전 리포트 탐색 + 파싱 결과 출력."""
    reports = find_previous_reports(student_name)
    if not reports:
        print(f"[INFO] {student_name}: 기존 리포트 없음 (최초 분석)")
        return

    print(f"[INFO] {student_name}: 기존 리포트 {len(reports)}건 발견")
    print()
    for path, ver, date, rnd in reports:
        print(f"  #{rnd} {path.name} (날짜: {date}, 버전: v{ver})")
    print()

    # 가장 최근 리포트 파싱
    latest_path, latest_ver, latest_date, latest_rnd = reports[-1]
    print(f"[PARSE] 최근 리포트 파싱: {latest_path.name}")
    prev = extract_previous_info(latest_path, round_num=latest_rnd)
    print()
    print(f"  핵심강점 ({len(prev.core_strengths)}개):")
    for i, s in enumerate(prev.core_strengths, 1):
        print(f"    {i}. {s[:80]}")
    print(f"  보완영역 ({len(prev.core_issues)}개):")
    for i, s in enumerate(prev.core_issues, 1):
        print(f"    {i}. {s[:80]}")
    print(f"  역량별보완법 fix_items: {len(prev.fix_items)}개")
    print(f"  세특 대표등급: {prev.setuek_grade} / 창체: {prev.changche_grade} / 행특: {prev.haengtuk_grade}")
    print()

    targets = build_tracking_targets(prev)
    print(f"[TARGETS] Claude 가 compare_data 에서 반드시 판정해야 할 대상:")
    print(f"  - strengths_tracking 최소 {len(targets.strengths_to_track)}개")
    print(f"  - issues_tracking 최소 {len(targets.issues_to_track)}개")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("사용법: python -m modules.compare_generator <학생명>")
        sys.exit(1)
    _print_analysis(sys.argv[1])
