"""입결 데이터 임포트 스크립트

Excel 파일(admission_db.xlsx)의 '수시입결RAW' + '정시입결RAW' 시트를 읽어
admission_data / jeongsi_admission_data 테이블에 저장합니다.

사용법:
    python -m backend.scripts.import_admission_data <xlsx_path> [--clear]

옵션:
    --clear : 기존 데이터를 모두 삭제 후 임포트
"""

import argparse
import asyncio
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import delete, text
from sqlalchemy.ext.asyncio import AsyncSession

# backend 디렉터리를 PYTHONPATH에 추가 (app 모듈 찾기)
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database import async_session, engine, Base  # noqa: E402
from app.models.admission_data import AdmissionData  # noqa: E402
from app.models.jeongsi_admission_data import JeongsiAdmissionData  # noqa: E402


def _parse_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _parse_float(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _parse_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def load_rows_from_xlsx(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # 첫 번째 시트 사용 (수시입결RAW)
    ws = wb[wb.sheetnames[0]]

    rows = []
    # 데이터 헤더가 3행, 컬럼 인덱스 힌트가 4행, 실제 데이터는 5행부터
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue
        university = _parse_str(row[0])
        admission_category = _parse_str(row[1])
        admission_name = _parse_str(row[2])
        major = _parse_str(row[3])
        year = _parse_int(row[4])
        recruit_count = _parse_int(row[5])
        applicants = _parse_int(row[6])
        competition_rate = _parse_float(row[7])
        chu_hap = _parse_int(row[8])
        result_50 = _parse_float(row[9])
        result_70 = _parse_float(row[10])
        note = _parse_str(row[11]) if len(row) > 11 else None

        if not (university and major and year):
            continue

        rows.append({
            "university": university,
            "admission_category": admission_category,
            "admission_name": admission_name,
            "major": major,
            "year": year,
            "recruit_count": recruit_count,
            "applicants": applicants,
            "competition_rate": competition_rate,
            "chu_hap": chu_hap,
            "result_50": result_50,
            "result_70": result_70,
            "note": note,
        })
    return rows


async def import_rows(rows: list[dict], clear_first: bool = False):
    # 테이블 생성 (없으면)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with async_session() as db:  # type: AsyncSession
        if clear_first:
            await db.execute(delete(AdmissionData))
            await db.commit()
            print("기존 admission_data 삭제 완료")

        # 배치 삽입
        BATCH = 1000
        total = len(rows)
        for i in range(0, total, BATCH):
            batch = rows[i : i + BATCH]
            db.add_all([AdmissionData(**r) for r in batch])
            await db.commit()
            print(f"  삽입: {i + len(batch)} / {total}")

        # 요약
        year_result = await db.execute(
            text("SELECT year, COUNT(*) FROM admission_data GROUP BY year ORDER BY year DESC")
        )
        print("\n학년도별 저장 결과:")
        for year, cnt in year_result.all():
            print(f"  {year}학년도: {cnt}건")


def load_jeongsi_rows_from_xlsx(xlsx_path: str):
    """정시입결RAW 시트 로드. 헤더 6행, 데이터 8행부터."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "정시입결RAW" not in wb.sheetnames:
        print("'정시입결RAW' 시트가 없습니다. 정시 데이터 건너뜀.")
        return []

    ws = wb["정시입결RAW"]
    rows = []
    # 0:수준, 1:분야, 2:일반/지역, 3:구분2, 4:계열, 5:(빈), 6:대학, 7:모집단위명
    # 8:연도, 9:군, 10:최초인원, 11:이월인원, 12:최종인원, 13:지원자, 14:경쟁률
    # 15:추합, 16:충원율, 17:환산점수, 18:백분위70%, 19:대학발표백분위
    for row in ws.iter_rows(min_row=8, values_only=True):
        university = _parse_str(row[6]) if len(row) > 6 else None
        major = _parse_str(row[7]) if len(row) > 7 else None
        year = _parse_int(row[8]) if len(row) > 8 else None
        if not (university and major and year):
            continue
        rows.append({
            "tier": _parse_str(row[0]),
            "field": _parse_str(row[1]),
            "general_local": _parse_str(row[2]),
            "category2": _parse_str(row[3]),
            "track": _parse_str(row[4]),
            "university": university,
            "major": major,
            "year": year,
            "gun": _parse_str(row[9]) if len(row) > 9 else None,
            "initial_count": _parse_int(row[10]) if len(row) > 10 else None,
            "transfer_count": _parse_int(row[11]) if len(row) > 11 else None,
            "final_count": _parse_int(row[12]) if len(row) > 12 else None,
            "applicants": _parse_int(row[13]) if len(row) > 13 else None,
            "competition_rate": _parse_float(row[14]) if len(row) > 14 else None,
            "chu_hap": _parse_int(row[15]) if len(row) > 15 else None,
            "chung_won_rate": _parse_float(row[16]) if len(row) > 16 else None,
            "converted_score": _parse_float(row[17]) if len(row) > 17 else None,
            "percentile_70": _parse_float(row[18]) if len(row) > 18 else None,
            "univ_percentile": _parse_float(row[19]) if len(row) > 19 else None,
        })
    return rows


async def import_jeongsi_rows(rows: list[dict], clear_first: bool = False):
    """정시 입결 데이터를 DB에 임포트."""
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with async_session() as db:
        if clear_first:
            await db.execute(delete(JeongsiAdmissionData))
            await db.commit()
            print("기존 jeongsi_admission_data 삭제 완료")

        BATCH = 1000
        total = len(rows)
        for i in range(0, total, BATCH):
            batch = rows[i : i + BATCH]
            db.add_all([JeongsiAdmissionData(**r) for r in batch])
            await db.commit()
            print(f"  정시 삽입: {i + len(batch)} / {total}")

        year_result = await db.execute(
            text("SELECT year, COUNT(*) FROM jeongsi_admission_data GROUP BY year ORDER BY year DESC")
        )
        print("\n정시 학년도별 저장 결과:")
        for year, cnt in year_result.all():
            print(f"  {year}년: {cnt}건")


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", help="입결 xlsx 파일 경로")
    parser.add_argument("--clear", action="store_true", help="기존 데이터 삭제 후 임포트")
    args = parser.parse_args()

    print(f"파일 로드 중: {args.xlsx_path}")

    # 수시 입결
    rows = load_rows_from_xlsx(args.xlsx_path)
    print(f"수시 파싱 완료: {len(rows)}건")
    await import_rows(rows, clear_first=args.clear)

    # 정시 입결
    jeongsi_rows = load_jeongsi_rows_from_xlsx(args.xlsx_path)
    print(f"\n정시 파싱 완료: {len(jeongsi_rows)}건")
    if jeongsi_rows:
        await import_jeongsi_rows(jeongsi_rows, clear_first=args.clear)

    print("\n전체 임포트 완료!")


if __name__ == "__main__":
    asyncio.run(main())
