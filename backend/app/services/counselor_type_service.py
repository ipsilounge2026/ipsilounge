"""
C4 상담사 유형 판정 자동 생성 서비스

학생의 내신 평균등급 → 수시 입결 DB → 수시 가능 대학 티어
학생의 모의 백분위   → 정시 입결 DB → 정시 가능 대학 티어
두 티어 비교 → 내신형 / 균형형 / 수능형 자동 판정

데이터 소스 우선순위:
1. PostgreSQL DB (서버 배포용 - import_admission_data.py로 임포트)
2. Excel ���일 (로컬 개발용 - ADMISSION_DB_PATH 환경변수 또는 상대경로)
"""
from __future__ import annotations

import os
import functools
import logging
from typing import Any

logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


# ============================================================
# 대학 티어 정의 (높을수록 상위)
# ============================================================

TIER_ORDER = {
    "상위6": 6,
    "의치한약수": 5,
    "상위15": 5,
    "교대": 4,
    "지거국": 3,
    "2그룹": 2,
    "기타": 1,
}

TIER_LABELS = {
    6: "최상위권 (SKY급)",
    5: "상위권 (상위15 / 의약학)",
    4: "중상위권 (교대급)",
    3: "중위권 (지거국급)",
    2: "중하위권 (2그룹)",
    1: "기타",
}


# ============================================================
# 입결 DB 로드 (캐싱)
# ============================================================

def _get_admission_db_path() -> str:
    """admission_db.xlsx 경로."""
    # 환경변수 우선, 없으면 school-record-analyzer 상대경로
    env_path = os.environ.get("ADMISSION_DB_PATH")
    if env_path and os.path.exists(env_path):
        return env_path

    # 프로젝트 상위 디렉토리에서 탐색
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    candidates = [
        os.path.join(base, "data", "admission_db.xlsx"),
        os.path.join(base, "..", "school-record-analyzer", "data", "admission_db.xlsx"),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c

    return ""


@functools.lru_cache(maxsize=1)
def _load_susi_db() -> list[dict]:
    """수시 입결 데이터 로드. DB 우선, Excel 폴백."""
    # 1) DB 시도
    try:
        records = _load_susi_from_db()
        if records:
            logger.info("수시 입결 DB 로드: %d건 (PostgreSQL)", len(records))
            return records
    except Exception as e:
        logger.debug("���시 DB 로드 실패, Excel 폴백: %s", e)

    # 2) Excel 폴백
    return _load_susi_from_excel()


def _load_susi_from_db() -> list[dict]:
    """PostgreSQL에서 수시 입결 로드 (동기)."""
    from sqlalchemy import create_engine, text
    from app.config import settings
    sync_url = settings.DATABASE_URL.replace("postgresql+asyncpg", "postgresql+psycopg2").replace("postgresql://", "postgresql+psycopg2://") if "asyncpg" in settings.DATABASE_URL else settings.DATABASE_URL
    engine = create_engine(sync_url)
    with engine.connect() as conn:
        result = conn.execute(text(
            "SELECT university, admission_category, admission_name, major, year, result_70 "
            "FROM admission_data WHERE result_70 IS NOT NULL AND result_70 > 0"
        ))
        return [
            {"univ": r[0], "type": r[1] or "", "name": r[2] or "", "dept": r[3], "year": r[4], "cut70": float(r[5])}
            for r in result
        ]


def _load_susi_from_excel() -> list[dict]:
    """Excel에서 수시 입결 로드."""
    if openpyxl is None:
        return []
    path = _get_admission_db_path()
    if not path:
        return []

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "수시입결RAW" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["수시입결RAW"]
    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        cut70 = _parse_num(row[10])
        if cut70 is None or cut70 <= 0:
            continue
        records.append({
            "univ": str(row[0] or "").strip(),
            "type": str(row[1] or "").strip(),
            "name": str(row[2] or "").strip(),
            "dept": str(row[3] or "").strip(),
            "year": _parse_int(row[4]),
            "cut70": cut70,
        })
    wb.close()
    logger.info("수시 입결 Excel 로드: %d건", len(records))
    return records


@functools.lru_cache(maxsize=1)
def _load_jeongsi_db() -> list[dict]:
    """정시 입결 데이터 로드. DB 우선, Excel 폴백."""
    # 1) DB 시도
    try:
        records = _load_jeongsi_from_db()
        if records:
            logger.info("정시 입결 DB 로드: %d건 (PostgreSQL)", len(records))
            return records
    except Exception as e:
        logger.debug("정시 DB 로드 실패, Excel 폴백: %s", e)

    # 2) Excel 폴백
    return _load_jeongsi_from_excel()


def _load_jeongsi_from_db() -> list[dict]:
    """PostgreSQL에서 정시 입결 로드 (동기)."""
    from sqlalchemy import create_engine, text
    from app.config import settings
    sync_url = settings.DATABASE_URL.replace("postgresql+asyncpg", "postgresql+psycopg2").replace("postgresql://", "postgresql+psycopg2://") if "asyncpg" in settings.DATABASE_URL else settings.DATABASE_URL
    engine = create_engine(sync_url)
    with engine.connect() as conn:
        result = conn.execute(text(
            "SELECT tier, field, track, university, major, year, percentile_70, converted_score, univ_percentile "
            "FROM jeongsi_admission_data"
        ))
        return [
            {
                "tier": r[0] or "", "field": r[1] or "", "track": r[2] or "",
                "univ": r[3], "dept": r[4], "year": r[5],
                "pct70": float(r[6]) if r[6] else None,
                "conv_score": float(r[7]) if r[7] else None,
                "univ_pct": float(r[8]) if r[8] else None,
            }
            for r in result
        ]


def _load_jeongsi_from_excel() -> list[dict]:
    """Excel에서 정시 입결 로드."""
    if openpyxl is None:
        return []
    path = _get_admission_db_path()
    if not path:
        return []

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "정시입결RAW" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["정시입결RAW"]
    records = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        pct70 = _parse_num(row[18])
        tier_raw = str(row[0] or "").strip()
        univ = str(row[6] or "").strip()
        if not univ:
            continue
        records.append({
            "tier": tier_raw,
            "field": str(row[1] or "").strip(),
            "track": str(row[4] or "").strip(),
            "univ": univ,
            "dept": str(row[7] or "").strip(),
            "year": _parse_int(row[8]),
            "pct70": pct70,
            "conv_score": _parse_num(row[17]),
            "univ_pct": _parse_num(row[19]),
        })
    wb.close()
    logger.info("정시 입결 Excel 로드: %d건", len(records))
    return records


@functools.lru_cache(maxsize=1)
def _build_univ_tier_map() -> dict[str, int]:
    """대학→티어 매핑 (정시 DB의 수준 컬럼 기반)."""
    jeongsi = _load_jeongsi_db()
    univ_tiers: dict[str, list[int]] = {}
    for r in jeongsi:
        tier_val = TIER_ORDER.get(r["tier"], 1)
        univ_tiers.setdefault(r["univ"], []).append(tier_val)
    # 대학별 최빈 티어 (또는 최고 티어)
    return {u: max(vals) for u, vals in univ_tiers.items()}


# ============================================================
# C4 유형 판정 핵심 로직
# ============================================================

def determine_counselor_type(answers: dict) -> dict:
    """
    학생 답변에서 내신/모의 데이터를 추출하여
    수시/정시 가능 대학 라인을 비교하고 유형을 판정.

    Returns:
        {
            "type": "내신형" | "균형형" | "수능형",
            "recommended_admission": "수시" | "정시",
            "naesin_avg": float,
            "mock_pct_avg": float,
            "susi_reachable_tier": int,
            "susi_reachable_label": str,
            "susi_sample_univs": list[str],
            "jeongsi_reachable_tier": int,
            "jeongsi_reachable_label": str,
            "jeongsi_sample_univs": list[str],
            "comparison_table": list[dict],
            "reasoning": str,
        }
    """
    # 1. 내신 평균등급 산출
    naesin_avg = _extract_naesin_avg(answers)

    # 2. 모의 백분위 평균 산출
    mock_pct_avg = _extract_mock_pct_avg(answers)

    # 3. 수시 가능 대학 라인 산출
    susi_result = _calc_susi_reachable(naesin_avg)

    # 4. 정시 가능 대학 라인 산출
    jeongsi_result = _calc_jeongsi_reachable(mock_pct_avg)

    # 5. 비교 & 판정
    susi_tier = susi_result["tier"]
    jeongsi_tier = jeongsi_result["tier"]

    if susi_tier - jeongsi_tier >= 1:
        c4_type = "내신형"
        recommended = "수시"
    elif jeongsi_tier - susi_tier >= 1:
        c4_type = "수능형"
        recommended = "정시"
    else:
        c4_type = "균형형"
        # 균형형일 때 내신이 더 나으면 수시, 아니면 정시
        recommended = "수시" if susi_tier >= jeongsi_tier else "정시"

    # 6. 과목별 비교 테이블 생성
    comparison_table = _build_comparison_table(answers)

    # 7. 판정 근거 텍스트 생성
    reasoning = _generate_reasoning(
        c4_type, recommended, naesin_avg, mock_pct_avg,
        susi_result, jeongsi_result,
    )

    return {
        "type": c4_type,
        "recommended_admission": recommended,
        "naesin_avg": naesin_avg,
        "mock_pct_avg": mock_pct_avg,
        "susi_reachable_tier": susi_tier,
        "susi_reachable_label": susi_result["label"],
        "susi_sample_univs": susi_result["sample_univs"],
        "jeongsi_reachable_tier": jeongsi_tier,
        "jeongsi_reachable_label": jeongsi_result["label"],
        "jeongsi_sample_univs": jeongsi_result["sample_univs"],
        "comparison_table": comparison_table,
        "reasoning": reasoning,
    }


# ============================================================
# 내부 헬퍼 — 학생 데이터 추출
# ============================================================

def _extract_naesin_avg(answers: dict) -> float:
    """B 카테고리에서 최근 학기 전과목 평균등급 산출."""
    cat_b = answers.get("B", {})
    semesters = ["B4", "B3", "B2", "B1"]  # 최근 학기 우선

    for sem_key in semesters:
        sem_data = cat_b.get(sem_key)
        if not sem_data or not isinstance(sem_data, dict):
            continue
        grades = []
        for subj_key, subj_data in sem_data.items():
            if not isinstance(subj_data, dict):
                continue
            g = _parse_num(subj_data.get("rank_grade"))
            if g is not None and 1 <= g <= 5:
                grades.append(g)
        if grades:
            return round(sum(grades) / len(grades), 2)
    return 3.0  # 데이터 없으면 중간값


def _extract_mock_pct_avg(answers: dict) -> float:
    """C1에서 최근 회차 국수탐 백분위 평균 산출."""
    cat_c = answers.get("C", {})
    mock_data = cat_c.get("C1")
    if not mock_data or not isinstance(mock_data, dict):
        return 50.0

    areas = ["korean", "math", "inquiry1", "inquiry2"]
    sorted_sessions = sorted(mock_data.items(), key=lambda x: x[0])

    # 최근 회차
    for session_key, session in reversed(sorted_sessions):
        if not isinstance(session, dict):
            continue
        pctiles = []
        for area in areas:
            area_data = session.get(area, {})
            if isinstance(area_data, dict):
                p = _parse_num(area_data.get("percentile"))
                if p is not None:
                    pctiles.append(p)
        if pctiles:
            return round(sum(pctiles) / len(pctiles), 1)
    return 50.0


# ============================================================
# 내부 헬퍼 — 입결 비교
# ============================================================

def _calc_susi_reachable(naesin_avg: float) -> dict:
    """내신 등급으로 수시 입결 매칭 → 가능 대학 티어 산출."""
    susi_db = _load_susi_db()
    univ_tier_map = _build_univ_tier_map()

    if not susi_db:
        return {"tier": 0, "label": "데이터 없음", "sample_univs": []}

    # 최신 2개년 필터
    years = sorted(set(r["year"] for r in susi_db if r["year"]), reverse=True)
    recent_years = years[:2] if years else []
    filtered = [r for r in susi_db if r["year"] in recent_years] if recent_years else susi_db

    # 내신 avg <= 입결70% 인 대학 = 학생이 경쟁 가능한 곳
    # 내신은 낮을수록 좋으므로, 학생 등급 <= 입결컷 이면 가능
    reachable = [r for r in filtered if r["cut70"] >= naesin_avg]

    if not reachable:
        # 완화: 0.3 등급 여유
        reachable = [r for r in filtered if r["cut70"] >= naesin_avg - 0.3]

    if not reachable:
        return {"tier": 1, "label": TIER_LABELS[1], "sample_univs": []}

    # 가능 대학별 티어 산출
    reachable_univs = set(r["univ"] for r in reachable)
    tiers = [univ_tier_map.get(u, 1) for u in reachable_univs]

    # 상위 20% 대학의 평균 티어 (낙관적 + 현실적 중간)
    tiers_sorted = sorted(tiers, reverse=True)
    top_count = max(1, len(tiers_sorted) // 5)
    representative_tier = round(sum(tiers_sorted[:top_count]) / top_count)

    # 해당 티어 대표 대학 3개
    sample = [u for u in reachable_univs if univ_tier_map.get(u, 1) >= representative_tier][:5]

    return {
        "tier": representative_tier,
        "label": TIER_LABELS.get(representative_tier, "기타"),
        "sample_univs": sorted(sample),
    }


def _calc_jeongsi_reachable(mock_pct_avg: float) -> dict:
    """모의 백분위로 정시 입결 매칭 → 가능 대학 티어 산출."""
    jeongsi_db = _load_jeongsi_db()

    if not jeongsi_db:
        return {"tier": 0, "label": "데이터 없음", "sample_univs": []}

    # 최신 2개년 + 백분위70% 데이터 있는 것만
    years = sorted(set(r["year"] for r in jeongsi_db if r["year"]), reverse=True)
    recent_years = years[:2] if years else []
    filtered = [
        r for r in jeongsi_db
        if r["year"] in recent_years and r["pct70"] is not None and r["pct70"] > 0
    ] if recent_years else [
        r for r in jeongsi_db if r["pct70"] is not None and r["pct70"] > 0
    ]

    # 학생 백분위 >= 입결 백분위70% 이면 가능
    reachable = [r for r in filtered if mock_pct_avg >= r["pct70"]]

    if not reachable:
        # 완화: 2% 여유
        reachable = [r for r in filtered if mock_pct_avg >= r["pct70"] - 2]

    if not reachable:
        return {"tier": 1, "label": TIER_LABELS[1], "sample_univs": []}

    # 가능 대학별 티어
    tier_vals = [TIER_ORDER.get(r["tier"], 1) for r in reachable]
    tiers_sorted = sorted(tier_vals, reverse=True)
    top_count = max(1, len(tiers_sorted) // 5)
    representative_tier = round(sum(tiers_sorted[:top_count]) / top_count)

    # 대표 대학
    sample_univs = set()
    for r in reachable:
        if TIER_ORDER.get(r["tier"], 1) >= representative_tier:
            sample_univs.add(r["univ"])
    sample = sorted(sample_univs)[:5]

    return {
        "tier": representative_tier,
        "label": TIER_LABELS.get(representative_tier, "기타"),
        "sample_univs": sample,
    }


# ============================================================
# 내부 헬퍼 — 비교 테이블 & 판정 근거
# ============================================================

def _build_comparison_table(answers: dict) -> list[dict]:
    """과목별 내신 vs 모의고사 비교 테이블."""
    cat_b = answers.get("B", {})
    cat_c = answers.get("C", {})
    mock_data = cat_c.get("C1", {})

    subject_map = {
        "ko": {"name": "국어", "mock_key": "korean"},
        "ma": {"name": "수학", "mock_key": "math"},
        "en": {"name": "영어", "mock_key": "english"},
        "sc1": {"name": "탐구1", "mock_key": "inquiry1"},
        "sc2": {"name": "탐구2", "mock_key": "inquiry2"},
    }

    # 최근 내신
    naesin_grades = {}
    for sem_key in ["B4", "B3", "B2", "B1"]:
        sem_data = cat_b.get(sem_key)
        if not sem_data or not isinstance(sem_data, dict):
            continue
        for subj_key, subj_data in sem_data.items():
            if not isinstance(subj_data, dict):
                continue
            if subj_key not in naesin_grades:
                g = _parse_num(subj_data.get("rank_grade"))
                if g is not None:
                    naesin_grades[subj_key] = g
        if naesin_grades:
            break

    # 최근 모의
    mock_scores = {}
    sorted_sessions = sorted(mock_data.items(), key=lambda x: x[0]) if isinstance(mock_data, dict) else []
    for session_key, session in reversed(sorted_sessions):
        if not isinstance(session, dict):
            continue
        for area_key, area_data in session.items():
            if not isinstance(area_data, dict):
                continue
            if area_key not in mock_scores:
                mock_scores[area_key] = {
                    "rank": area_data.get("rank"),
                    "percentile": area_data.get("percentile"),
                    "raw_score": area_data.get("raw_score"),
                }
        if mock_scores:
            break

    table = []
    for subj_code, info in subject_map.items():
        row = {
            "subject": info["name"],
            "naesin_grade": naesin_grades.get(subj_code),
            "mock_rank": None,
            "mock_percentile": None,
            "mock_raw_score": None,
        }
        ms = mock_scores.get(info["mock_key"], {})
        if ms:
            row["mock_rank"] = ms.get("rank")
            row["mock_percentile"] = ms.get("percentile")
            row["mock_raw_score"] = ms.get("raw_score")
        table.append(row)

    return table


def _generate_reasoning(
    c4_type: str,
    recommended: str,
    naesin_avg: float,
    mock_pct_avg: float,
    susi_result: dict,
    jeongsi_result: dict,
) -> str:
    """판정 근거 자동 생성 텍스트."""
    lines = []

    lines.append(f"[내신 분석] 최근 학기 평균등급 {naesin_avg}등급으로, "
                 f"수시 기준 {susi_result['label']} 수준의 대학에 지원 가능합니다.")
    if susi_result["sample_univs"]:
        lines.append(f"  - 수시 가능 대학 예시: {', '.join(susi_result['sample_univs'][:3])}")

    lines.append(f"\n[모의고사 분석] 최근 국수탐 백분위 평균 {mock_pct_avg}%로, "
                 f"정시 기준 {jeongsi_result['label']} 수준의 대학에 지원 가능합니다.")
    if jeongsi_result["sample_univs"]:
        lines.append(f"  - 정시 가능 대학 예시: {', '.join(jeongsi_result['sample_univs'][:3])}")

    tier_diff = susi_result["tier"] - jeongsi_result["tier"]
    lines.append("\n[유형 판정]")

    if c4_type == "내신형":
        lines.append(
            f"수시 가능 대학 라인({susi_result['label']})이 "
            f"정시 가능 라인({jeongsi_result['label']})보다 높아 '내신형'으로 판정되었습니다. "
            f"내신 경쟁력을 활용한 수시 전형이 유리합니다."
        )
    elif c4_type == "수능형":
        lines.append(
            f"정시 가능 대학 라인({jeongsi_result['label']})이 "
            f"수시 가능 라인({susi_result['label']})보다 높아 '수능형'으로 판정되었습니다. "
            f"모의고사 성적을 활용한 정시 전형이 유리합니다."
        )
    else:
        lines.append(
            f"수시({susi_result['label']})와 정시({jeongsi_result['label']}) "
            f"가능 대학 라인이 유사하여 '균형형'으로 판정되었습니다. "
            f"수시·정시 병행 전략을 권장합니다."
        )

    lines.append(f"\n→ 추천 전형 방향: {recommended}")

    return "\n".join(lines)


# ============================================================
# 유틸리티
# ============================================================

def _parse_num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _parse_int(v: Any) -> int:
    if v is None:
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0
