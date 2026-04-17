"""
권장 이수 과목 매칭 서비스

- course_requirements.xlsx (Wide+쉼표 구분 정규화 포맷) 로딩 및 캐싱
- 학생의 목표 대학/학과(E2) + 이수/수강예정 과목(B1_B4, E5) 기반 매칭
- 결과: 핵심과목 이수현황 + 권장과목 이수현황

xlsx 스키마 (단일 시트: 권장과목):
    A: 대학
    B: 모집단위
    C: 핵심과목 (쉼표 구분)
    D: 권장과목 (쉼표 구분)
    E: 비고 (선택)

※ analyzer/modules/grade_analyzer.py 의 load_course_requirements() 도 동일 파일을
   동일 스키마로 로드한다 (통합 이식 후 ipsilounge/analyzer/data/ 단일 소스).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

from ..config import settings

logger = logging.getLogger(__name__)


# ── 파일 경로 ──

_DB_PATH = settings.DATA_ROOT / "course_requirements.xlsx"
_SHEET_NAME = "권장과목"


# ── 캐시 ──

_DB_CACHE: list[dict] | None = None


def _split_courses(cell_value: Any) -> list[str]:
    """쉼표 구분 셀 값을 과목 리스트로 변환."""
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text or text == "-":
        return []
    return [p.strip() for p in text.split(",") if p.strip()]


def _normalize(name: str) -> str:
    """과목명 정규화 — 비교용 (공백·특수문자 제거, 소문자)."""
    if not name:
        return ""
    return "".join(name.split()).lower()


def _load_db() -> list[dict]:
    """권장과목 DB 로드 (캐싱)."""
    global _DB_CACHE
    if _DB_CACHE is not None:
        return _DB_CACHE

    if not _DB_PATH.exists():
        logger.warning("권장과목 DB 파일 없음: %s", _DB_PATH)
        _DB_CACHE = []
        return _DB_CACHE

    try:
        wb = openpyxl.load_workbook(str(_DB_PATH), data_only=True, read_only=True)
        ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active
    except Exception as e:
        logger.error("권장과목 DB 로드 실패: %s", e)
        _DB_CACHE = []
        return _DB_CACHE

    rows: list[dict] = []
    for i, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # row_cells = (대학, 모집단위, 핵심과목, 권장과목, 비고, ...)
        if not row_cells or len(row_cells) < 2:
            continue
        univ = row_cells[0]
        major = row_cells[1] if len(row_cells) > 1 else None
        if not univ or not major:
            continue
        univ_s = str(univ).strip()
        major_s = str(major).strip()
        # 주석 행 스킵 (※ 로 시작)
        if univ_s.startswith("※"):
            continue
        core = _split_courses(row_cells[2] if len(row_cells) > 2 else None)
        recommended = _split_courses(row_cells[3] if len(row_cells) > 3 else None)
        note = str(row_cells[4]).strip() if len(row_cells) > 4 and row_cells[4] else ""
        rows.append(
            {
                "대학": univ_s,
                "모집단위": major_s,
                "핵심과목": core,
                "권장과목": recommended,
                "비고": note,
            }
        )

    _DB_CACHE = rows
    logger.info("권장과목 DB 로드 완료: %d개 모집단위", len(rows))
    return _DB_CACHE


def reset_cache() -> None:
    """DB 캐시 초기화 (xlsx 갱신 후 호출)."""
    global _DB_CACHE
    _DB_CACHE = None


# ── 목표 대학/학과 파싱 ──


def _parse_target(target: str) -> tuple[str, str] | None:
    """E2 target_universities 항목("서울대 컴퓨터공학부")을 (대학, 모집단위)로 분리.

    공백 기준 첫 토큰을 대학, 나머지를 모집단위로 처리.
    모집단위가 없으면 None 반환 (매칭 불가).
    """
    if not target or not isinstance(target, str):
        return None
    parts = target.strip().split(None, 1)
    if len(parts) < 2:
        return None
    return parts[0].strip(), parts[1].strip()


def _lookup(db: list[dict], univ: str, major: str) -> dict | None:
    """대학+모집단위로 DB 행 조회. 정확 매칭 우선, 실패 시 부분 매칭."""
    univ_n = _normalize(univ)
    major_n = _normalize(major)

    # 정확 매칭
    for row in db:
        if _normalize(row["대학"]) == univ_n and _normalize(row["모집단위"]) == major_n:
            return row

    # 대학 정확 + 모집단위 부분 매칭 (예: "경제학부" vs "경제학")
    for row in db:
        if _normalize(row["대학"]) == univ_n:
            r_major = _normalize(row["모집단위"])
            if major_n in r_major or r_major in major_n:
                return row

    return None


# ── 학생 이수 과목 추출 ──


def _extract_student_courses(answers: dict) -> dict[str, list[str]]:
    """설문 답변에서 이수/예정 과목 리스트 추출.

    Returns:
        {
            "completed": [...],  # B1_B4 에서 추출된 이수 과목명
            "planned": [...],    # E5.next_courses 수강 예정 과목
        }
    """
    completed: list[str] = []
    planned: list[str] = []

    # B1_B4: 학기별 과목 성적 행
    cat_b = answers.get("B", {}) if isinstance(answers, dict) else {}
    b_data = cat_b.get("B1_B4", []) if isinstance(cat_b, dict) else []
    if isinstance(b_data, list):
        for row in b_data:
            if not isinstance(row, dict):
                continue
            name = row.get("course_name") or row.get("과목명")
            if name:
                completed.append(str(name).strip())

    # 일부 스키마에서는 학기 키 아래 리스트인 경우도 있음
    if not completed and isinstance(b_data, dict):
        for sem_key, sem_rows in b_data.items():
            if isinstance(sem_rows, list):
                for row in sem_rows:
                    if isinstance(row, dict):
                        name = row.get("course_name") or row.get("과목명")
                        if name:
                            completed.append(str(name).strip())

    # E5.next_courses
    cat_e = answers.get("E", {}) if isinstance(answers, dict) else {}
    e5 = cat_e.get("E5", {}) if isinstance(cat_e, dict) else {}
    next_list = e5.get("next_courses", []) if isinstance(e5, dict) else []
    if isinstance(next_list, list):
        for nc in next_list:
            if nc:
                planned.append(str(nc).strip())

    return {"completed": completed, "planned": planned}


# ── 매칭 ──


def _match_courses(
    required: list[str], student_all: set[str]
) -> tuple[list[str], list[str]]:
    """DB 요구 과목 리스트 vs 학생 과목 집합 매칭.

    과목명 정규화 후 교집합/차집합 계산.
    Returns: (이수완료, 미이수)
    """
    student_norm = {_normalize(c): c for c in student_all}
    matched: list[str] = []
    missing: list[str] = []
    for req in required:
        req_n = _normalize(req)
        # 정확 매칭 또는 포함 매칭 (예: "미적분" ⊂ "미적분1")
        hit = False
        for s_n in student_norm:
            if s_n == req_n or req_n in s_n or s_n in req_n:
                hit = True
                break
        if hit:
            matched.append(req)
        else:
            missing.append(req)
    return matched, missing


def build_matching(answers: dict) -> dict:
    """설문 답변 기반 권장과목 이수 현황 생성.

    Returns:
        {
            "available": bool,      # 매칭 수행 가능 여부
            "reason": str | None,   # available=False 일 때 사유
            "results": [            # 목표별 매칭 결과
                {
                    "target": "서울대 컴퓨터공학부",
                    "대학": "서울대",
                    "모집단위": "컴퓨터공학부",
                    "found": bool,              # DB 내 데이터 존재 여부
                    "핵심_이수": [...],
                    "핵심_미이수": [...],
                    "권장_이수": [...],
                    "권장_미이수": [...],
                    "비고": str,
                    "학생_과목": {"completed": [...], "planned": [...]},
                }
            ]
        }
    """
    cat_e = answers.get("E", {}) if isinstance(answers, dict) else {}
    e2 = cat_e.get("E2", {}) if isinstance(cat_e, dict) else {}
    targets = e2.get("target_universities", []) if isinstance(e2, dict) else []
    if not isinstance(targets, list):
        targets = []
    targets = [t for t in targets if t and isinstance(t, str) and t.strip()]

    if not targets:
        return {
            "available": False,
            "reason": "학생이 목표 대학/학과를 입력하지 않아 권장과목 매칭이 불가합니다.",
            "results": [],
        }

    db = _load_db()
    if not db:
        return {
            "available": False,
            "reason": "권장과목 DB(course_requirements_v2.xlsx)가 비어 있거나 존재하지 않습니다.",
            "results": [],
        }

    student = _extract_student_courses(answers)
    student_all = set(student["completed"]) | set(student["planned"])

    results: list[dict] = []
    for target in targets:
        parsed = _parse_target(target)
        if not parsed:
            results.append(
                {
                    "target": target,
                    "대학": None,
                    "모집단위": None,
                    "found": False,
                    "reason": "대학·모집단위 형식이 아님 (예: '서울대 컴퓨터공학부')",
                    "핵심_이수": [],
                    "핵심_미이수": [],
                    "권장_이수": [],
                    "권장_미이수": [],
                    "비고": "",
                    "학생_과목": student,
                }
            )
            continue

        univ, major = parsed
        row = _lookup(db, univ, major)
        if row is None:
            results.append(
                {
                    "target": target,
                    "대학": univ,
                    "모집단위": major,
                    "found": False,
                    "reason": "해당 대학·모집단위가 권장과목 DB에 등록되지 않았습니다",
                    "핵심_이수": [],
                    "핵심_미이수": [],
                    "권장_이수": [],
                    "권장_미이수": [],
                    "비고": "",
                    "학생_과목": student,
                }
            )
            continue

        core_matched, core_missing = _match_courses(row["핵심과목"], student_all)
        rec_matched, rec_missing = _match_courses(row["권장과목"], student_all)

        results.append(
            {
                "target": target,
                "대학": row["대학"],
                "모집단위": row["모집단위"],
                "found": True,
                "핵심_이수": core_matched,
                "핵심_미이수": core_missing,
                "권장_이수": rec_matched,
                "권장_미이수": rec_missing,
                "비고": row.get("비고", ""),
                "학생_과목": student,
            }
        )

    return {"available": True, "reason": None, "results": results}
