"""대학어디가 자동 수집 입결 Excel → DB import 서비스.

지원 형식 (자동 감지):
1. 표준 형식 (legacy): 시트 '전년도입결' 1개, 37컬럼 고정 순서
2. 정규화 3시트 형식: 시트 '입결_종합'/'입결_교과'/'입결_수능',
   컬럼 = 대학명/대학코드/전형유형/전형명/모집단위/구분/모집인원/경쟁률/충원인원
   + (종합·교과) 학생부 환산점수·등급 50/70% + (수능) 환산점수 50/70% + 백분위 50/70%
3. 원시 3시트 형식: 같은 시트명이지만 대학별 자체발표 표를 그대로 합친 형태
   (컬럼명이 대학마다 제각각, 경쟁률 '3.75:1' 텍스트, 종합/교과 시트에 구분 없음)
   → 헤더 패턴 + 값 범위(등급 ≤ 9.5 / 점수 > 9.5)로 best-effort 추출

운영 정책:
- 같은 학년도 재업로드 시 → 해당 학년도 row 전부 삭제 후 새로 INSERT
- 다른 학년도는 영향 없음
- 트랜잭션 처리 (중간 실패 시 롤백)

용도:
- CLI 스크립트(import_adiga_results.py) 에서 호출
- 관리자 업로드 API (admin_adiga_import.py) 에서 호출
"""

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.adiga_admission_result import AdigaAdmissionResult

logger = logging.getLogger(__name__)

# ── 형식 1: 표준 (legacy) ──────────────────────────────────────────

# Excel 시트명
SHEET_NAME = "전년도입결"

# Excel 컬럼 인덱스 (0-based)
COL_UNIVERSITY = 0
COL_UNIVERSITY_CODE = 1
COL_ADMISSION_CATEGORY = 2
COL_ADMISSION_NAME = 3
COL_RECRUITMENT_TYPE = 4
COL_MAJOR = 5
COL_RECRUIT_COUNT = 6
COL_COMPETITION_RATE = 7
COL_ADDITIONAL_COUNT = 8
COL_GPA_SCORE_50 = 9
COL_GPA_SCORE_70 = 10
COL_GPA_GRADE_50 = 11
COL_GPA_GRADE_70 = 12
COL_CONV_SCORE_50 = 13
COL_CONV_SCORE_70 = 14
# 백분위 50%: 15~25 (11개)
# 백분위 70%: 26~36 (11개)
PERCENTILE_50_RANGE = range(15, 26)
PERCENTILE_70_RANGE = range(26, 37)

# 백분위 컬럼 키 (헤더 텍스트 → JSON 키 매핑용)
# Excel 헤더 예: "백분위 50% 국어", "백분위 70% 평균백분위" 등
PERCENTILE_KEY_MAP = {
    "국어": "korean",
    "수학": "math",
    "탐구1 사탐": "investigation1_social",
    "탐구1 과탐": "investigation1_science",
    "탐구1 직탐": "investigation1_vocational",
    "탐구2 사탐": "investigation2_social",
    "탐구2 과탐": "investigation2_science",
    "탐구2 직탐": "investigation2_vocational",
    "평균백분위": "average_percentile",
    "한국사(등급)": "korean_history_grade",
    "영어등급": "english_grade",
}

# 표준 형식 헤더 (앞 9개 컬럼, 공백 제거 후 비교)
EXPECTED_HEADERS = ["대학명", "대학코드", "전형유형", "전형명", "구분", "모집단위", "모집인원", "경쟁률", "충원인원"]

# ── 형식 2/3: 3시트 (입결_종합/입결_교과/입결_수능) ──────────────────

UNIFIED_SHEET_PREFIX = "입결_"

# 정규화 형식의 점수·등급 컬럼 (공백 제거 후 정확 일치)
# 고정 매핑: 헤더가 학생부/등급을 명시 → 필드 확정
CLEAN_FIXED_HEADERS = {
    "학생부환산점수50%": "gpa_score_50",
    "학생부환산점수70%": "gpa_score_70",
    "학생부환산등급50%": "gpa_grade_50",
    "학생부환산등급70%": "gpa_grade_70",
    "등급50": "gpa_grade_50",
    "등급70": "gpa_grade_70",
}
# 범용 환산점수: 행의 구분에 따라 수시 → gpa_score, 정시 → conv_score 로 배정
CLEAN_GENERIC_SCORE_HEADERS = {
    "환산점수50%": "50",
    "환산점수50": "50",
    "환산점수70%": "70",
    "환산점수70": "70",
}
# 비고 컬럼
NOTE_HEADERS = {"기타", "비고"}

# 정리 형식 백분위: '백분위50_국어' 식 (공백 제거 후) suffix → JSON 키
NEW_PCT_SUFFIX_MAP = {
    "국어": "korean",
    "수학": "math",
    "탐구1": "investigation1",
    "탐구2": "investigation2",
    "영어": "english_grade",
    "한국사": "korean_history_grade",
    "평균": "average_percentile",
}

# 원시 형식 수능 백분위: '... 50% ... 과목별 백분위 ... 국' 식 헤더의 끝글자 → JSON 키
RAW_PCT_SUFFIX_MAP = {
    "국": "korean",
    "수": "math",
    "탐1": "investigation1",
    "탐2": "investigation2",
    "영": "english_grade",
    "한": "korean_history_grade",
}


# 파일명에서 학년도 추출: "adiga_입결_2027.xlsx" → 2027
RE_YEAR_FROM_FILENAME = re.compile(r"(\d{4})\.xlsx?$", re.IGNORECASE)

# 경쟁률 텍스트: "3.75:1" / "5.25 : 1"
RE_RATE_TEXT = re.compile(r"^\s*([\d.]+)\s*:\s*1\s*$")


def extract_year_from_filename(filename: str) -> int | None:
    """파일명에서 학년도 추출."""
    m = RE_YEAR_FROM_FILENAME.search(filename)
    return int(m.group(1)) if m else None


def _validate_headers(headers: list, sheetnames: list[str]) -> None:
    """표준(legacy) 형식 컬럼 구조 검증. 다르면 명확한 메시지로 ValueError.

    형식이 다른 파일을 그대로 import 하면 컬럼이 밀려서
    학과명이 '구분'(VARCHAR(10)) 컬럼에 들어가는 등 DB 에러가 난다.
    """
    norm = [_norm_header(h) for h in headers[: len(EXPECTED_HEADERS)]]
    mismatches = [
        f"{i + 1}번째 컬럼: 기대 '{exp}' / 실제 '{norm[i] if i < len(norm) and norm[i] else '(없음)'}'"
        for i, exp in enumerate(EXPECTED_HEADERS)
        if i >= len(norm) or norm[i] != exp
    ]
    if mismatches:
        raise ValueError(
            "Excel 컬럼 구조가 표준 형식과 다릅니다. "
            f"(시트: {sheetnames}) "
            "표준 형식: 시트명 '전년도입결' 1개, 컬럼 순서 = "
            + " → ".join(EXPECTED_HEADERS)
            + " → 학생부 환산점수/등급/백분위... (총 37컬럼). "
            "불일치: " + " / ".join(mismatches[:5])
        )


def _norm_header(h: Any) -> str:
    """헤더 정규화: 공백 제거 (대학마다 '모집 인원'/'모집인원' 등 표기 차이 흡수)."""
    return str(h).replace(" ", "").strip() if h is not None else ""


def _to_float(v: Any) -> float | None:
    """셀 값을 float으로 변환. 텍스트("0.0", "미제출 사유 : 3명이하") 처리."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _to_rate(v: Any) -> float | None:
    """경쟁률 변환: 숫자 또는 '3.75:1' 텍스트."""
    f = _to_float(v)
    if f is not None:
        return f
    if isinstance(v, str):
        m = RE_RATE_TEXT.match(v.strip())
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                return None
    return None


def _to_int(v: Any) -> int | None:
    """셀 값을 int 로 변환."""
    if v is None or v == "":
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _build_percentile_dict(
    headers: list[str], row: tuple, col_range: range, prefix: str
) -> dict:
    """(legacy) 백분위 영역 컬럼을 JSON 객체로 변환.

    headers: Excel 첫 행 (헤더)
    row: 데이터 행
    col_range: 백분위 컬럼 인덱스 범위
    prefix: "백분위 50%" 또는 "백분위 70%"
    """
    result = {}
    for idx in col_range:
        if idx >= len(headers):
            break
        header = headers[idx]
        if not isinstance(header, str):
            continue
        # 헤더에서 prefix 제거 → 키 정규화
        key_raw = header.replace(prefix, "").strip()
        key = PERCENTILE_KEY_MAP.get(key_raw)
        if key is None:
            continue
        value = row[idx] if idx < len(row) else None
        # 등급 컬럼은 정수, 백분위는 실수
        if "grade" in key:
            result[key] = _to_int(value)
        else:
            result[key] = _to_float(value)
    return result


def _parse_legacy_rows(headers: list, rows_iter, year: int | None, filename: str) -> list[dict]:
    """표준 형식 (전년도입결 37컬럼) 행 파싱 — 기존 로직 그대로."""
    parsed: list[dict] = []
    for row in rows_iter:
        if row is None:
            continue
        # university 가 비어 있으면 무효 행
        univ = _to_str(row[COL_UNIVERSITY] if COL_UNIVERSITY < len(row) else None)
        if not univ:
            continue

        # 학생부 환산점수/등급 셀에 "0.0" 같이 0 인 케이스, 또는 텍스트(미제출 사유) 케이스 처리.
        # 텍스트가 들어가 있으면 note 로 보존.
        gpa_score_50_raw = row[COL_GPA_SCORE_50] if COL_GPA_SCORE_50 < len(row) else None
        gpa_score_70_raw = row[COL_GPA_SCORE_70] if COL_GPA_SCORE_70 < len(row) else None
        note_parts = []
        for v in (gpa_score_50_raw, gpa_score_70_raw):
            if isinstance(v, str) and _to_float(v) is None and v.strip():
                note_parts.append(v.strip())

        item = {
            "university": univ,
            "university_code": _to_str(row[COL_UNIVERSITY_CODE]) or "",
            "year": year or 0,  # 0 이면 import 시점에 덮어쓰기
            "admission_category": _to_str(row[COL_ADMISSION_CATEGORY]),
            "admission_name": _to_str(row[COL_ADMISSION_NAME]),
            "recruitment_type": _to_str(row[COL_RECRUITMENT_TYPE]),
            "major": _to_str(row[COL_MAJOR]) or "",
            "recruit_count": _to_int(row[COL_RECRUIT_COUNT]),
            "competition_rate": _to_float(row[COL_COMPETITION_RATE]),
            "additional_count": _to_int(row[COL_ADDITIONAL_COUNT]),
            "gpa_score_50": _to_float(gpa_score_50_raw),
            "gpa_score_70": _to_float(gpa_score_70_raw),
            "gpa_grade_50": _to_float(row[COL_GPA_GRADE_50] if COL_GPA_GRADE_50 < len(row) else None),
            "gpa_grade_70": _to_float(row[COL_GPA_GRADE_70] if COL_GPA_GRADE_70 < len(row) else None),
            "conv_score_50": _to_float(row[COL_CONV_SCORE_50] if COL_CONV_SCORE_50 < len(row) else None),
            "conv_score_70": _to_float(row[COL_CONV_SCORE_70] if COL_CONV_SCORE_70 < len(row) else None),
            "percentile_50": _build_percentile_dict(headers, row, PERCENTILE_50_RANGE, "백분위 50%"),
            "percentile_70": _build_percentile_dict(headers, row, PERCENTILE_70_RANGE, "백분위 70%"),
            "note": " | ".join(note_parts) if note_parts else None,
            "source_file": filename,
        }
        parsed.append(item)
    return parsed


# ── 3시트 형식 파싱 ─────────────────────────────────────────────────


def _normalize_recruitment_type(v: Any, admission_name: str | None, sheet_name: str) -> str:
    """구분 정규화: '수시' / '정시(가)' 형태로 통일.

    원시 형식의 "' 나 ' 군" → '정시(나)', 구분 컬럼이 없으면
    전형명 prefix('수시 ...'/'정시 ...') 또는 시트명(수능→정시)으로 유추.
    """
    s = _to_str(v)
    if s:
        s2 = re.sub(r"[\s'‘’\"]+", "", s)
        # 원문자 변형: ㉮군/㉯군/㉰군
        s2 = s2.translate(str.maketrans("㉮㉯㉰", "가나다"))
        if "수시" in s2:
            return "수시"
        m = re.search(r"정시\(([가나다])\)", s2)
        if m:
            return f"정시({m.group(1)})"
        m = re.search(r"([가나다])군", s2)
        if m:
            return f"정시({m.group(1)})"
        if s2 in ("가", "나", "다"):
            return f"정시({s2})"
        return s2[:10]
    name = admission_name or ""
    if name.startswith("수시"):
        return "수시"
    if name.startswith("정시"):
        return "정시"
    if "수능" in sheet_name:
        return "정시"
    return "수시"


def _strip_period_prefix(name: str | None) -> str | None:
    """전형명의 '수시 '/'정시 ' prefix 제거 (원시 형식: '수시 지역균형전형' → '지역균형전형')."""
    if not name:
        return name
    for p in ("수시 ", "정시 "):
        if name.startswith(p):
            stripped = name[len(p):].strip()
            return stripped or name
    return name


def _collect_core_candidates(headers_norm: list[str]) -> dict[str, list[int]]:
    """기본 필드별 후보 컬럼 인덱스 수집 (우선순위 순). 원시 형식의 표기 변형 흡수."""
    cand: dict[str, list[int]] = {
        "university": [], "university_code": [], "admission_category": [],
        "admission_name": [], "major": [], "recruitment_type": [],
        "recruit_count": [], "competition_rate": [], "additional_count": [],
    }
    for idx, h in enumerate(headers_norm):
        if not h:
            continue
        if h == "대학명":
            cand["university"].append(idx)
        elif h == "대학코드":
            cand["university_code"].append(idx)
        elif h == "전형유형":
            cand["admission_category"].append(idx)
        elif h == "전형명":
            cand["admission_name"].append(idx)
        elif h == "모집단위":
            cand["major"].append(idx)
        elif h == "구분":
            cand["recruitment_type"].append(idx)
        elif "모집인원" in h and "이월" not in h and "최초" not in h:
            # 정확 일치('모집인원') 또는 최종 인원이 최우선
            if h == "모집인원" or "최종" in h:
                cand["recruit_count"].insert(0, idx)
            else:
                cand["recruit_count"].append(idx)
        elif h in ("최종(A+B)", "모집"):
            cand["recruit_count"].append(idx)
        elif "경쟁률" in h:
            cand["competition_rate"].append(idx)
        elif "충원" in h:
            if h == "충원인원":
                cand["additional_count"].insert(0, idx)
            else:
                cand["additional_count"].append(idx)
    return cand


def _collect_cut_candidates(headers_norm: list[str], pct: str) -> list[int]:
    """50%/70% cut 점수·등급 후보 컬럼 수집 (원시 형식용).

    '최종등록자 교과성적 학생부등급 50% cut', '대학별환산 최종등록자 환산점수 50% cut',
    '50% cut' (bare) 등 대학별 표기 변형을 모두 포함.
    값 범위(≤9.5 → 등급 / >9.5 → 점수)로 행 단위 분류한다.
    """
    cols = []
    exclude_tokens = ("백분위", "분포도", "응시비율", "총점", "경쟁률", "모집", "충원", "지원", "등록률")
    for idx, h in enumerate(headers_norm):
        if not h or pct not in h:
            continue
        if any(t in h for t in exclude_tokens):
            continue
        if "80" in h or "90" in h:
            continue
        if "cut" in h.lower() or h.endswith(f"{pct}%") or h.endswith(pct):
            cols.append(idx)
    return cols


def _collect_pct_candidates(
    headers: list, headers_norm: list[str], pct: str
) -> dict[str, list[int]]:
    """백분위 컬럼 매핑: {json_key: [후보 인덱스...]}.

    3가지 헤더 가족 지원:
    1. '백분위 50% 국어' (구 정규화 형식)
    2. '백분위50_국어' (신 정리 형식)
    3. '... 50% ... 과목별 백분위 ... 국' (원시 형식)
    """
    cols: dict[str, list[int]] = {}
    prefix = f"백분위 {pct}%"
    # 1) 구 정규화 형식
    for idx, h in enumerate(headers):
        if not isinstance(h, str):
            continue
        hs = h.strip()
        if hs.startswith(prefix):
            key = PERCENTILE_KEY_MAP.get(hs.replace(prefix, "").strip())
            if key:
                cols.setdefault(key, []).append(idx)
    if cols:
        return cols
    # 2) 신 정리 형식: '백분위50_국어'
    new_prefix = f"백분위{pct}_"
    for idx, h in enumerate(headers_norm):
        if h.startswith(new_prefix):
            key = NEW_PCT_SUFFIX_MAP.get(h[len(new_prefix):])
            if key:
                cols.setdefault(key, []).append(idx)
    if cols:
        return cols
    # 3) 원시 형식
    for idx, h in enumerate(headers_norm):
        if f"{pct}%" in h and "과목별백분위" in h:
            for suffix, key in RAW_PCT_SUFFIX_MAP.items():
                if h.endswith(suffix):
                    cols.setdefault(key, []).append(idx)
                    break
    return cols


def _first_value(row: tuple, idxs: list[int], converter) -> Any:
    """후보 컬럼들 중 첫 번째 유효 값."""
    for idx in idxs:
        v = row[idx] if idx < len(row) else None
        conv = converter(v)
        if conv is not None:
            return conv
    return None


def _parse_unified_sheet(sheet, sheet_name: str, year: int | None, filename: str) -> list[dict]:
    """3시트 형식(정규화/원시 공통) 한 시트 파싱."""
    rows_iter = sheet.iter_rows(values_only=True)
    headers = list(next(rows_iter, ()))
    headers_norm = [_norm_header(h) for h in headers]

    core = _collect_core_candidates(headers_norm)
    required = {
        "university": "대학명", "university_code": "대학코드",
        "admission_category": "전형유형", "admission_name": "전형명", "major": "모집단위",
    }
    missing = [label for field, label in required.items() if not core[field]]
    if missing:
        raise ValueError(
            f"시트 '{sheet_name}' 에 필수 컬럼이 없습니다: {', '.join(missing)} "
            f"(발견된 헤더 앞부분: {[h for h in headers[:8]]})"
        )

    # 정규화 형식의 점수·등급 정확 일치 컬럼
    fixed_cols: dict[str, list[int]] = {}
    generic_score_cols: dict[str, list[int]] = {}  # "50"/"70" → 인덱스 (구분에 따라 gpa/conv 배정)
    note_cols: list[int] = []
    for idx, h in enumerate(headers_norm):
        field = CLEAN_FIXED_HEADERS.get(h)
        if field:
            fixed_cols.setdefault(field, []).append(idx)
            continue
        pct = CLEAN_GENERIC_SCORE_HEADERS.get(h)
        if pct:
            generic_score_cols.setdefault(pct, []).append(idx)
            continue
        if h in NOTE_HEADERS:
            note_cols.append(idx)

    # 원시 형식 fallback: 50/70 cut 후보 (정확 일치 컬럼이 전혀 없을 때만 사용)
    use_raw_cut = not fixed_cols and not generic_score_cols
    cut50 = _collect_cut_candidates(headers_norm, "50") if use_raw_cut else []
    cut70 = _collect_cut_candidates(headers_norm, "70") if use_raw_cut else []

    pct50 = _collect_pct_candidates(headers, headers_norm, "50")
    pct70 = _collect_pct_candidates(headers, headers_norm, "70")

    is_suneung_sheet = "수능" in sheet_name

    def _build_pct(row: tuple, colmap: dict[str, list[int]]) -> dict:
        out = {}
        for key, idxs in colmap.items():
            conv = _to_int if "grade" in key else _to_float
            out[key] = _first_value(row, idxs, conv)
        return out

    def _pct_text_note(row: tuple, colmap: dict[str, list[int]]) -> str | None:
        """백분위 컬럼에 숫자 대신 텍스트('미제출 사유: 3명이하 미산출' 등)가 있으면 반환."""
        for idxs in colmap.values():
            for idx in idxs:
                v = row[idx] if idx < len(row) else None
                if isinstance(v, str):
                    s = v.strip()
                    if s and s != "-" and _to_float(s) is None:
                        return s
        return None

    def _read_clean(row: tuple, idxs: list[int], note_parts: list[str]) -> float | None:
        """정확 일치 컬럼 읽기. 숫자가 아니면 텍스트를 note 로 보존."""
        val = _first_value(row, idxs, _to_float)
        if val is None:
            for idx in idxs:
                v = row[idx] if idx < len(row) else None
                if isinstance(v, str) and v.strip() and v.strip() != "-":
                    note_parts.append(v.strip())
                    break
        return val

    def _extract_cut(row: tuple, idxs: list[int]) -> tuple[float | None, float | None, list[str]]:
        """원시 형식: 후보 컬럼 값들을 등급(≤9.5)/점수(>9.5)로 분류."""
        grade = score = None
        texts: list[str] = []
        for idx in idxs:
            v = row[idx] if idx < len(row) else None
            if v is None or v == "" or v == "-":
                continue
            f = _to_float(v)
            if f is None:
                if isinstance(v, str) and v.strip() and v.strip() != "-":
                    texts.append(v.strip())
                continue
            if f <= 0:
                continue
            if f <= 9.5:
                if grade is None:
                    grade = f
            else:
                if score is None:
                    score = f
        return grade, score, texts

    parsed: list[dict] = []
    for row in rows_iter:
        if row is None:
            continue
        univ = _first_value(row, core["university"], _to_str)
        if not univ:
            continue

        raw_name = _first_value(row, core["admission_name"], _to_str)
        rt_raw = _first_value(row, core["recruitment_type"], _to_str) if core["recruitment_type"] else None
        recruitment_type = _normalize_recruitment_type(rt_raw, raw_name, sheet_name)
        admission_name = _strip_period_prefix(raw_name)

        is_jeongsi_row = recruitment_type.startswith("정시") or (
            recruitment_type not in ("수시",) and is_suneung_sheet
        )

        note_parts: list[str] = []
        values: dict[str, float | None] = {
            "gpa_score_50": None, "gpa_score_70": None,
            "gpa_grade_50": None, "gpa_grade_70": None,
            "conv_score_50": None, "conv_score_70": None,
        }

        if fixed_cols or generic_score_cols:
            # 정규화 형식: 정확 일치 컬럼 직접 읽기 (+텍스트 메모 보존)
            for field, idxs in fixed_cols.items():
                values[field] = _read_clean(row, idxs, note_parts)
            for pct, idxs in generic_score_cols.items():
                # 범용 환산점수: 정시 행 → conv_score, 수시 행 → gpa_score (학생부 환산)
                target = f"conv_score_{pct}" if is_jeongsi_row else f"gpa_score_{pct}"
                if values[target] is None:
                    values[target] = _read_clean(row, idxs, note_parts)
        if use_raw_cut:
            # 원시 형식: 값 범위로 등급/점수 분류
            g50, s50, t50 = _extract_cut(row, cut50)
            g70, s70, t70 = _extract_cut(row, cut70)
            note_parts.extend(t50)
            note_parts.extend(t70)
            values["gpa_grade_50"], values["gpa_grade_70"] = g50, g70
            if is_jeongsi_row:
                values["conv_score_50"], values["conv_score_70"] = s50, s70
            else:
                values["gpa_score_50"], values["gpa_score_70"] = s50, s70

        # 비고('기타') 컬럼 보존
        for idx in note_cols:
            v = row[idx] if idx < len(row) else None
            s = _to_str(v)
            if s and s != "-":
                note_parts.append(s)

        # 백분위 컬럼의 텍스트(미제출 사유 등) 보존
        pct50_dict = _build_pct(row, pct50)
        pct70_dict = _build_pct(row, pct70)
        if pct50 and not any(v is not None for v in pct50_dict.values()):
            t = _pct_text_note(row, pct50)
            if t:
                note_parts.append(t)

        # note 중복 제거 (순서 보존)
        seen: set[str] = set()
        notes = [n for n in note_parts if not (n in seen or seen.add(n))]

        item = {
            "university": univ,
            "university_code": _first_value(row, core["university_code"], _to_str) or "",
            "year": year or 0,
            "admission_category": _first_value(row, core["admission_category"], _to_str),
            "admission_name": admission_name,
            "recruitment_type": recruitment_type,
            "major": _first_value(row, core["major"], _to_str) or "",
            "recruit_count": _first_value(row, core["recruit_count"], _to_int),
            "competition_rate": _first_value(row, core["competition_rate"], _to_rate),
            "additional_count": _first_value(row, core["additional_count"], _to_int),
            **values,
            "percentile_50": pct50_dict,
            "percentile_70": pct70_dict,
            "note": " | ".join(notes) if notes else None,
            "source_file": filename,
        }
        parsed.append(item)
    return parsed


def parse_excel(file_path: str | Path) -> dict:
    """
    Excel 파일을 읽어 (year, rows) 반환. 형식 자동 감지.

    Returns:
        {
            "year": int | None,
            "filename": str,
            "format": "legacy" | "unified",
            "headers": list[str],
            "rows": [{...}],  # AdigaAdmissionResult kwargs
            "total_rows": int,
        }
    """
    file_path = Path(file_path)
    filename = file_path.name

    # 파일명에서 학년도 추출 (default)
    year = extract_year_from_filename(filename)

    wb = load_workbook(file_path, read_only=True, data_only=True)

    if SHEET_NAME in wb.sheetnames:
        # 형식 1: 표준 (legacy)
        sheet = wb[SHEET_NAME]
        rows_iter = sheet.iter_rows(values_only=True)
        headers = list(next(rows_iter, ()))
        _validate_headers(headers, wb.sheetnames)
        parsed = _parse_legacy_rows(headers, rows_iter, year, filename)
        fmt = "legacy"
    else:
        # 형식 2/3: 입결_* 3시트
        unified_sheets = [sn for sn in wb.sheetnames if sn.startswith(UNIFIED_SHEET_PREFIX)]
        if not unified_sheets:
            wb.close()
            raise ValueError(
                f"지원하지 않는 Excel 형식입니다 (시트: {wb.sheetnames}). "
                "시트명 '전년도입결' (표준 37컬럼) 또는 "
                "'입결_종합/입결_교과/입결_수능' 3시트 형식만 업로드할 수 있습니다."
            )
        parsed = []
        headers = []
        for sn in unified_sheets:
            sheet_rows = _parse_unified_sheet(wb[sn], sn, year, filename)
            logger.info(f"adiga parse[{filename}] 시트 '{sn}': {len(sheet_rows)}행")
            parsed.extend(sheet_rows)
            if not headers:
                first_sheet = wb[unified_sheets[0]]
                headers = list(next(first_sheet.iter_rows(values_only=True), ()))
        fmt = "unified"

    wb.close()

    return {
        "year": year,
        "filename": filename,
        "format": fmt,
        "headers": headers,
        "rows": parsed,
        "total_rows": len(parsed),
    }


async def import_to_db(
    db: AsyncSession,
    parsed: dict,
    *,
    override_year: int | None = None,
    chunk_size: int = 500,
    mode: str = "full",
) -> dict:
    """
    parse_excel 결과를 DB에 import.

    - override_year 가 주어지면 모든 row 의 year 를 그 값으로 강제 설정 (파일명 학년도 불일치 대응)
    - mode="full": 해당 학년도 데이터 전체 삭제 후 새로 INSERT
    - mode="partial": 파일에 포함된 대학(university_code)의 해당 학년도 데이터만 삭제 후 INSERT
      → 파일에 없는 대학의 기존 데이터는 유지 (일부 대학만 수정해 올리는 워크플로용)
    - chunk_size 단위로 batch insert (대용량 24,000행 대응)

    Returns: {"year": int, "deleted": int, "inserted": int, "total": int, "mode": str}
    """
    rows = parsed["rows"]
    if not rows:
        raise ValueError("Excel 에 유효한 행이 없습니다")
    if mode not in ("full", "partial"):
        raise ValueError(f"지원하지 않는 import mode: {mode} (full/partial)")

    # 학년도 확정
    year = override_year or parsed.get("year")
    if not year:
        raise ValueError("학년도를 결정할 수 없습니다 (파일명 또는 override_year 필요)")

    # 기존 row 삭제
    if mode == "partial":
        codes = sorted({r["university_code"] for r in rows if r["university_code"]})
        del_result = await db.execute(
            delete(AdigaAdmissionResult).where(
                AdigaAdmissionResult.year == year,
                AdigaAdmissionResult.university_code.in_(codes),
            )
        )
    else:
        del_result = await db.execute(
            delete(AdigaAdmissionResult).where(AdigaAdmissionResult.year == year)
        )
    deleted = del_result.rowcount or 0

    # 새 row INSERT (chunk 단위)
    inserted = 0
    chunk: list[AdigaAdmissionResult] = []
    for item in rows:
        item["year"] = year  # override
        # university_code 가 비어 있으면 무효
        if not item["university_code"]:
            continue
        chunk.append(AdigaAdmissionResult(**item))
        if len(chunk) >= chunk_size:
            db.add_all(chunk)
            await db.flush()
            inserted += len(chunk)
            chunk = []
    if chunk:
        db.add_all(chunk)
        await db.flush()
        inserted += len(chunk)

    await db.commit()

    logger.info(
        f"adiga import(year={year}, mode={mode}): deleted={deleted}, inserted={inserted}"
    )
    return {
        "year": year,
        "display_year": year + 1,  # 사용자 페이지의 어느 학년도에 표시되는지 (입결 연도 + 1)
        "mode": mode,
        "deleted": deleted,
        "inserted": inserted,
        "total": inserted,
        "filename": parsed.get("filename"),
    }


async def get_year_summary(db: AsyncSession) -> list[dict]:
    """현재 DB 에 들어있는 학년도별 요약 (학년도별 행 수 + 가장 최근 import 파일)."""
    from sqlalchemy import func

    result = await db.execute(
        select(
            AdigaAdmissionResult.year,
            func.count(AdigaAdmissionResult.id).label("count"),
            func.max(AdigaAdmissionResult.created_at).label("last_imported"),
            func.max(AdigaAdmissionResult.source_file).label("source_file"),
        )
        .group_by(AdigaAdmissionResult.year)
        .order_by(AdigaAdmissionResult.year.desc())
    )
    return [
        {
            "year": row[0],
            "count": row[1],
            "last_imported": row[2].isoformat() if row[2] else None,
            "source_file": row[3],
        }
        for row in result.all()
    ]
